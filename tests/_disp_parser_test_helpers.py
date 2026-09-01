"""Helpers compartidos por los tests de los 6 disp parsers (Fase 5).

Convención: cada parser (``DispEDParser``, ``DispEAParser``,
``DispSAParser``, ``DispVParser``, ``DispMParser``, ``DispM_VFParser``)
tiene su propio ``test_disp_<hw>_parser.py`` que reutiliza estos
helpers. La duplicación entre los 6 archivos de tests es mínima
gracias a este módulo compartido.
"""
from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


# Cabeceras completas del Excel legacy del corporativo. Se usan como
# default en ``save_xlsx_with_disp_table``.
_FULL_DISP_HEADERS: dict[str, list[str]] = {
    "ED": [
        "UID", "Numero", "PLC.Tag", "PLC.Comentario", "Descripcion",
        "Tag", "FAT", "E.Byte", "E.Bit", "Gr.Alarma", "Cuadro",
        "Observaciones", "PLC.Tipo", "PLC.Index", "Hmi.Index",
        "Hmi.Texto", "Cfg.Habilitar", "Cfg.ByteEntrada",
        "Cfg.BitEntrada", "Cfg.GrupoAlarma", "ComentarioDB",
    ],
    "EA": [
        "UID", "Numero", "PLC.Tag", "PLC.Comentario", "Descripcion",
        "Tag", "FAT", "E.Byte", "UNIDADES", "RII", "RSI", "Gr.Alarma",
        "Cuadro", "Observaciones", "PLC.Tipo", "PLC.Index", "Hmi.Index",
        "Hmi.Texto", "Cfg.Habilitar", "Cfg.ByteEntrada",
        "Cfg.EscaladoMin", "Cfg.EscaladoMax", "Cfg.GrupoAlarma",
        "ComentarioDB",
    ],
    "SA": [
        "UID", "Numero", "PLC.Tag", "PLC.Comentario", "Descripcion",
        "Tag", "FAT", "E.Byte", "UNIDADES", "RII", "RSI", "Gr.Alarma",
        "Cuadro", "Observaciones", "PLC.Tipo", "PLC.Index", "Hmi.Index",
        "Hmi.Texto", "Cfg.Habilitar", "Cfg.ByteEntrada",
        "Cfg.EscaladoMin", "Cfg.EscaladoMax", "Cfg.GrupoAlarma",
        "ComentarioDB",
    ],
    "V": [
        "UID", "Numero", "PLC.Tag", "PLC.Comentario", "Descripcion",
        "Tag", "FAT", "S.Byte", "S.Bit", "RR.Byte", "RR.Bit",
        "RT.Byte", "RT.Bit", "Gr.Alarma", "Cuadro", "Observaciones",
        "PLC.Tipo", "PLC.Index", "Hmi.Index", "Hmi.Texto",
        "Cfg.Habilitar", "Cfg.ByteRetornoReposo", "Cfg.BitRetornoReposo",
        "Cfg.ByteRetornoTrabajo", "Cfg.BitRetornoTrabajo",
        "Cfg.ByteActivacion", "Cfg.BitActivacion", "Cfg.HabRetReposo",
        "Cfg.HabRetTrabajo", "Cfg.GrupoAlarma", "ComentarioDB",
    ],
    "M": [
        "UID", "Numero", "PLC.Tag", "PLC.Comentario", "Descripcion",
        "Tag", "FAT", "S.Byte", "S.Bit", "RT.Byte", "RT.Bit",
        "RM.Byte", "RM.Bit", "Gr.Alarma", "Cuadro", "Observaciones",
        "PLC.Tipo", "PLC.Index", "Hmi.Index", "Hmi.Texto",
        "Cfg.Habilitar", "Cfg.ByteRetornoTermico", "Cfg.BitRetornoTermico",
        "Cfg.ByteConfMarcha", "Cfg.BitConfMarcha", "Cfg.ByteActivacion",
        "Cfg.BitActivacion", "Cfg.HabRetTermico", "Cfg.HabRetConfMarcha",
        "Cfg.GrupoAlarma", "ComentarioDB",
    ],
    "M_VF": [
        "UID", "Numero", "PLC.Tag", "PLC.Comentario", "Descripcion",
        "Tag", "FAT", "S.Byte", "S.Bit", "RT.Byte", "RT.Bit",
        "RM.Byte", "RM.Bit", "SA.Byte", "Gr.Alarma", "Cuadro",
        "Observaciones", "PLC.Tipo", "PLC.Index", "Hmi.Index",
        "Hmi.Texto", "Cfg.Habilitar", "Cfg.ByteRetornoTermico",
        "Cfg.BitRetornoTermico", "Cfg.ByteConfMarcha", "Cfg.BitConfMarcha",
        "Cfg.ByteActivacion", "Cfg.BitActivacion", "Cfg.ByteAnalogica",
        "Cfg.HabRetTermico", "Cfg.HabRetConfMarcha", "Cfg.GrupoAlarma",
        "ComentarioDB",
    ],
}


def get_headers_for_hw(hw: str) -> list[str]:
    """Devuelve las cabeceras legacy del Excel para ``hw`` (``"ED"``/...)."""
    return list(_FULL_DISP_HEADERS[hw])


def _column_letter(col_idx_1based: int) -> str:
    """Convierte índice 1-based a letra de columna A-Z, AA-AZ, BA-... .

    ``_column_letter(1) == "A"``, ``_column_letter(27) == "AA"``,
    ``_column_letter(52) == "AZ"``, ``_column_letter(53) == "BA"``.
    """
    result = ""
    n = col_idx_1based
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(ord("A") + rem) + result
    return result


def save_xlsx_with_disp_table(
    tmp_path,
    hw: str,
    rows: list[list] | None,
    *,
    sheet_name: str | None = None,
    table_name: str | None = None,
    headers: list[str] | None = None,
) -> str:
    """Crea un .xlsx sintético con la ``Tabla_Disp_<hw>``."""
    if sheet_name is None:
        sheet_name = f"DISP_{hw}"
    if table_name is None:
        table_name = f"Tabla_Disp_{hw}"
    if headers is None:
        headers = get_headers_for_hw(hw)
    if rows is None:
        rows = []

    xlsx_path = tmp_path / f"{table_name}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    last_col_letter = _column_letter(len(headers))
    last_row = 1 + len(rows)
    ref = f"A1:{last_col_letter}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(str(xlsx_path))
    return str(xlsx_path)


def load_workbook_safe(path: str) -> Workbook:
    """Carga un .xlsx desde disco preservando ``worksheet.tables``."""
    from openpyxl import load_workbook
    return load_workbook(path)


def build_full_row(hw: str, **overrides: Any) -> list:
    """Construye una fila ``list`` con TODOS los campos del hw, opcionalmente
    sobreescritos por ``overrides``.

    Útil para tests que quieren verificar el DTO completo con un solo
    builder call.

    Example::

        row = build_full_row("ED", UID="ED_001", Numero=1, plc_tag="X")
    """
    headers = get_headers_for_hw(hw)
    # Defaults coherentes con la primera fila típica del Excel legacy.
    defaults: dict[str, Any] = {
        "UID":            f"{hw}_001",
        "Numero":         1,
        "PLC.Tag":        f"V_{hw}_001",
        "PLC.Comentario": f"Comentario {hw}",
        "Descripcion":    f"Desc {hw}",
        "Tag":            f"T_{hw}",
        "FAT":            f"F_{hw}",
        "E.Byte":         0, "E.Bit": 0,
        "Gr.Alarma":      0,
        "Cuadro":         "C1",
        "Observaciones":  "Obs",
        "PLC.Tipo":       "Bool",
        "PLC.Index":      0,
        "Hmi.Index":      0,
        "Hmi.Texto":      f"Texto HMI {hw}",
        "Cfg.Habilitar":  "cfg_hab := 1;",
        "ComentarioDB":   f"DB {hw} comentario",
    }
    # Defaults específicos por hw (campos únicos).
    if hw in ("EA", "SA"):
        defaults.update({
            "UNIDADES": "BAR",
            "RII": 0.0,
            "RSI": 100.0,
            "Cfg.EscaladoMin": "cfg_min := 0.0;",
            "Cfg.EscaladoMax": "cfg_max := 100.0;",
        })
    if hw == "V":
        defaults.update({
            "S.Byte": 0, "S.Bit": 0,
            "RR.Byte": 0, "RR.Bit": 0,
            "RT.Byte": 0, "RT.Bit": 0,
        })
    if hw in ("M", "M_VF"):
        defaults.update({
            "S.Byte": 0, "S.Bit": 0,
            "RT.Byte": 0, "RT.Bit": 0,
            "RM.Byte": 0, "RM.Bit": 0,
        })
    if hw == "M_VF":
        defaults.update({
            "SA.Byte": 0,
            "Cfg.ByteAnalogica": "cfg_ana := 0;",
        })
    # Override.
    for k, v in overrides.items():
        defaults[k] = v
    return [defaults.get(h, "") for h in headers]


__all__ = [
    "get_headers_for_hw",
    "save_xlsx_with_disp_table",
    "load_workbook_safe",
    "build_full_row",
]
