"""Tests del tool MCP ``tia_upload_excel`` wired al cache (Fase 5).

Cubre:
  * La tool popula ``state.excel_cache`` y ``state.excel_path``.
  * Path inexistente → ``FileNotFoundError``.
  * Response tiene la misma shape que el endpoint web.

Patrón: hereda la fixture ``mcp_deps`` de
``test_mcp_alimentacion_tools.py`` (vía import local) y solo añade
los tests específicos del cache.
"""
from __future__ import annotations

import asyncio
import json
from io import BytesIO
from pathlib import Path
from typing import Any
from unittest.mock import MagicMock

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from areas.alimentacion.interfaces.mcp import tools as mcp_tools
from core.application.log_buffer import get_log_buffer
from core.application.state import get_app_state
from core.infrastructure.gateway import TIAProcessGateway
from core.interfaces import mcp_server
from core.infrastructure.config_manager import ConfigManager


# ── Config JSON fixture (mismo que test_mcp_alimentacion_tools) ──────


_FULL_CONFIG: dict[str, Any] = {
    "departments": {
        "alimentacion": {
            "global_config_table_name": "000_Config_Dispositivos",
            "tia_folders": {
                "proceso":      "003_Procesos",
                "dispositivos": "2000_Dispositivos",
                "nmax":         "000_Sistema",
            },
            "n_max_catalog": [
                {"name": "N_MAX_DISP_ED",   "value": 10},
                {"name": "N_MAX_DISP_EA",   "value": 10},
                {"name": "N_MAX_DISP_SA",   "value": 10},
                {"name": "N_MAX_DISP_V",    "value": 10},
                {"name": "N_MAX_DISP_M",    "value": 10},
                {"name": "N_MAX_DISP_M_VF", "value": 10},
            ],
            "Dispositivos": {
                "ed": {
                    "db_name": "DB2000_ED", "db_array_name": "ED",
                    "tag_table": "2000_Disp_ED",
                    "config_table": "000_Config_Dispositivos",
                },
                "ea": {
                    "db_name": "DB2001_EA", "db_array_name": "EA",
                    "tag_table": "2000_Disp_EA",
                    "config_table": "000_Config_Dispositivos",
                },
                "sa": {
                    "db_name": "DB2006_SA", "db_array_name": "SA",
                    "tag_table": "2000_Disp_SA",
                    "config_table": "000_Config_Dispositivos",
                },
                "v": {
                    "db_name": "DB2010_V", "db_array_name": "V",
                    "tag_table": "2000_Disp_V",
                    "config_table": "000_Config_Dispositivos",
                },
                "m": {
                    "db_name": "DB2015_M", "db_array_name": "M",
                    "tag_table": "2000_Disp_M",
                    "config_table": "000_Config_Dispositivos",
                },
                "m_vf": {
                    "db_name": "DB2016_M_VF", "db_array_name": "M_VF",
                    "tag_table": "2000_Disp_M_VF",
                    "config_table": "000_Config_Dispositivos",
                },
            },
        }
    }
}


def _write_config(tmp_path: Path) -> Path:
    p = tmp_path / "config.json"
    p.write_text(json.dumps(_FULL_CONFIG), encoding="utf-8")
    return p


def _add_table(
    wb: Workbook, sheet_name: str, table_name: str,
    headers: list[str], rows: list[list],
) -> None:
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    last_col_letter = chr(ord("A") + len(headers) - 1)
    last_row = 1 + len(rows)
    ref = f"A1:{last_col_letter}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)


def _build_minimal_xlsx(target: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    tables = [
        ("DISP_ED",   "Tabla_Disp_ED",   "ED_001",   1, "V_ED_001"),
        ("DISP_EA",   "Tabla_Disp_EA",   "EA_001",   1, "V_EA_001"),
        ("DISP_SA",   "Tabla_Disp_SA",   "SA_001",   1, "V_SA_001"),
        ("DISP_V",    "Tabla_Disp_V",    "V_001",    1, "V_V_001"),
        ("DISP_M",    "Tabla_Disp_M",    "M_001",    1, "V_M_001"),
        ("DISP_M_VF", "Tabla_Disp_M_VF", "MVF_001",  1, "V_MVF_001"),
    ]
    headers = ["UID", "Numero", "PLC.Tag", "Descripcion"]
    for sheet_name, table_name, uid, numero, tag in tables:
        _add_table(wb, sheet_name, table_name, headers,
                   [[uid, numero, tag, f"Desc {uid}"]])
    wb.save(target)
    return target


# ── Capturador de tools (mock mínimo de FastMCP) ──────────────────────


class _McpToolCapture:
    def __init__(self) -> None:
        self._tools_by_name: dict[str, Any] = {}

    def tool(self):
        def decorator(fn):
            self._tools_by_name[fn.__name__] = fn
            return fn
        return decorator


def _get_tool(mcp_mock, name: str):
    tools = getattr(mcp_mock, "_tools_by_name", {})
    if name not in tools:
        raise AssertionError(
            f"Tool {name!r} no encontrada. Disponibles: {sorted(tools.keys())}"
        )
    return tools[name]


# ── Fixture ────────────────────────────────────────────────────────────


@pytest.fixture
def mcp_deps(tmp_path: Path):
    config_path = _write_config(tmp_path)
    cm = ConfigManager(config_path=config_path)
    gw = MagicMock(spec=TIAProcessGateway)
    state = get_app_state()
    # Reset del Singleton (puede venir contaminado de otros tests).
    state.reset()
    state.excel_cache = None
    state.excel_path = None
    state.dimensiones = None

    mcp_server._deps = {
        "gateway": gw,
        "config_manager": cm,
        "app_state": state,
        "logger": get_log_buffer(),
    }

    captured_mcp = _McpToolCapture()
    mcp_tools.register(captured_mcp)
    yield {
        "gateway": gw,
        "config_manager": cm,
        "app_state": state,
        "logger": get_log_buffer(),
        "mcp": captured_mcp,
    }
    # Reset al final para que tests posteriores vean estado limpio.
    state.reset()
    state.excel_cache = None
    state.excel_path = None
    state.dimensiones = None


# ── Tests ──────────────────────────────────────────────────────────────


def test_tia_upload_excel_popula_cache(
    mcp_deps: dict[str, Any], tmp_path: Path
) -> None:
    """``tia_upload_excel`` popula ``state.excel_cache`` y ``state.excel_path``."""
    tool = _get_tool(mcp_deps["mcp"], "tia_upload_excel")
    xlsx_path = _build_minimal_xlsx(tmp_path / "fixture.xlsx")
    state = mcp_deps["app_state"]

    # Sanity: vacío al inicio.
    assert state.excel_cache is None
    assert state.excel_path is None

    result = asyncio.run(tool(str(xlsx_path)))

    # Cache populado.
    assert state.excel_cache is not None
    assert state.excel_path is not None
    assert state.excel_path.endswith("fixture.xlsx")
    # Volcado data-driven: los 6 tipos populados.
    for hw in ("ed", "ea", "sa", "v", "m", "m_vf"):
        assert len(state.get_devices(hw)) == 1, hw
    # ``state.dimensiones`` también populado.
    assert state.dimensiones is not None
    # Response shape legacy.
    assert result["ok"] is True
    assert result["total_dispositivos"] == 6
    assert "summary" in result
    assert "dimensiones" in result


def test_tia_upload_excel_path_inexistente_lanza_error(
    mcp_deps: dict[str, Any], tmp_path: Path
) -> None:
    """Path que no existe → ``FileNotFoundError``."""
    tool = _get_tool(mcp_deps["mcp"], "tia_upload_excel")
    missing = tmp_path / "no_existe.xlsx"
    assert not missing.is_file()

    with pytest.raises(FileNotFoundError):
        asyncio.run(tool(str(missing)))


def test_tia_upload_excel_response_shape(
    mcp_deps: dict[str, Any], tmp_path: Path
) -> None:
    """Response de la tool = mismo shape que el endpoint web."""
    tool = _get_tool(mcp_deps["mcp"], "tia_upload_excel")
    xlsx_path = _build_minimal_xlsx(tmp_path / "fixture.xlsx")

    result = asyncio.run(tool(str(xlsx_path)))

    # Mismas claves que el endpoint web.
    assert set(result.keys()) == {
        "ok", "summary", "total_dispositivos", "dimensiones",
    }
    # ``dimensiones`` tiene los 6 canónicos (no ``extras``).
    assert set(result["dimensiones"].keys()) == {
        "num_disp_ed", "num_disp_ea", "num_disp_sa",
        "num_disp_v", "num_disp_m", "num_disp_m_vf",
    }
