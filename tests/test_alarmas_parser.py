"""Tests del parser ``AlarmasParser`` (Fase 4 del plan).

Cubre la extracción de ``Tabla_Alarmas`` (hoja ``ALARMAS``) y, en
particular, la invariante R-F4.1: ``AlarmaPLC`` NO tiene atributo
``visibilidad`` y el parser ignora silenciosamente la columna
``Visibilidad`` del Excel si esta existe.

Es la **implementación de referencia** que los 6 mini parsers de
dispositivos de Fase 5 imitarán; los tests aquí son la plantilla
que los tests de Fase 5 extenderán.

Convenciones:
    * Excel sintético construido en ``tmp_path`` con ``Table`` real
      (R5 del plan).
    * Se carga con ``load_workbook`` para garantizar que la
      ``ListObject`` se registre en ``worksheet.tables``.
"""
from __future__ import annotations

from dataclasses import fields

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from areas.alimentacion.domain.models.excel_cache import AlarmaPLC
from areas.alimentacion.infrastructure.parsers.alarmas import AlarmasParser


# ── Helpers de construcción de Excels sintéticos ────────────────────────


def _save_xlsx_with_alarmas_table(
    tmp_path,
    rows: list[list] | None,
    *,
    sheet_name: str = "ALARMAS",
    table_name: str = "Tabla_Alarmas",
    headers: list[str] | None = None,
) -> str:
    """Crea un .xlsx sintético con la tabla de alarmas.

    Args:
        tmp_path: fixture pytest de path temporal.
        rows: lista de filas de datos (``None`` o ``[]`` → solo
            cabeceras, útil para verificar que la tabla existe
            pero está vacía).
        sheet_name: nombre de la hoja (default ``ALARMAS``).
        table_name: nombre de la ``ListObject`` (default
            ``Tabla_Alarmas``).
        headers: cabeceras (default: las del Excel legacy del
            corporativo, 6 columnas). Si se pasa otra lista, el
            test puede simular schema drift (p. ej. añadir
            ``Visibilidad`` para verificar R-F4.1).

    Returns:
        Path absoluto del .xlsx en str.
    """
    if headers is None:
        # Cabeceras del legacy: 6 columnas, SIN Visibilidad.
        headers = [
            "UID",
            "Numero",
            "Proceso",
            "Num.DB",
            "Descripcion",
            "ComentarioDB",
        ]
    if rows is None:
        rows = []

    xlsx_path = tmp_path / f"{table_name}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)

    # Registrar la Table (R5 del plan: sin esto, ``worksheet.tables``
    # no contiene la ``ListObject`` y el parser no la encontraría).
    last_col_letter = chr(ord("A") + len(headers) - 1)
    last_row = 1 + len(rows)
    ref = f"A1:{last_col_letter}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(str(xlsx_path))
    return str(xlsx_path)


def _load(path: str) -> Workbook:
    """Carga un .xlsx desde disco preservando ``worksheet.tables``."""
    from openpyxl import load_workbook

    return load_workbook(path)


# ── Tests del parser ``AlarmasParser`` ──────────────────────────────────


def test_extrae_alarmas_basico(tmp_path) -> None:
    """Excel con 1 fila válida → DTO con los 6 campos correctos.

    El test cubre los 6 campos del DTO ``AlarmaPLC``:
    ``uid``, ``numero``, ``proceso``, ``num_db``, ``descripcion``,
    ``comentario_db``. ``num_db`` se castea a ``int`` vía
    ``_safe_int``. El resto son ``str`` con defaults tolerantes.
    """
    xlsx_path = _save_xlsx_with_alarmas_table(
        tmp_path,
        rows=[
            [
                "AL_1_001",  # UID
                "001",  # Numero
                "Proceso Uno",  # Proceso
                5001,  # Num.DB (DB alarmas = 5000+uid)
                "Presion alta en linea A",  # Descripcion
                "DB alarmas proceso 1",  # ComentarioDB
            ],
        ],
    )
    wb = _load(xlsx_path)

    result = AlarmasParser().extraer(wb)

    assert len(result) == 1
    a = result[0]
    assert isinstance(a, AlarmaPLC)
    assert a.uid == "AL_1_001"
    assert a.numero == "001"
    assert a.proceso == "Proceso Uno"
    assert a.num_db == 5001
    assert a.descripcion == "Presion alta en linea A"
    assert a.comentario_db == "DB alarmas proceso 1"


def test_hoja_inexistente_devuelve_lista_vacia(tmp_path) -> None:
    """Sheet ``ALARMAS`` no existe → ``[]`` (no lanza).

    Política coherente con R1 del plan y con el resto de
    parsers de software: si la hoja no está, no es un error
    (el Excel puede no traerla en alguna configuración).
    """
    xlsx_path = _save_xlsx_with_alarmas_table(
        tmp_path,
        rows=[["AL_1_001", "001", "P", 5001, "", ""]],
        sheet_name="OTRA_HOJA",
    )
    wb = _load(xlsx_path)

    result = AlarmasParser().extraer(wb)

    assert result == []


def test_tabla_inexistente_devuelve_lista_vacia(tmp_path) -> None:
    """Sheet existe pero ``Tabla_Alarmas`` no → ``[]`` (no lanza)."""
    # Construimos un .xlsx con OTRA tabla en ALARMAS.
    xlsx_path = tmp_path / "otro.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "ALARMAS"
    ws.append(["X", "Y"])
    ws.append([1, 2])
    table = Table(displayName="Tabla_Otra", ref="A1:B2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(str(xlsx_path))

    wb2 = _load(xlsx_path)
    result = AlarmasParser().extraer(wb2)
    assert result == []


def test_fila_sin_uid_se_descarta(tmp_path) -> None:
    """Fila con UID vacío (``None``) NO aparece en el resultado.

    Política consistente con ``ProcesosParser``/``PRealParser``/
    ``PIntParser`` (dropna por UID). Evita alarmas fantasma sin
    UID en el cache.
    """
    xlsx_path = _save_xlsx_with_alarmas_table(
        tmp_path,
        rows=[
            [None, "001", "P", 5001, "desc", "coment"],  # sin UID
            [
                "AL_1_001",  # válida
                "001",
                "Proceso Uno",
                5001,
                "desc",
                "coment",
            ],
        ],
    )
    wb = _load(xlsx_path)

    result = AlarmasParser().extraer(wb)

    # Solo la fila con UID válido se queda.
    assert len(result) == 1
    assert result[0].uid == "AL_1_001"


def test_sin_visibilidad_en_dto() -> None:
    """``AlarmaPLC`` NO expone atributo ``visibilidad`` (R-F4.1).

    Esta es una **invariante contractual**: la tabla de Alarmas
    del Excel corporativo no incluye ``Visibilidad`` (consistente
    con el legacy ``_legacy_reference/ZC_ALM_TOOLS/infrastructure/
    parsers/software/alarmas.py:21-31``), por lo que el DTO no
    expone ese campo. El test bloquea cualquier regresión que
    añada ``visibilidad`` al DTO por confusión con
    ``ParamRealPLC``/``ParamIntPLC``.

    Se verifica doble:
        * ``hasattr(AlarmaPLC(...), 'visibilidad') == False``
          (invariante de instancia).
        * ``'visibilidad' not in [f.name for f in fields(AlarmaPLC)]``
          (invariante de clase / dataclass).
    """
    a = AlarmaPLC(
        uid="AL_1_001",
        numero="001",
        proceso="P",
        num_db=5001,
        descripcion="d",
        comentario_db="c",
    )
    # Invariante de instancia.
    assert not hasattr(a, "visibilidad")
    # Invariante de clase.
    field_names = {f.name for f in fields(AlarmaPLC)}
    assert "visibilidad" not in field_names
    # Sanity: los 6 campos esperados sí están.
    assert field_names == {
        "uid",
        "numero",
        "proceso",
        "num_db",
        "descripcion",
        "comentario_db",
    }


def test_columna_visibilidad_en_excel_se_ignora(tmp_path) -> None:
    """Columna extra ``Visibilidad`` en el Excel se ignora (R-F4.1).

    Defensa contra schema drift: si el corporativo añade una
    columna ``Visibilidad`` a ``Tabla_Alarmas`` en el futuro, el
    parser la **ignora silenciosamente**. El DTO se construye
    correctamente con sus 6 campos conocidos y el resultado NO
    tiene atributo ``visibilidad``.

    Por qué funciona: ``AlarmasParser.extraer`` llama al
    constructor de ``AlarmaPLC`` con kwargs explícitos
    (``uid=...``, ``numero=...``, etc.) y NO usa ``**row``. La
    columna extra queda en el ``dict`` que devuelve
    ``extract_list_object_rows``, pero el constructor de la
    dataclass ``frozen=True`` nunca la lee. No se emite WARNING
    ni se descarta la fila: la alarma se conserva con sus 6
    campos originales.

    Esto es crítico para la **back-compat con el legacy** y para
    que ``AlarmasParser`` siga siendo drop-in si el schema del
    Excel evoluciona.
    """
    # Cabecera con Visibilidad añadida al final (columna extra).
    headers_con_visibilidad = [
        "UID",
        "Numero",
        "Proceso",
        "Num.DB",
        "Descripcion",
        "ComentarioDB",
        "Visibilidad",  # columna extra (schema drift)
    ]
    # Una fila válida: el campo "Si" en Visibilidad debe ignorarse.
    fila = [
        "AL_1_001",
        "001",
        "Proceso Uno",
        5001,
        "Presion alta",
        "DB alarmas",
        "Si",  # valor de la columna ignorada
    ]
    xlsx_path = _save_xlsx_with_alarmas_table(
        tmp_path,
        rows=[fila],
        headers=headers_con_visibilidad,
    )
    wb = _load(xlsx_path)

    result = AlarmasParser().extraer(wb)

    # El parser devuelve 1 alarma: la fila es válida para los 6
    # campos que conoce, y la columna "Visibilidad" simplemente
    # se ignora.
    assert len(result) == 1
    a = result[0]
    # El DTO tiene los 6 campos esperados con los valores correctos.
    assert a.uid == "AL_1_001"
    assert a.numero == "001"
    assert a.proceso == "Proceso Uno"
    assert a.num_db == 5001
    assert a.descripcion == "Presion alta"
    assert a.comentario_db == "DB alarmas"
    # El valor de "Visibilidad" NO aparece en el DTO.
    assert not hasattr(a, "visibilidad")
    # El dict crudo sí traía la clave (extract_list_object_rows la
    # recoge), pero el constructor de AlarmaPLC solo lee los 6
    # kwargs explícitos. Verificamos que la fila cruda tiene la
    # clave para confirmar que el parser la "vio" pero no la usó.
    from areas.alimentacion.infrastructure.parsers._xlsx_helpers import (
        extract_list_object_rows,
    )
    raw_rows = extract_list_object_rows(wb, "ALARMAS", "Tabla_Alarmas")
    assert len(raw_rows) == 1
    assert raw_rows[0].get("Visibilidad") == "Si"
