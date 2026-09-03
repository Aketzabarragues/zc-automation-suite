"""Tests de las tools MCP del área de alimentación (PR 6).

Cubre las 4 tools que dan paridad con los endpoints web del área:
  - tia_sync_disp_preview      → DispSyncInstancesUseCase.generar_prevision
  - tia_sync_disp_commit       → DispSyncInstancesUseCase.ejecutar_transaccion
  - tia_apply_disp_comentarios → DispComentariosSyncUseCase.apply_comentarios_disp
  - tia_upload_excel           → AlimentacionExcelParser + volcado data-driven

Patrón:
  - Mockear el gateway con ``MagicMock(spec=TIAProcessGateway)`` (no
    instancias reales: tests offline).
  - ``ConfigManager`` con fixture JSON en ``tmp_path`` (mismo patrón
    que ``test_config_manager.py``).
  - ``AppState`` real (con state_extensions del área instaladas) —
    verificar que ``tia_upload_excel`` popula los 6 atributos legacy.
"""
from __future__ import annotations

import asyncio
import json
from pathlib import Path
from typing import Any
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from areas.alimentacion.interfaces.mcp import tools as mcp_tools
from core.application.log_buffer import get_log_buffer
from core.application.progress_buffer import ProgressTracker
from core.application.state import get_app_state
from core.infrastructure.gateway import TIAProcessGateway
from core.interfaces import mcp_server


# ── Helpers locales ────────────────────────────────────────────────────


def _arun(coro):
    """Ejecuta una coroutine y devuelve su resultado (tests síncronos)."""
    return asyncio.run(coro)


def _get_registered_tool(mcp_mock: Any, name: str):
    """Recupera una tool registrada vía ``@mcp.tool()``.

    ``@mcp.tool()`` se aplica como decorador SIN argumentos al wrapper
    ``FastMCP.tool()``: Python invoca primero ``mcp.tool()`` (que
    devuelve un decorador interno) y luego llama a ese decorador
    con la función como argumento. Para capturar la función, el mock
    debe recordar las llamadas al decorador interno.

    El fixture (``mcp_capture_factory``) devuelve un objeto que se
    comporta como FastMCP y guarda cada tool en ``_tools_by_name``.
    """
    tools = getattr(mcp_mock, "_tools_by_name", {})
    if name not in tools:
        raise AssertionError(
            f"Tool {name!r} no encontrada entre las registradas. "
            f"Disponibles: {sorted(tools.keys())}"
        )
    return tools[name]


# ── Configuración JSON fixture ─────────────────────────────────────────

# Mínimo viable: 6 tipos (ed/ea/sa/v/m/m_vf) con los campos que
# ``ConfigManager`` necesita para resolver el path data-driven del
# volcado Excel → AppState.
_FULL_CONFIG: dict[str, Any] = {
    "_comment": "Test fixture for MCP tools",
    "departments": {
        "alimentacion": {
            "global_config_table_name": "000_Config_Dispositivos",
            "tia_folders": {
                "proceso":      "003_Procesos",
                "dispositivos": "2000_Dispositivos",
                "nmax":         "000_Sistema",
            },
            "n_max_catalog": [
                {"name": "N_MAX_DISP_ED",   "value": 10},
                {"name": "N_MAX_DISP_EA",   "value": 10},
                {"name": "N_MAX_DISP_SA",   "value": 10},
                {"name": "N_MAX_DISP_V",    "value": 10},
                {"name": "N_MAX_DISP_M",    "value": 10},
                {"name": "N_MAX_DISP_M_VF", "value": 10},
            ],
            "Dispositivos": {
                "ed": {
                    "db_name": "DB2000_ED", "db_array_name": "ED",
                    "tag_table": "2000_Disp_ED", "config_table": "000_Config_Dispositivos",
                },
                "ea": {
                    "db_name": "DB2001_EA", "db_array_name": "EA",
                    "tag_table": "2000_Disp_EA", "config_table": "000_Config_Dispositivos",
                },
                "sa": {
                    "db_name": "DB2006_SA", "db_array_name": "SA",
                    "tag_table": "2000_Disp_SA", "config_table": "000_Config_Dispositivos",
                },
                "v": {
                    "db_name": "DB2010_V", "db_array_name": "V",
                    "tag_table": "2000_Disp_V", "config_table": "000_Config_Dispositivos",
                },
                "m": {
                    "db_name": "DB2015_M", "db_array_name": "M",
                    "tag_table": "2000_Disp_M", "config_table": "000_Config_Dispositivos",
                },
                "m_vf": {
                    "db_name": "DB2016_M_VF", "db_array_name": "M_VF",
                    "tag_table": "2000_Disp_M_VF",
                    "config_table": "000_Config_Dispositivos",
                },
            },
        },
    },
}


def _write_config(tmp_path: Path) -> Path:
    p = tmp_path / "config.json"
    p.write_text(json.dumps(_FULL_CONFIG), encoding="utf-8")
    return p


# ── Fixture helper: xlsx con 6 ListObjects y 1 fila mínima por tabla ──


def _build_minimal_xlsx(target: Path) -> Path:
    """Genera un xlsx con las 6 ListObjects, 1 fila mínima por tabla.

    Cada tabla tiene solo 4 columnas (UID, Numero, PLC.Tag, Descripcion):
    lo mínimo que el parser necesita para construir un dispositivo
    (cada builder exige UID o Numero; el resto se rellena con defaults
    vacíos/0 vía ``_safe_str`` / ``_safe_int``).
    """
    wb = Workbook()
    wb.remove(wb.active)  # Quitar la hoja por defecto.

    tables = [
        ("DISP_ED",   "Tabla_Disp_ED",   "ED_001",   1, "V_ED_001"),
        ("DISP_EA",   "Tabla_Disp_EA",   "EA_001",   1, "V_EA_001"),
        ("DISP_SA",   "Tabla_Disp_SA",   "SA_001",   1, "V_SA_001"),
        ("DISP_V",    "Tabla_Disp_V",    "V_001",    1, "V_V_001"),
        ("DISP_M",    "Tabla_Disp_M",    "M_001",    1, "V_M_001"),
        ("DISP_M_VF", "Tabla_Disp_M_VF", "MVF_001",  1, "V_MVF_001"),
    ]
    headers = ["UID", "Numero", "PLC.Tag", "Descripcion"]

    for sheet_name, table_name, uid, numero, tag in tables:
        ws = wb.create_sheet(sheet_name)
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        ws.cell(row=2, column=1, value=uid)
        ws.cell(row=2, column=2, value=numero)
        ws.cell(row=2, column=3, value=tag)
        ws.cell(row=2, column=4, value=f"Desc {uid}")
        ref = "A1:D2"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(table)

    wb.save(target)
    return target


# ── Capturador de tools (mock mínimo de FastMCP) ──────────────────────


class _McpToolCapture:
    """Mock mínimo de FastMCP que captura las tools registradas por nombre.

    Simula el contrato ``@mcp.tool()`` que usa ``mcp_tools.register``:
    cada ``@mcp.tool()`` aplicado a una función async la guarda en
    ``_tools_by_name`` indexada por ``fn.__name__``.
    """

    def __init__(self) -> None:
        self._tools_by_name: dict[str, Any] = {}

    def tool(self):
        """Devuelve un decorador que captura la función en ``_tools_by_name``."""

        def decorator(fn):
            self._tools_by_name[fn.__name__] = fn
            return fn

        return decorator


# ── Fixtures ───────────────────────────────────────────────────────────


@pytest.fixture
def mcp_deps(tmp_path: Path):
    """Puebla el dict ``_deps`` del shell con mocks listos para los tools.

    El shell (``create_mcp_server``) lo haría en producción, pero los
    tests usan un capturador ligero (``_McpToolCapture``) para evitar
    instanciar un FastMCP real (más rápido, mismo shape de API).
    """
    from core.infrastructure.config_manager import ConfigManager

    config_path = _write_config(tmp_path)
    cm = ConfigManager(config_path=config_path)
    gw = MagicMock(spec=TIAProcessGateway)
    state = get_app_state()  # Singleton (con state_extensions aplicadas).

    mcp_server._deps = {
        "gateway": gw,
        "config_manager": cm,
        "app_state": state,
        "logger": get_log_buffer(),
    }

    captured_mcp = _McpToolCapture()
    mcp_tools.register(captured_mcp)
    return {
        "gateway": gw,
        "config_manager": cm,
        "app_state": state,
        "logger": get_log_buffer(),
        "mcp": captured_mcp,
    }


# ── Tests ──────────────────────────────────────────────────────────────


def test_tia_sync_disp_preview_invokes_generar_prevision(
    mcp_deps: dict[str, Any],
) -> None:
    """``tia_sync_disp_preview`` delega en el use case ``generar_prevision``."""
    tool = _get_registered_tool(mcp_deps["mcp"], "tia_sync_disp_preview")
    expected = {"agregados": [], "eliminados": [], "summary": {}}
    with patch(
        "areas.alimentacion.application.use_cases."
        "disp_sync_instances.DispSyncInstancesUseCase"
    ) as UCCls:
        uc = MagicMock()
        uc.generar_prevision = AsyncMock(return_value=expected)
        UCCls.return_value = uc
        result = _arun(tool("PLC_PRUEBAS"))

    UCCls.assert_called_once_with(
        gateway=mcp_deps["gateway"],
        config_manager=mcp_deps["config_manager"],
        state=mcp_deps["app_state"],
    )
    uc.generar_prevision.assert_awaited_once_with("PLC_PRUEBAS")
    assert result is expected


def test_tia_sync_disp_commit_invokes_ejecutar_transaccion(
    mcp_deps: dict[str, Any],
) -> None:
    """``tia_sync_disp_commit`` delega en el use case ``ejecutar_transaccion``."""
    tool = _get_registered_tool(mcp_deps["mcp"], "tia_sync_disp_commit")
    prevision = {"agregados": [{"uid": "X"}]}
    expected = {"plc_name": "PLC_PRUEBAS", "operations": 3}
    with patch(
        "areas.alimentacion.application.use_cases."
        "disp_sync_instances.DispSyncInstancesUseCase"
    ) as UCCls:
        uc = MagicMock()
        uc.ejecutar_transaccion = AsyncMock(return_value=expected)
        UCCls.return_value = uc
        result = _arun(tool("PLC_PRUEBAS", prevision))

    UCCls.assert_called_once_with(
        gateway=mcp_deps["gateway"],
        config_manager=mcp_deps["config_manager"],
        state=mcp_deps["app_state"],
    )
    uc.ejecutar_transaccion.assert_awaited_once_with("PLC_PRUEBAS", prevision)
    assert result is expected


def test_tia_apply_disp_comentarios_invokes_apply_comentarios_disp(
    mcp_deps: dict[str, Any],
) -> None:
    """``tia_apply_disp_comentarios`` delega en ``apply_comentarios_disp``."""
    tool = _get_registered_tool(
        mcp_deps["mcp"], "tia_apply_disp_comentarios"
    )
    expected = {
        "plc_name": "PLC_PRUEBAS",
        "success": True,
        "applied": True,
        "operations_executed": 6,
        "summary": {"disp_dbs_updated": 6, "total_ops": 6},
        "details": [],
        "warnings": [],
    }
    with patch(
        "areas.alimentacion.application.use_cases."
        "disp_sync_comentarios.DispComentariosSyncUseCase"
    ) as UCCls:
        uc = MagicMock()
        uc.apply_comentarios_disp = AsyncMock(return_value=expected)
        UCCls.return_value = uc
        with patch(
            "core.application.progress_buffer.get_progress_tracker",
            return_value=MagicMock(spec=ProgressTracker),
        ):
            result = _arun(tool("PLC_PRUEBAS"))

    UCCls.assert_called_once()
    uc.apply_comentarios_disp.assert_awaited_once_with("PLC_PRUEBAS")
    assert result is expected


def test_tia_upload_excel_populates_app_state_with_6_types(
    mcp_deps: dict[str, Any], tmp_path: Path
) -> None:
    """``tia_upload_excel`` parsea el xlsx y popula AppState con los 6 tipos.

    Verifica que el volcado data-driven (mismo path que el router
    web) escribe en los 6 atributos legacy del AppState.
    """
    tool = _get_registered_tool(mcp_deps["mcp"], "tia_upload_excel")
    xlsx_path = _build_minimal_xlsx(tmp_path / "fixture.xlsx")
    state = mcp_deps["app_state"]

    # Sanity: al principio el AppState está vacío para los 6 tipos.
    for hw in ("ed", "ea", "sa", "v", "m", "m_vf"):
        assert state.get_devices(hw) == []

    result = _arun(tool(str(xlsx_path)))

    # Devuelve la shape esperada.
    assert result["ok"] is True
    assert result["total_dispositivos"] == 6  # 1 por cada uno de los 6 tipos
    assert set(result["summary"].keys()) == {
        "DispED", "DispEA", "DispSA", "DispV", "DispM", "DispM_VF",
    }
    for tipo, count in result["summary"].items():
        assert count == 1, f"Tipo {tipo!r} esperaba 1 dispositivo, obtuve {count}"

    # AppState populado con los 6 tipos (atributos legacy instalados
    # por state_extensions del área de alimentación).
    for hw in ("ed", "ea", "sa", "v", "m", "m_vf"):
        devices = state.get_devices(hw)
        assert len(devices) == 1, (
            f"Tipo {hw!r} esperaba 1 dispositivo en AppState, "
            f"obtuve {len(devices)}"
        )


def test_tia_upload_excel_raises_filenotfound_for_missing_path(
    mcp_deps: dict[str, Any], tmp_path: Path
) -> None:
    """``tia_upload_excel`` lanza ``FileNotFoundError`` si el path no existe."""
    tool = _get_registered_tool(mcp_deps["mcp"], "tia_upload_excel")
    missing = tmp_path / "no_existe.xlsx"
    assert not missing.is_file()

    with pytest.raises(FileNotFoundError):
        _arun(tool(str(missing)))
