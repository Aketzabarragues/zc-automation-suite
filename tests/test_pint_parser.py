"""Tests del parser ``PIntParser`` (Fase 3 del plan).

Cubre la extracción de ``Tabla_PInt`` (hoja ``P_INT``) y la
preservación de ``num_lista`` como ``int | str`` (helper
``_safe_num_lista``), ya tratado en Fase 2 pero re-verificado en
el contexto del parser de enteros. Además incluye el test de
tipos distintos (R4 del plan): ``ParamIntPLC`` y ``ParamRealPLC``
son nominalmente distintos aunque compartan shape.

Convenciones:
    * Excel sintético construido en ``tmp_path`` con ``Table`` real
      (R5 del plan).
    * Se carga con ``load_workbook`` para garantizar que la
      ``ListObject`` se registre en ``worksheet.tables``.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from areas.alimentacion.domain.models.excel_cache import (
    ParamIntPLC,
    ParamRealPLC,
)
from areas.alimentacion.infrastructure.parsers.proc_pint import PIntParser


# ── Helpers de construcción de Excels sintéticos ────────────────────────


def _save_xlsx_with_pint_table(
    tmp_path,
    rows: list[list] | None,
    *,
    sheet_name: str = "P_INT",
    table_name: str = "Tabla_PInt",
    headers: list[str] | None = None,
) -> str:
    """Crea un .xlsx sintético con la tabla de parámetros enteros.

    Args:
        tmp_path: fixture pytest de path temporal.
        rows: lista de filas de datos (``None`` o ``[]`` → solo
            cabeceras, útil para verificar que la tabla existe
            pero está vacía).
        sheet_name: nombre de la hoja (default ``P_INT``).
        table_name: nombre de la ``ListObject`` (default
            ``Tabla_PInt``).
        headers: cabeceras (default: las del Excel legacy del
            corporativo, 12 columnas).

    Returns:
        Path absoluto del .xlsx en str.
    """
    if headers is None:
        headers = [
            "UID",
            "Numero",
            "Proceso",
            "Codigo",
            "Num.DB",
            "Producto",
            "Tipo",
            "Descripcion",
            "ComentarioDB",
            "Visibilidad",
            "Num.Lista",
            "Txt.Lista",
        ]
    if rows is None:
        rows = []

    xlsx_path = tmp_path / f"{table_name}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)

    # Registrar la Table (R5 del plan: sin esto, ``worksheet.tables``
    # no contiene la ``ListObject`` y el parser no la encontraría).
    last_col_letter = chr(ord("A") + len(headers) - 1)
    last_row = 1 + len(rows)
    ref = f"A1:{last_col_letter}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(str(xlsx_path))
    return str(xlsx_path)


def _load(path: str) -> Workbook:
    """Carga un .xlsx desde disco preservando ``worksheet.tables``."""
    from openpyxl import load_workbook

    return load_workbook(path)


# ── Tests del parser ``PIntParser`` ──────────────────────────────────────


def test_extrae_pint_basico(tmp_path) -> None:
    """Excel con 1 fila válida → DTO con los 12 campos correctos."""
    xlsx_path = _save_xlsx_with_pint_table(
        tmp_path,
        rows=[
            [
                "PI_1_001",  # UID
                "001",  # Numero
                "Proceso Uno",  # Proceso
                "PI1",  # Codigo
                3002,  # Num.DB (PInt DB = 3000+uid+1)
                "Linea A",  # Producto
                "Contador",  # Tipo
                "Numero de ciclos",  # Descripcion
                "DB PINT proceso 1",  # ComentarioDB
                "Si",  # Visibilidad
                3,  # Num.Lista (int)
                "Lista 3 - Ciclos",  # Txt.Lista
            ],
        ],
    )
    wb = _load(xlsx_path)

    result = PIntParser().extraer(wb)

    assert len(result) == 1
    p = result[0]
    assert isinstance(p, ParamIntPLC)
    assert p.uid == "PI_1_001"
    assert p.numero == "001"
    assert p.proceso == "Proceso Uno"
    assert p.codigo == "PI1"
    assert p.num_db == 3002
    assert p.producto == "Linea A"
    assert p.tipo == "Contador"
    assert p.descripcion == "Numero de ciclos"
    assert p.comentario_db == "DB PINT proceso 1"
    assert p.visibilidad == "Si"
    assert p.num_lista == 3  # int preservado
    assert p.txt_lista == "Lista 3 - Ciclos"


def test_hoja_inexistente_devuelve_lista_vacia(tmp_path) -> None:
    """Sheet ``P_INT`` no existe → ``[]`` (no lanza)."""
    xlsx_path = _save_xlsx_with_pint_table(
        tmp_path,
        rows=[["PI_1_001", "001", "P", "PI1", 3002, "", "", "", "", "Si", 0, ""]],
        sheet_name="OTRA_HOJA",
    )
    wb = _load(xlsx_path)

    result = PIntParser().extraer(wb)

    assert result == []


def test_tabla_inexistente_devuelve_lista_vacia(tmp_path) -> None:
    """Sheet existe pero ``Tabla_PInt`` no → ``[]`` (no lanza)."""
    # Construimos un .xlsx con OTRA tabla en P_INT.
    xlsx_path = tmp_path / "otro.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "P_INT"
    ws.append(["X", "Y"])
    ws.append([1, 2])
    table = Table(displayName="Tabla_Otra", ref="A1:B2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(str(xlsx_path))

    wb2 = _load(xlsx_path)
    result = PIntParser().extraer(wb2)
    assert result == []


def test_fila_sin_uid_se_descarta(tmp_path) -> None:
    """Fila con UID vacío (None) NO aparece en el resultado.

    Política consistente con ``ProcesosParser`` y ``PRealParser``
    (dropna por UID). Evita parámetros fantasma sin UID en el
    cache.
    """
    xlsx_path = _save_xlsx_with_pint_table(
        tmp_path,
        rows=[
            [None, "001", "P", "PI1", 3002, "", "", "", "", "Si", 0, ""],  # sin UID
            [
                "PI_1_001",
                "001",
                "Proceso Uno",
                "PI1",
                3002,
                "",
                "",
                "",
                "",
                "Si",
                0,
                "",
            ],  # válida
        ],
    )
    wb = _load(xlsx_path)

    result = PIntParser().extraer(wb)

    # Solo la fila con UID válido se queda.
    assert len(result) == 1
    assert result[0].uid == "PI_1_001"


def test_paramint_y_paramreal_son_tipos_distintos() -> None:
    """``ParamIntPLC`` y ``ParamRealPLC`` son tipos nominalmente distintos (R4).

    Aunque el shape (12 campos) es idéntico, las dos dataclasses
    son tipos distintos en Python. ``isinstance(ParamIntPLC(...),
    ParamRealPLC)`` devuelve ``False``. Esto permite extender
    cada una con campos específicos (p. ej. ``rango_min``/
    ``rango_max`` solo en ``ParamRealPLC``) sin tocar la otra.

    Decisión del operario (R4 del plan, resuelto el 2026-09-01):
    mantener los DTOs separados aunque hoy coincidan en campos.
    """
    p_int = ParamIntPLC(
        uid="PI_1_001",
        numero="001",
        proceso="P",
        codigo="PI1",
        num_db=3002,
        producto="",
        tipo="",
        descripcion="",
        comentario_db="",
        visibilidad="Si",
        num_lista=0,
        txt_lista="",
    )
    p_real = ParamRealPLC(
        uid="PR_1_001",
        numero="001",
        proceso="P",
        codigo="PR1",
        num_db=3001,
        producto="",
        tipo="",
        descripcion="",
        comentario_db="",
        visibilidad="Si",
        num_lista=0,
        txt_lista="",
    )
    # El test crítico de R4: cada DTO NO es instancia del otro.
    assert not isinstance(p_int, ParamRealPLC)
    assert not isinstance(p_real, ParamIntPLC)
    # Y sí son instancias de sí mismos (control de sanidad).
    assert isinstance(p_int, ParamIntPLC)
    assert isinstance(p_real, ParamRealPLC)


def test_num_lista_preserva_texto(tmp_path) -> None:
    """Fila con ``Num.Lista="TODOS"`` → ``dto.num_lista == "TODOS"``.

    Mismo contrato que ``PRealParser``: ``_safe_num_lista``
    preserva los marcadores semánticos del operario (``"N/A"``,
    ``"TODOS"``) como ``str`` en lugar de caer a ``0``.
    """
    xlsx_path = _save_xlsx_with_pint_table(
        tmp_path,
        rows=[
            [
                "PI_1_002",
                "002",
                "Proceso Uno",
                "PI1",
                3002,
                "Linea A",
                "Indice",
                "Todos los items",
                "",
                "Si",
                "TODOS",  # str preservado
                "Lista completa",
            ],
        ],
    )
    wb = _load(xlsx_path)

    result = PIntParser().extraer(wb)

    assert len(result) == 1
    p = result[0]
    assert p.num_lista == "TODOS"
    # El tipo debe ser exactamente ``str`` (no int caído a 0).
    assert isinstance(p.num_lista, str)
    assert not isinstance(p.num_lista, int)
