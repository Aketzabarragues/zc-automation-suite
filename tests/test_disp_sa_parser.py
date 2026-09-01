"""Tests del ``DispSAParser`` (Fase 5 del plan).

Estructura idéntica a ``DispEA`` (mismos campos y semántica;
solo cambia el sentido: salida vs entrada). Mismos 5 tests.
"""
from __future__ import annotations

from areas.alimentacion.domain.models.excel_cache import DispSA
from areas.alimentacion.infrastructure.parsers.disp_sa import DispSAParser

from tests._disp_parser_test_helpers import (
    build_full_row,
    load_workbook_safe,
    save_xlsx_with_disp_table,
)


def test_extrae_disp_sa_basico(tmp_path) -> None:
    """Excel con 1 fila válida → ``DispSA`` con ``rii``/``rsi`` float."""
    xlsx_path = save_xlsx_with_disp_table(
        tmp_path, "SA",
        rows=[build_full_row("SA", UID="SA_001", Numero=1,
                              plc_tag="V_SA_001",
                              UNIDADES="HZ", RII=0.0, RSI=50.0)],
    )
    wb = load_workbook_safe(xlsx_path)
    result = DispSAParser().extraer(wb)
    assert len(result) == 1
    d = result[0]
    assert isinstance(d, DispSA)
    assert d.uid == "SA_001"
    assert d.plc_tag == "V_SA_001"
    assert d.unidades == "HZ"
    assert d.rii == 0.0
    assert d.rsi == 50.0


def test_hoja_inexistente_devuelve_lista_vacia(tmp_path) -> None:
    xlsx_path = save_xlsx_with_disp_table(
        tmp_path, "SA",
        rows=[build_full_row("SA")],
        sheet_name="OTRA_HOJA",
    )
    wb = load_workbook_safe(xlsx_path)
    assert DispSAParser().extraer(wb) == []


def test_tabla_inexistente_devuelve_lista_vacia(tmp_path) -> None:
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo

    xlsx_path = tmp_path / "otro.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "DISP_SA"
    ws.append(["X", "Y"])
    ws.append([1, 2])
    table = Table(displayName="Tabla_Otra", ref="A1:B2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(str(xlsx_path))

    wb2 = load_workbook_safe(xlsx_path)
    assert DispSAParser().extraer(wb2) == []


def test_fila_sin_uid_se_descarta(tmp_path) -> None:
    xlsx_path = save_xlsx_with_disp_table(
        tmp_path, "SA",
        rows=[
            [None, None, "X", "X", "X"],
            build_full_row("SA", UID="SA_001"),
        ],
    )
    wb = load_workbook_safe(xlsx_path)
    result = DispSAParser().extraer(wb)
    assert len(result) == 1
    assert result[0].uid == "SA_001"


def test_defaults_when_only_uid_and_numero(tmp_path) -> None:
    """Solo UID+Numero+PLC.Tag → resto defaults."""
    xlsx_path = save_xlsx_with_disp_table(
        tmp_path, "SA",
        rows=[["SA_001", 1, "V_SA_001", "c", "d"]],
        headers=["UID", "Numero", "PLC.Tag", "PLC.Comentario", "Descripcion"],
    )
    wb = load_workbook_safe(xlsx_path)
    result = DispSAParser().extraer(wb)
    assert len(result) == 1
    d = result[0]
    assert d.unidades == ""
    assert d.rii == 0.0
    assert d.rsi == 0.0
