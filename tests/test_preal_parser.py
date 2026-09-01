"""Tests del parser ``PRealParser`` y del helper ``_safe_num_lista``.

Cubre la extracción de ``Tabla_PReal`` (hoja ``P_REAL``) y la
preservación de ``num_lista`` como ``int | str`` (helper
``_safe_num_lista``), que es el punto crítico de Fase 2 del plan
``_plan/04_excel_cache_phased_plan.md``: el operario usa valores
como ``"N/A"`` o ``"TODOS"`` como marcadores semánticos y el
parser no debe destruirlos cayendo a ``0``.

Convenciones:
    * Excel sintético construido en ``tmp_path`` con ``Table`` real
      (R5 del plan).
    * Se carga con ``load_workbook`` para garantizar que la
      ``ListObject`` se registre en ``worksheet.tables``.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from areas.alimentacion.domain.models.excel_cache import ParamRealPLC
from areas.alimentacion.infrastructure.parsers._xlsx_helpers import (
    _safe_num_lista,
)
from areas.alimentacion.infrastructure.parsers.preal import PRealParser


# ── Helpers de construcción de Excels sintéticos ────────────────────────


def _save_xlsx_with_preal_table(
    tmp_path,
    rows: list[list] | None,
    *,
    sheet_name: str = "P_REAL",
    table_name: str = "Tabla_PReal",
    headers: list[str] | None = None,
) -> str:
    """Crea un .xlsx sintético con la tabla de parámetros reales.

    Args:
        tmp_path: fixture pytest de path temporal.
        rows: lista de filas de datos (``None`` o ``[]`` → solo
            cabeceras, útil para verificar que la tabla existe
            pero está vacía).
        sheet_name: nombre de la hoja (default ``P_REAL``).
        table_name: nombre de la ``ListObject`` (default
            ``Tabla_PReal``).
        headers: cabeceras (default: las del Excel legacy del
            corporativo, 12 columnas).

    Returns:
        Path absoluto del .xlsx en str.
    """
    if headers is None:
        headers = [
            "UID",
            "Numero",
            "Proceso",
            "Codigo",
            "Num.DB",
            "Producto",
            "Tipo",
            "Descripcion",
            "ComentarioDB",
            "Visibilidad",
            "Num.Lista",
            "Txt.Lista",
        ]
    if rows is None:
        rows = []

    xlsx_path = tmp_path / f"{table_name}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)

    # Registrar la Table (R5 del plan: sin esto, ``worksheet.tables``
    # no contiene la ``ListObject`` y el parser no la encontraría).
    last_col_letter = chr(ord("A") + len(headers) - 1)
    last_row = 1 + len(rows)
    ref = f"A1:{last_col_letter}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(str(xlsx_path))
    return str(xlsx_path)


def _load(path: str) -> Workbook:
    """Carga un .xlsx desde disco preservando ``worksheet.tables``."""
    from openpyxl import load_workbook

    return load_workbook(path)


# ── Tests del helper ``_safe_num_lista`` ─────────────────────────────────


def test_safe_num_lista_int() -> None:
    """Valores numéricos (int, float, str-numérico) → int."""
    assert _safe_num_lista(1) == 1
    assert _safe_num_lista(1.0) == 1
    assert _safe_num_lista("5") == 5
    assert _safe_num_lista("5.0") == 5
    # float se trunca, no redondea
    assert _safe_num_lista(1.9) == 1
    # str con whitespace se castea correctamente
    assert _safe_num_lista("  7  ") == 7


def test_safe_num_lista_str_preserva_texto() -> None:
    """Texto no numérico se preserva literal como ``str`` (R4 del plan)."""
    assert _safe_num_lista("N/A") == "N/A"
    assert _safe_num_lista("TODOS") == "TODOS"
    # También preserva case y whitespace interno
    assert _safe_num_lista("n/a") == "n/a"
    assert _safe_num_lista("Todos") == "Todos"


def test_safe_num_lista_none_vacio_nan_devuelve_cero() -> None:
    """None / vacío / ``"nan"`` / whitespace → ``0`` (int)."""
    assert _safe_num_lista(None) == 0
    assert _safe_num_lista("") == 0
    assert _safe_num_lista("nan") == 0
    assert _safe_num_lista("NaN") == 0  # case-insensitive
    assert _safe_num_lista("None") == 0
    assert _safe_num_lista("null") == 0
    assert _safe_num_lista("  ") == 0  # whitespace puro


# ── Tests del parser ``PRealParser`` ─────────────────────────────────────


def test_extrae_preal_basico(tmp_path) -> None:
    """Excel con 1 fila válida → DTO con los 12 campos correctos."""
    xlsx_path = _save_xlsx_with_preal_table(
        tmp_path,
        rows=[
            [
                "PR_1_001",  # UID
                "001",  # Numero
                "Proceso Uno",  # Proceso
                "PR1",  # Codigo
                3001,  # Num.DB
                "Linea A",  # Producto
                "Setpoint",  # Tipo
                "Temperatura objetivo",  # Descripcion
                "DB PREAL proceso 1",  # ComentarioDB
                "Si",  # Visibilidad
                2,  # Num.Lista (int)
                "Lista 2 - Temp",  # Txt.Lista
            ],
        ],
    )
    wb = _load(xlsx_path)

    result = PRealParser().extraer(wb)

    assert len(result) == 1
    p = result[0]
    assert isinstance(p, ParamRealPLC)
    assert p.uid == "PR_1_001"
    assert p.numero == "001"
    assert p.proceso == "Proceso Uno"
    assert p.codigo == "PR1"
    assert p.num_db == 3001
    assert p.producto == "Linea A"
    assert p.tipo == "Setpoint"
    assert p.descripcion == "Temperatura objetivo"
    assert p.comentario_db == "DB PREAL proceso 1"
    assert p.visibilidad == "Si"
    assert p.num_lista == 2  # int preservado
    assert p.txt_lista == "Lista 2 - Temp"


def test_num_lista_preserva_texto(tmp_path) -> None:
    """Fila con ``Num.Lista="N/A"`` → ``dto.num_lista == "N/A"`` (R4).

    Caso crítico: el operario usa ``"N/A"`` como marcador
    semántico. Si el parser usara ``_safe_int``, caería a ``0`` y
    el dato se perdería. ``_safe_num_lista`` lo preserva como
    ``str``.
    """
    xlsx_path = _save_xlsx_with_preal_table(
        tmp_path,
        rows=[
            [
                "PR_1_002",  # UID
                "002",  # Numero
                "Proceso Uno",  # Proceso
                "PR1",  # Codigo
                3001,  # Num.DB
                "Linea A",  # Producto
                "Limite",  # Tipo
                "Sin lista",  # Descripcion
                "",  # ComentarioDB
                "No",  # Visibilidad
                "N/A",  # Num.Lista (str preservado)
                "",  # Txt.Lista
            ],
        ],
    )
    wb = _load(xlsx_path)

    result = PRealParser().extraer(wb)

    assert len(result) == 1
    p = result[0]
    assert p.num_lista == "N/A"
    # El tipo debe ser exactamente ``str`` (no int caído a 0).
    assert isinstance(p.num_lista, str)
    assert not isinstance(p.num_lista, int)


def test_num_lista_preserva_todos(tmp_path) -> None:
    """Fila con ``Num.Lista="TODOS"`` → ``dto.num_lista == "TODOS"``."""
    xlsx_path = _save_xlsx_with_preal_table(
        tmp_path,
        rows=[
            [
                "PR_1_003",
                "003",
                "Proceso Uno",
                "PR1",
                3001,
                "Linea A",
                "Setpoint",
                "Todos los items",
                "",
                "Si",
                "TODOS",  # str preservado
                "Lista completa",
            ],
        ],
    )
    wb = _load(xlsx_path)

    result = PRealParser().extraer(wb)

    assert len(result) == 1
    assert result[0].num_lista == "TODOS"
    assert isinstance(result[0].num_lista, str)


def test_hoja_inexistente_devuelve_lista_vacia(tmp_path) -> None:
    """Sheet ``P_REAL`` no existe → ``[]`` (no lanza)."""
    xlsx_path = _save_xlsx_with_preal_table(
        tmp_path,
        rows=[["PR_1_001", "001", "P", "PR1", 3001, "", "", "", "", "Si", 0, ""]],
        sheet_name="OTRA_HOJA",
    )
    wb = _load(xlsx_path)

    result = PRealParser().extraer(wb)

    assert result == []


def test_tabla_inexistente_devuelve_lista_vacia(tmp_path) -> None:
    """Sheet existe pero ``Tabla_PReal`` no → ``[]`` (no lanza)."""
    # Construimos un .xlsx con OTRA tabla en P_REAL.
    xlsx_path = tmp_path / "otro.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "P_REAL"
    ws.append(["X", "Y"])
    ws.append([1, 2])
    table = Table(displayName="Tabla_Otra", ref="A1:B2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(str(xlsx_path))

    wb2 = _load(xlsx_path)
    result = PRealParser().extraer(wb2)
    assert result == []


def test_fila_sin_uid_se_descarta(tmp_path) -> None:
    """Fila con UID vacío (None) NO aparece en el resultado.

    Política consistente con ``ProcesosParser`` (dropna por UID).
    Evita parámetros fantasma sin UID en el cache.
    """
    xlsx_path = _save_xlsx_with_preal_table(
        tmp_path,
        rows=[
            [None, "001", "P", "PR1", 3001, "", "", "", "", "Si", 0, ""],  # sin UID
            [
                "PR_1_001",
                "001",
                "Proceso Uno",
                "PR1",
                3001,
                "",
                "",
                "",
                "",
                "Si",
                0,
                "",
            ],  # válida
        ],
    )
    wb = _load(xlsx_path)

    result = PRealParser().extraer(wb)

    # Solo la fila con UID válido se queda.
    assert len(result) == 1
    assert result[0].uid == "PR_1_001"


def test_paramreal_no_tiene_properties_derivadas() -> None:
    """``ParamRealPLC`` NO tiene properties (a diferencia de ``ProcesoPLC``).

    Documentado en el plan §6.3: la lógica de derivación de
    nombres de DB vive en ``ProcesoPLC`` (que conoce el ``num_db``
    raíz del proceso). ``ParamRealPLC`` solo tiene los 12 campos
    del Excel, sin properties.
    """
    p = ParamRealPLC(
        uid="PR_1_001",
        numero="001",
        proceso="P",
        codigo="PR1",
        num_db=3001,
        producto="",
        tipo="",
        descripcion="",
        comentario_db="",
        visibilidad="Si",
        num_lista=0,
        txt_lista="",
    )
    # Verificar que NO existen properties de derivación de DB.
    assert not hasattr(p, "db_numero")
    assert not hasattr(p, "db_nombre")
    # El campo ``num_db`` sigue siendo accesible como atributo normal.
    assert p.num_db == 3001
