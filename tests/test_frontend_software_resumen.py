"""Smoke test del wiring SPA ↔ secciones de software (Fase 6).

La SPA es Vue 3 ESM sin build step y sin infra de tests JS
(Jest/Vitest no están en el repo). En lugar de importar los
``.js`` desde pytest (que necesitaría Node + ESM + DOM mock),
este test verifica la **forma textual** del wiring: que los
símbolos públicos esperados aparezcan en los archivos correctos
(``store.js`` documenta el shape nuevo,
``DefinicionProgramacion.js`` renderiza 4 secciones + banner ámbar,
``diagnostics.py`` expone el response con los 4 campos nuevos).

Es un contract check barato. Si en el futuro se añade infra JS,
se puede sustituir por tests unitarios de verdad sobre el
componente ``DefinicionProgramacion``.

Marcado con ``@pytest.mark.frontend_smoke`` para permitir
filtrado (``pytest -m frontend_smoke``).

Plan canónico de referencia:
``_plan/04_excel_cache_phased_plan.md`` §10 (Fase 6).
"""
from __future__ import annotations

import json
import sys
from io import BytesIO
from pathlib import Path
from typing import Any
from unittest.mock import MagicMock

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from core.application.state import AppState, get_app_state  # noqa: E402
from core.infrastructure.config_manager import ConfigManager  # noqa: E402
from core.infrastructure.gateway import TIAProcessGateway  # noqa: E402
from interfaces.web_server.app import create_app  # noqa: E402

REPO_ROOT = Path(__file__).resolve().parent.parent

STORE_JS = (
    REPO_ROOT
    / "interfaces"
    / "web_server"
    / "static"
    / "js"
    / "store.js"
)
DEFINICION_PROGRAMACION_JS = (
    REPO_ROOT
    / "areas"
    / "alimentacion"
    / "frontend"
    / "components"
    / "DefinicionProgramacion.js"
)
# Espejo del panel donde, tras el rediseño Opción A (tabs principales),
# vive realmente el banner ámbar y las 4 secciones de software
# (Procesos / PInt / PReal / Alarmas). Antes vivían inline en
# ``DefinicionProgramacion.js``; el refactor los movió a este
# sub-componente para dar simetría visual con ``DispositivosPanel``.
# Los contract checks textuales que verifican banner + 4 secciones
# apuntan aquí, no a ``DefinicionProgramacion.js``.
SOFTWARE_PANEL_JS = (
    REPO_ROOT
    / "areas"
    / "alimentacion"
    / "frontend"
    / "components"
    / "SoftwarePanel.js"
)
DIAGNOSTICS_PY = (
    REPO_ROOT
    / "interfaces"
    / "web_server"
    / "routers"
    / "diagnostics.py"
)
STYLES_CSS = REPO_ROOT / "interfaces" / "web_server" / "static" / "styles.css"


pytestmark = pytest.mark.frontend_smoke


# ── Helpers ──────────────────────────────────────────────────────────────


def _read(path: Path) -> str:
    assert path.exists(), f"Missing required file: {path}"
    return path.read_text(encoding="utf-8")


# ── Configuración JSON fixture (mínima) ─────────────────────────────────


_FULL_CONFIG: dict[str, Any] = {
    "departments": {
        "alimentacion": {
            "global_config_table_name": "000_Config_Dispositivos",
            "tia_folders": {
                "proceso":      "003_Procesos",
                "dispositivos": "2000_Dispositivos",
                "nmax":         "000_Sistema",
            },
            "n_max_catalog": [
                {"name": "N_MAX_DISP_ED",   "value": 10},
                {"name": "N_MAX_DISP_EA",   "value": 10},
                {"name": "N_MAX_DISP_SA",   "value": 10},
                {"name": "N_MAX_DISP_V",    "value": 10},
                {"name": "N_MAX_DISP_M",    "value": 10},
                {"name": "N_MAX_DISP_M_VF", "value": 10},
            ],
            "Dispositivos": {
                "ed": {
                    "db_name": "DB2000_ED", "db_array_name": "ED",
                    "tag_table": "2000_Disp_ED",
                    "config_table": "000_Config_Dispositivos",
                },
                "ea": {
                    "db_name": "DB2001_EA", "db_array_name": "EA",
                    "tag_table": "2000_Disp_EA",
                    "config_table": "000_Config_Dispositivos",
                },
                "sa": {
                    "db_name": "DB2006_SA", "db_array_name": "SA",
                    "tag_table": "2000_Disp_SA",
                    "config_table": "000_Config_Dispositivos",
                },
                "v": {
                    "db_name": "DB2010_V", "db_array_name": "V",
                    "tag_table": "2000_Disp_V",
                    "config_table": "000_Config_Dispositivos",
                },
                "m": {
                    "db_name": "DB2015_M", "db_array_name": "M",
                    "tag_table": "2000_Disp_M",
                    "config_table": "000_Config_Dispositivos",
                },
                "m_vf": {
                    "db_name": "DB2016_M_VF", "db_array_name": "M_VF",
                    "tag_table": "2000_Disp_M_VF",
                    "config_table": "000_Config_Dispositivos",
                },
            },
        }
    }
}


# ── Excel sintético (1 fila por tipo de dispositivo) ─────────────────────


def _add_table(
    wb: Workbook, sheet_name: str, table_name: str,
    headers: list[str], rows: list[list],
) -> None:
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    last_col_letter = chr(ord("A") + len(headers) - 1)
    last_row = 1 + len(rows)
    ref = f"A1:{last_col_letter}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)


def _build_minimal_xlsx_bytes() -> bytes:
    """Genera un xlsx en memoria con 1 fila de cada uno de los 6 tipos."""
    wb = Workbook()
    wb.remove(wb.active)
    tables = [
        ("DISP_ED",   "Tabla_Disp_ED",   "ED_001",   1, "V_ED_001"),
        ("DISP_EA",   "Tabla_Disp_EA",   "EA_001",   1, "V_EA_001"),
        ("DISP_SA",   "Tabla_Disp_SA",   "SA_001",   1, "V_SA_001"),
        ("DISP_V",    "Tabla_Disp_V",    "V_001",    1, "V_V_001"),
        ("DISP_M",    "Tabla_Disp_M",    "M_001",    1, "V_M_001"),
        ("DISP_M_VF", "Tabla_Disp_M_VF", "MVF_001",  1, "V_MVF_001"),
    ]
    headers = ["UID", "Numero", "PLC.Tag", "Descripcion"]
    for sheet_name, table_name, uid, numero, tag in tables:
        _add_table(wb, sheet_name, table_name, headers,
                   [[uid, numero, tag, f"Desc {uid}"]])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Fixtures ──────────────────────────────────────────────────────────────


@pytest.fixture
def app_with_overrides(tmp_path: Path):
    """App con ``ConfigManager`` apuntando al fixture JSON y estado limpio."""
    config_path = tmp_path / "config.json"
    config_path.write_text(json.dumps(_FULL_CONFIG), encoding="utf-8")
    cm = ConfigManager(config_path=config_path)
    gateway = MagicMock(spec=TIAProcessGateway)
    app = create_app(gateway)
    app.state.config_manager = cm

    state = get_app_state()
    state.reset()
    state.excel_cache = None
    state.excel_path = None
    state.dimensiones = None

    yield app, state

    state.reset()
    state.excel_cache = None
    state.excel_path = None
    state.dimensiones = None


# ── Tests ────────────────────────────────────────────────────────────────


def test_memory_state_shape_tras_upload_incluye_software(
    app_with_overrides,
) -> None:
    """El response de ``/state/dispositivos`` tiene los 4 campos + el flag.

    Simula un upload (POST) y luego verifica que el GET de
    ``/state/dispositivos`` expone los 4 nuevos campos de software
    + el flag ``software_parsers_implemented=true``.

    Este test verifica el **contrato de la API** desde la
    perspectiva del frontend: si mañana alguien borra un campo
    del response, este test rompe.
    """
    app, _state = app_with_overrides
    xlsx_bytes = _build_minimal_xlsx_bytes()
    with TestClient(app) as client:
        # Upload (popula el cache).
        upload_resp = client.post(
            "/api/v1/excel/upload",
            files={"file": ("test.xlsx", xlsx_bytes,
                            "application/octet-stream")},
        )
        assert upload_resp.status_code == 200

        # Get state (la SPA hace esto tras un upload para repintar).
        state_resp = client.get("/api/v1/state/dispositivos")
        assert state_resp.status_code == 200
        body = state_resp.json()

    # Las 4 keys de software + el flag están presentes.
    assert "procesos" in body
    assert "parametros_int" in body
    assert "parametros_real" in body
    assert "alarmas" in body
    assert "software_parsers_implemented" in body

    # El flag es ``true`` (cache real tiene el atributo en True).
    assert body["software_parsers_implemented"] is True

    # Los arrays son listas (pueden ser vacías si el Excel sintético
    # no tiene las hojas de software — la SPA renderiza sin error).
    assert isinstance(body["procesos"], list)
    assert isinstance(body["parametros_int"], list)
    assert isinstance(body["parametros_real"], list)
    assert isinstance(body["alarmas"], list)


def test_software_parsers_implemented_false_sin_upload(
    app_with_overrides,
) -> None:
    """Pre-upload: el flag es ``false`` y los 4 arrays están vacíos.

    Contrato back-compat: si un frontend cacheado (service worker)
    pide ``/state/dispositivos`` antes de que el operario suba un
    Excel, debe recibir ``software_parsers_implemented=false`` y
    4 arrays vacíos. La SPA pinta el banner ámbar en ese caso.
    """
    app, _state = app_with_overrides
    # Sin upload previo.
    with TestClient(app) as client:
        resp = client.get("/api/v1/state/dispositivos")
        assert resp.status_code == 200
        body = resp.json()

    assert body["software_parsers_implemented"] is False
    assert body["procesos"] == []
    assert body["parametros_int"] == []
    assert body["parametros_real"] == []
    assert body["alarmas"] == []


def test_banner_amber_eliminado_para_simetria_con_dispositivos() -> None:
    """El panel de software NO tiene banner ámbar: el operario pidió
    eliminarlo por consistencia visual con ``DispositivosPanel``.

    Cuando no hay datos, ambos tabs deben mostrar el mismo
    patrón: las tablas con su mensaje "Sin X definidos" / "La
    pestaña X no contiene dispositivos" en cada fila. No hay
    "Inspector vacío" ni banner ámbar: la única indicación de
    "datos faltantes" son los mensajes de las propias tablas.

    El banner ámbar existía como indicación visual para el caso
    ``software_parsers_implemented === false`` (backend degradado
    o Excel sin las 4 hojas de software). El operario lo eliminó
    para que ambos tabs respondan igual a la ausencia de datos.

    Contract check textual: verifica que el banner NO está en el
    archivo (clases ámbar, computed softwareImplemented, copy).
    """
    text = _read(SOFTWARE_PANEL_JS)
    # Las clases del banner ámbar ya no deben estar.
    assert "bg-amber-100" not in text, (
        "El banner ámbar (bg-amber-100) debería estar eliminado"
    )
    assert "border-amber-500" not in text, (
        "El banner ámbar (border-amber-500) debería estar eliminado"
    )
    # El computed softwareImplemented ya no debe existir (huérfano
    # tras la eliminación del banner).
    assert "softwareImplemented" not in text, (
        "El computed softwareImplemented debería estar eliminado (ya no se usa)"
    )
    # El copy del banner ya no debe estar.
    assert "Datos de software pendientes" not in text, (
        "El texto del banner ámbar debería estar eliminado"
    )


def test_definicion_programacion_tiene_4_secciones_software() -> None:
    """El panel de software declara 4 secciones de datos
    (Procesos / PInt / PReal / Alarmas), una por dominio.

    Tras el rediseño Opción A, las 4 secciones ya NO son
    ``<details>`` plegables: el sub-componente ``SoftwarePanel``
    las renderiza como **4 ``<table>`` con ``v-if``** (uno por
    sub-tab, mutuamente excluyentes con ``v-else-if``) más 4
    ``<tr>`` de "fila vacía" (``v-if="<dominio>.length === 0"``).
    Esto le da simetría visual con ``DispositivosPanel`` (que
    también es tabla + sub-tab).

    El contrato que verifica este test es: el panel expone los 4
    dominios de software, no la forma concreta ``<details>`` vs
    ``<table>``. Si en el futuro se cambia el layout (p. ej. a
    listas, cards, etc.) basta con adaptar las aserciones
    mecánicas; el espíritu "4 secciones, una por DTO" se mantiene.
    """
    text = _read(SOFTWARE_PANEL_JS)
    # 4 ``<table>`` (uno por dominio), cada uno con la clase firma
    # ``class="w-full text-xs"``. El test cuenta aperturas que
    # contengan esa clase exacta; los ``v-else-if`` también abren
    # un ``<table>`` (es la forma del template encadenado en Vue 3).
    # Patrón usado: ``<table v-if=... class="w-full text-xs">`` o
    # ``<table v-else-if=... class="w-full text-xs">``.
    table_open = text.count('class="w-full text-xs"')
    assert table_open == 4, (
        f"Esperaba 4 <table> con class='w-full text-xs', encontré {table_open}"
    )
    # 4 ``</table>`` de cierre.
    assert text.count("</table>") == 4, (
        f"Esperaba 4 </table>, encontré {text.count('</table>')}"
    )
    # 4 referencias a los nombres de dominio en los v-if de "fila vacía"
    # (uno por tabla). Patrón: ``v-if="<dominio>.length === 0"``.
    for domain in ("procesos", "parametrosInt", "parametrosReal", "alarmas"):
        assert f"v-if=\"{domain}.length === 0\"" in text, (
            f"Falta la fila vacía v-if='{domain}.length === 0' "
            f"para el dominio {domain}"
        )
    # Los 4 computed properties existen en el setup (siguen
    # siendo la fuente de datos de las tablas).
    for ref in ("procesos", "parametrosInt", "parametrosReal", "alarmas"):
        assert f"const {ref} = computed" in text, (
            f"Falta el computed {ref} en el setup"
        )


def test_store_js_documenta_shape_de_software() -> None:
    """``store.js`` documenta los 4 nuevos campos + el flag en ``memoryState``.

    Es un comment JSDoc: no afecta al runtime pero el contrato
    escrito es la fuente de verdad para IDEs y agentes AI.
    """
    text = _read(STORE_JS)
    # Las 4 keys aparecen en el docstring de ``memoryState``.
    for key in (
        "procesos",
        "parametros_int",
        "parametros_real",
        "alarmas",
        "software_parsers_implemented",
    ):
        assert key in text, (
            f"Falta la key {key!r} en el docstring de memoryState"
        )


def test_diagnostics_py_retorna_5_keys_de_software() -> None:
    """``diagnostics.py`` extrae y devuelve los 4 arrays + el flag.

    Contract check textual: el helper ``_extract_software_from_cache``
    existe y el endpoint lo usa (``** _extract_software_from_cache(state)``).
    """
    text = _read(DIAGNOSTICS_PY)
    assert "_extract_software_from_cache" in text, (
        "Falta el helper _extract_software_from_cache"
    )
    # El endpoint lo invoca con ``**``.
    assert "** _extract_software_from_cache(state)" in text, (
        "Falta la invocación ** _extract_software_from_cache(state) "
        "en el return del endpoint"
    )
    # El helper maneja los 4 dominios.
    for domain in (
        "cache.procesos",
        "cache.parametros_int",
        "cache.parametros_real",
        "cache.alarmas",
    ):
        assert domain in text, (
            f"Falta la referencia a {domain} en el helper"
        )
    # Y el flag.
    assert "software_parsers_implemented" in text, (
        "Falta la referencia a software_parsers_implemented"
    )


def test_styles_css_regenerated_post_fase_6() -> None:
    """El bundle CSS existe y tiene tamaño no trivial tras la recompilación.

    El input.css ya incluye el glob ``areas/**/frontend/**/*.js``
    (ver ``interfaces/web_server/static/src/input.css``), así que
    las clases nuevas de Fase 6 (incluido ``bg-amber-100`` del
    banner) se detectan automáticamente. Aquí solo validamos que
    la recompilación efectivamente regeneró el bundle.
    """
    assert STYLES_CSS.exists(), (
        "styles.css no se regeneró — ejecutar run_tailwind.bat "
        "antes de commit"
    )
    size = STYLES_CSS.stat().st_size
    assert size > 10_000, (
        f"styles.css demasiado pequeño ({size} bytes); "
        "recompilar Tailwind antes de commit"
    )
    # Tras la recompilación, la clase ``bg-amber-100`` debe existir
    # en el bundle (es nueva en Fase 6).
    css = STYLES_CSS.read_text(encoding="utf-8")
    assert "bg-amber-100" in css, (
        "La clase bg-amber-100 (banner ámbar) no aparece en "
        "styles.css — recompilar Tailwind antes de commit"
    )
