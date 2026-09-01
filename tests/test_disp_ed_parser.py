"""Tests del ``DispEDParser`` (Fase 5 del plan).

Cubre la extracción de ``Tabla_Disp_ED`` (hoja ``DISP_ED``) y la
construcción de ``DispED``. Sigue el patrón de las tests de los
parsers de software (Fase 1-4): Excel sintético con ``Table`` real
en ``tmp_path``.

Convenciones:
    * Cabeceras literales del Excel (con puntos y mayúsculas).
    * ``extract_list_object_rows`` ya está cubierta por
      ``test_xlsx_helpers.py``; aquí solo verificamos que el parser
      rellena los campos del DTO correctamente.
    * Las filas sin UID ni Numero se descartan (criterio legacy).
    * Las filas inválidas se descartan con WARNING sin romper la tabla.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from areas.alimentacion.domain.models.excel_cache import DispED
from areas.alimentacion.infrastructure.parsers.disp_ed import DispEDParser


# ── Helpers ─────────────────────────────────────────────────────────────


def _save_xlsx_with_disp_table(
    tmp_path,
    rows: list[list] | None,
    *,
    sheet_name: str = "DISP_ED",
    table_name: str = "Tabla_Disp_ED",
    headers: list[str] | None = None,
) -> str:
    """Crea un .xlsx sintético con la ``Tabla_Disp_ED``."""
    if headers is None:
        # Cabeceras del Excel legacy (subset mínimo para los tests).
        headers = [
            "UID", "Numero", "PLC.Tag", "PLC.Comentario", "Descripcion",
            "Tag", "FAT", "E.Byte", "E.Bit", "Gr.Alarma", "Cuadro",
            "Observaciones", "PLC.Tipo", "PLC.Index", "Hmi.Index",
            "Hmi.Texto", "Cfg.Habilitar", "Cfg.ByteEntrada",
            "Cfg.BitEntrada", "Cfg.GrupoAlarma", "ComentarioDB",
        ]
    if rows is None:
        rows = []
    xlsx_path = tmp_path / f"{table_name}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    last_col_letter = chr(ord("A") + len(headers) - 1)
    last_row = 1 + len(rows)
    ref = f"A1:{last_col_letter}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(str(xlsx_path))
    return str(xlsx_path)


def _load(path: str) -> Workbook:
    from openpyxl import load_workbook
    return load_workbook(path)


# ── Tests ───────────────────────────────────────────────────────────────


def test_extrae_disp_ed_basico(tmp_path) -> None:
    """Excel con 1 fila válida → ``DispED`` con los campos correctos."""
    xlsx_path = _save_xlsx_with_disp_table(
        tmp_path,
        rows=[
            [
                "ED_001",  # UID
                1,  # Numero
                "V_ED_001",  # PLC.Tag
                "Comentario PLC",  # PLC.Comentario
                "Entrada digital 1",  # Descripcion
                "T_ED",  # Tag
                "F_ED",  # FAT
                0,  # E.Byte
                0,  # E.Bit
                1,  # Gr.Alarma
                "C1",  # Cuadro
                "Obs",  # Observaciones
                "Bool",  # PLC.Tipo
                0,  # PLC.Index
                1,  # Hmi.Index
                "Texto HMI",  # Hmi.Texto
                "cfg_hab := 1;",  # Cfg.Habilitar
                "cfg_byte := 0;",  # Cfg.ByteEntrada
                "cfg_bit := 0;",  # Cfg.BitEntrada
                "cfg_ga := 1;",  # Cfg.GrupoAlarma
                "DB ED comentario",  # ComentarioDB
            ],
        ],
    )
    wb = _load(xlsx_path)

    result = DispEDParser().extraer(wb)

    assert len(result) == 1
    d = result[0]
    assert isinstance(d, DispED)
    assert d.uid == "ED_001"
    assert d.numero == 1
    assert d.plc_tag == "V_ED_001"
    assert d.plc_comentario == "Comentario PLC"
    assert d.descripcion == "Entrada digital 1"
    assert d.tag == "T_ED"
    assert d.fat == "F_ED"
    assert d.e_byte == 0
    assert d.e_bit == 0
    assert d.gr_alarma == 1
    assert d.cuadro == "C1"
    assert d.observaciones == "Obs"
    assert d.plc_tipo == "Bool"
    assert d.plc_index == 0
    assert d.hmi_index == 1
    assert d.hmi_texto == "Texto HMI"
    # Los ``cfg_*`` preservan las líneas SCL crudas.
    assert d.cfg_habilitar == "cfg_hab := 1;"
    assert d.cfg_byte_entrada == "cfg_byte := 0;"
    assert d.cfg_bit_entrada == "cfg_bit := 0;"
    assert d.cfg_grupo_alarma == "cfg_ga := 1;"
    assert d.comentario_db == "DB ED comentario"


def test_hoja_inexistente_devuelve_lista_vacia(tmp_path) -> None:
    """Hoja ``DISP_ED`` no existe → ``[]`` (no lanza)."""
    xlsx_path = _save_xlsx_with_disp_table(
        tmp_path,
        rows=[["ED_001", 1, "X", "X", "X"]],
        sheet_name="OTRA_HOJA",
    )
    wb = _load(xlsx_path)
    result = DispEDParser().extraer(wb)
    assert result == []


def test_tabla_inexistente_devuelve_lista_vacia(tmp_path) -> None:
    """Hoja existe pero ``Tabla_Disp_ED`` no → ``[]`` (no lanza)."""
    xlsx_path = tmp_path / "otro.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "DISP_ED"
    ws.append(["X", "Y"])
    ws.append([1, 2])
    table = Table(displayName="Tabla_Otra", ref="A1:B2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(str(xlsx_path))

    wb2 = _load(xlsx_path)
    assert DispEDParser().extraer(wb2) == []


def test_fila_sin_uid_se_descarta(tmp_path) -> None:
    """Fila sin UID ni Numero se descarta (criterio legacy)."""
    xlsx_path = _save_xlsx_with_disp_table(
        tmp_path,
        rows=[
            [None, None, "X", "X", "X"],  # sin UID ni Numero
            ["ED_001", 1, "V_ED_001", "c", "d"],  # válida
        ],
    )
    wb = _load(xlsx_path)
    result = DispEDParser().extraer(wb)
    assert len(result) == 1
    assert result[0].uid == "ED_001"


def test_defaults_when_only_uid_and_numero(tmp_path) -> None:
    """Si solo hay UID+Numero+PLC.Tag, el resto son defaults tolerantes."""
    xlsx_path = _save_xlsx_with_disp_table(
        tmp_path,
        rows=[["ED_001", 1, "V_ED_001", "c", "d"]],
    )
    wb = _load(xlsx_path)
    result = DispEDParser().extraer(wb)
    assert len(result) == 1
    d = result[0]
    # El resto son defaults (str "" / int 0).
    assert d.tag == ""
    assert d.fat == ""
    assert d.e_byte == 0
    assert d.e_bit == 0
    assert d.gr_alarma == 0
    assert d.comentario_db == ""
