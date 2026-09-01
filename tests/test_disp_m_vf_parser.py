"""Tests del ``DispM_VFParser`` (Fase 5 del plan).

Cubre la extracción de ``Tabla_Disp_M_VF`` (hoja ``DISP_M_VF``) y la
construcción de ``DispM_VF`` (motores con variador de frecuencia).
Los campos exclusivos ``sa_byte`` y ``cfg_byteanalogica`` se
verifican explícitamente.
"""
from __future__ import annotations

from areas.alimentacion.domain.models.excel_cache import DispM_VF
from areas.alimentacion.infrastructure.parsers.disp_m_vf import DispM_VFParser

from tests._disp_parser_test_helpers import (
    build_full_row,
    load_workbook_safe,
    save_xlsx_with_disp_table,
)


def test_extrae_disp_m_vf_basico(tmp_path) -> None:
    """1 fila con ``sa_byte`` y ``cfg_byteanalogica`` populados."""
    xlsx_path = save_xlsx_with_disp_table(
        tmp_path, "M_VF",
        rows=[build_full_row("M_VF", UID="MVF_001", Numero=1,
                              plc_tag="V_MVF_001",
                              **{"SA.Byte": 42,
                                 "Cfg.ByteAnalogica": "cfg_ana := 1;"})],
    )
    wb = load_workbook_safe(xlsx_path)
    result = DispM_VFParser().extraer(wb)
    assert len(result) == 1
    d = result[0]
    assert isinstance(d, DispM_VF)
    assert d.uid == "MVF_001"
    assert d.sa_byte == 42
    assert d.cfg_byteanalogica == "cfg_ana := 1;"


def test_hoja_inexistente_devuelve_lista_vacia(tmp_path) -> None:
    xlsx_path = save_xlsx_with_disp_table(
        tmp_path, "M_VF",
        rows=[build_full_row("M_VF")],
        sheet_name="OTRA_HOJA",
    )
    wb = load_workbook_safe(xlsx_path)
    assert DispM_VFParser().extraer(wb) == []


def test_tabla_inexistente_devuelve_lista_vacia(tmp_path) -> None:
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo

    xlsx_path = tmp_path / "otro.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "DISP_M_VF"
    ws.append(["X", "Y"])
    ws.append([1, 2])
    table = Table(displayName="Tabla_Otra", ref="A1:B2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(str(xlsx_path))

    wb2 = load_workbook_safe(xlsx_path)
    assert DispM_VFParser().extraer(wb2) == []


def test_fila_sin_uid_se_descarta(tmp_path) -> None:
    xlsx_path = save_xlsx_with_disp_table(
        tmp_path, "M_VF",
        rows=[
            [None, None, "X", "X", "X"],
            build_full_row("M_VF", UID="MVF_001"),
        ],
    )
    wb = load_workbook_safe(xlsx_path)
    result = DispM_VFParser().extraer(wb)
    assert len(result) == 1
    assert result[0].uid == "MVF_001"


def test_defaults_when_only_uid_and_numero(tmp_path) -> None:
    """Solo UID+Numero → ``sa_byte`` y ``cfg_byteanalogica`` son defaults."""
    xlsx_path = save_xlsx_with_disp_table(
        tmp_path, "M_VF",
        rows=[["MVF_001", 1, "V_MVF_001", "c", "d"]],
        headers=["UID", "Numero", "PLC.Tag", "PLC.Comentario", "Descripcion"],
    )
    wb = load_workbook_safe(xlsx_path)
    result = DispM_VFParser().extraer(wb)
    assert len(result) == 1
    d = result[0]
    assert d.sa_byte == 0
    assert d.cfg_byteanalogica == ""
