"""Tests del ``DispMParser`` (Fase 5 del plan).

Cubre la extracción de ``Tabla_Disp_M`` (hoja ``DISP_M``) y la
construcción de ``DispM`` (motores digitales).
"""
from __future__ import annotations

from areas.alimentacion.domain.models.excel_cache import DispM
from areas.alimentacion.infrastructure.parsers.disp_m import DispMParser

from tests._disp_parser_test_helpers import (
    build_full_row,
    load_workbook_safe,
    save_xlsx_with_disp_table,
)


def test_extrae_disp_m_basico(tmp_path) -> None:
    """1 fila con los campos específicos ``rt_*``/``rm_*`` populados."""
    xlsx_path = save_xlsx_with_disp_table(
        tmp_path, "M",
        rows=[build_full_row("M", UID="M_001", Numero=1,
                              plc_tag="V_M_001",
                              **{"S.Byte": 0, "S.Bit": 0,
                                 "RT.Byte": 1, "RT.Bit": 2,
                                 "RM.Byte": 3, "RM.Bit": 4})],
    )
    wb = load_workbook_safe(xlsx_path)
    result = DispMParser().extraer(wb)
    assert len(result) == 1
    d = result[0]
    assert isinstance(d, DispM)
    assert d.uid == "M_001"
    assert d.s_byte == 0
    assert d.rt_byte == 1
    assert d.rt_bit == 2
    assert d.rm_byte == 3
    assert d.rm_bit == 4


def test_hoja_inexistente_devuelve_lista_vacia(tmp_path) -> None:
    xlsx_path = save_xlsx_with_disp_table(
        tmp_path, "M",
        rows=[build_full_row("M")],
        sheet_name="OTRA_HOJA",
    )
    wb = load_workbook_safe(xlsx_path)
    assert DispMParser().extraer(wb) == []


def test_tabla_inexistente_devuelve_lista_vacia(tmp_path) -> None:
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo

    xlsx_path = tmp_path / "otro.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "DISP_M"
    ws.append(["X", "Y"])
    ws.append([1, 2])
    table = Table(displayName="Tabla_Otra", ref="A1:B2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(str(xlsx_path))

    wb2 = load_workbook_safe(xlsx_path)
    assert DispMParser().extraer(wb2) == []


def test_fila_sin_uid_se_descarta(tmp_path) -> None:
    xlsx_path = save_xlsx_with_disp_table(
        tmp_path, "M",
        rows=[
            [None, None, "X", "X", "X"],
            build_full_row("M", UID="M_001"),
        ],
    )
    wb = load_workbook_safe(xlsx_path)
    result = DispMParser().extraer(wb)
    assert len(result) == 1
    assert result[0].uid == "M_001"


def test_defaults_when_only_uid_and_numero(tmp_path) -> None:
    xlsx_path = save_xlsx_with_disp_table(
        tmp_path, "M",
        rows=[["M_001", 1, "V_M_001", "c", "d"]],
        headers=["UID", "Numero", "PLC.Tag", "PLC.Comentario", "Descripcion"],
    )
    wb = load_workbook_safe(xlsx_path)
    result = DispMParser().extraer(wb)
    assert len(result) == 1
    d = result[0]
    assert d.s_byte == 0
    assert d.rt_byte == 0
    assert d.rm_byte == 0
    # Los ``cfg_*`` son strings vacíos.
    assert d.cfg_byteretornotermico == ""
