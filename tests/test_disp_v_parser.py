"""Tests del ``DispVParser`` (Fase 5 del plan).

Cubre la extracción de ``Tabla_Disp_V`` (hoja ``DISP_V``) y la
construcción de ``DispV``. Los campos específicos (S.Byte/S.Bit,
RR.Byte/RR.Bit, RT.Byte/RT.Bit) se verifican explícitamente.
"""
from __future__ import annotations

from areas.alimentacion.domain.models.excel_cache import DispV
from areas.alimentacion.infrastructure.parsers.disp_v import DispVParser

from tests._disp_parser_test_helpers import (
    build_full_row,
    load_workbook_safe,
    save_xlsx_with_disp_table,
)


def test_extrae_disp_v_basico(tmp_path) -> None:
    """1 fila con los 4 campos específicos ``rr_*``/``rt_*`` populados."""
    xlsx_path = save_xlsx_with_disp_table(
        tmp_path, "V",
        rows=[build_full_row("V", UID="V_001", Numero=1,
                              plc_tag="V_V_001",
                              **{"S.Byte": 0, "S.Bit": 0,
                                 "RR.Byte": 0, "RR.Bit": 1,
                                 "RT.Byte": 2, "RT.Bit": 3})],
    )
    wb = load_workbook_safe(xlsx_path)
    result = DispVParser().extraer(wb)
    assert len(result) == 1
    d = result[0]
    assert isinstance(d, DispV)
    assert d.uid == "V_001"
    assert d.s_byte == 0
    assert d.s_bit == 0
    assert d.rr_byte == 0
    assert d.rr_bit == 1
    assert d.rt_byte == 2
    assert d.rt_bit == 3


def test_hoja_inexistente_devuelve_lista_vacia(tmp_path) -> None:
    xlsx_path = save_xlsx_with_disp_table(
        tmp_path, "V",
        rows=[build_full_row("V")],
        sheet_name="OTRA_HOJA",
    )
    wb = load_workbook_safe(xlsx_path)
    assert DispVParser().extraer(wb) == []


def test_tabla_inexistente_devuelve_lista_vacia(tmp_path) -> None:
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo

    xlsx_path = tmp_path / "otro.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "DISP_V"
    ws.append(["X", "Y"])
    ws.append([1, 2])
    table = Table(displayName="Tabla_Otra", ref="A1:B2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(str(xlsx_path))

    wb2 = load_workbook_safe(xlsx_path)
    assert DispVParser().extraer(wb2) == []


def test_fila_sin_uid_se_descarta(tmp_path) -> None:
    xlsx_path = save_xlsx_with_disp_table(
        tmp_path, "V",
        rows=[
            [None, None, "X", "X", "X"],
            build_full_row("V", UID="V_001"),
        ],
    )
    wb = load_workbook_safe(xlsx_path)
    result = DispVParser().extraer(wb)
    assert len(result) == 1
    assert result[0].uid == "V_001"


def test_defaults_when_only_uid_and_numero(tmp_path) -> None:
    xlsx_path = save_xlsx_with_disp_table(
        tmp_path, "V",
        rows=[["V_001", 1, "V_V_001", "c", "d"]],
        headers=["UID", "Numero", "PLC.Tag", "PLC.Comentario", "Descripcion"],
    )
    wb = load_workbook_safe(xlsx_path)
    result = DispVParser().extraer(wb)
    assert len(result) == 1
    d = result[0]
    assert d.s_byte == 0
    assert d.rr_byte == 0
    assert d.rt_byte == 0
    # Los ``cfg_*`` son strings vacíos por default.
    assert d.cfg_byteretornoreposo == ""
    assert d.cfg_byteretornotrabajo == ""
