"""Tests del parser ``ProcesosParser``.

Cubre la extracción de ``Tabla_Procesos`` (hoja ``CONFIGURACION``) y
la fidelidad de los 8 campos del DTO ``ProcesoPLC`` con el Excel
corporativo.

Convenciones:
    * Excel sintético construido en ``tmp_path`` con ``Table`` real
      (R5 del plan ``_plan/04_excel_cache_phased_plan.md``).
    * Se carga con ``load_workbook`` para garantizar que la
      ``ListObject`` se registre en ``worksheet.tables``.
"""
from __future__ import annotations

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from areas.alimentacion.domain.models.excel_cache import ProcesoPLC
from areas.alimentacion.infrastructure.parsers.proc_procesos import ProcesosParser


# ── Helpers de construcción de Excels sintéticos ────────────────────────


def _save_xlsx_with_procesos_table(
    tmp_path,
    rows: list[list] | None,
    *,
    sheet_name: str = "CONFIGURACION",
    table_name: str = "Tabla_Procesos",
    headers: list[str] | None = None,
) -> str:
    """Crea un .xlsx sintético con la tabla de procesos.

    Args:
        tmp_path: fixture pytest de path temporal.
        rows: lista de filas de datos (``None`` o ``[]`` → solo cabeceras,
            útil para verificar que la tabla existe pero está vacía).
        sheet_name: nombre de la hoja (default ``CONFIGURACION``).
        table_name: nombre de la ``ListObject`` (default
            ``Tabla_Procesos``).
        headers: cabeceras (default: las del Excel legacy del corporativo).

    Returns:
        Path absoluto del .xlsx en str.
    """
    if headers is None:
        headers = [
            "UID",
            "Nombre",
            "Codigo",
            "PReal",
            "Index_Preal",
            "PInt",
            "Index_Pint",
            "Alarmas",
        ]
    if rows is None:
        rows = []

    xlsx_path = tmp_path / f"{table_name}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)

    # Registrar la Table (R5 del plan: sin esto, ``worksheet.tables``
    # no contiene la ``ListObject`` y el parser no la encontraría).
    last_col_letter = chr(ord("A") + len(headers) - 1)
    last_row = 1 + len(rows)
    if last_row == 1:
        # Tabla de solo cabeceras (1 fila). Aún así la registramos.
        ref = f"A1:{last_col_letter}{last_row}"
    else:
        ref = f"A1:{last_col_letter}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(str(xlsx_path))
    return str(xlsx_path)


def _load(path: str) -> Workbook:
    """Carga un .xlsx desde disco preservando ``worksheet.tables``."""
    from openpyxl import load_workbook

    return load_workbook(path)


# ── Tests del parser ─────────────────────────────────────────────────────


def test_extrae_procesos_basico(tmp_path) -> None:
    """Excel con 1 fila válida → DTO con campos correctos."""
    xlsx_path = _save_xlsx_with_procesos_table(
        tmp_path,
        rows=[
            [
                1,
                "Proceso Uno",
                "PR1",
                10,
                0,
                5,
                0,
                32,
            ],
        ],
    )
    wb = _load(xlsx_path)

    result = ProcesosParser().extraer(wb)

    assert len(result) == 1
    p = result[0]
    assert isinstance(p, ProcesoPLC)
    assert p.uid == 1
    assert p.nombre == "Proceso Uno"
    assert p.codigo == "PR1"
    assert p.preal == 10
    assert p.index_preal == 0
    assert p.pint == 5
    assert p.index_pint == 0
    assert p.alarmas == 32


def test_hoja_inexistente_devuelve_lista_vacia(tmp_path) -> None:
    """Sheet ``CONFIGURACION`` no existe → ``[]`` (no lanza)."""
    xlsx_path = _save_xlsx_with_procesos_table(
        tmp_path,
        rows=[[1, "P", "P", 0, 0, 0, 0, 0]],
        sheet_name="OTRA_HOJA",
    )
    wb = _load(xlsx_path)

    result = ProcesosParser().extraer(wb)

    assert result == []


def test_tabla_inexistente_devuelve_lista_vacia(tmp_path) -> None:
    """Sheet existe pero ``Tabla_Procesos`` no → ``[]`` (no lanza)."""
    # Construimos un .xlsx con OTRA tabla en CONFIGURACION.
    xlsx_path = tmp_path / "otro.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "CONFIGURACION"
    ws.append(["X", "Y"])
    ws.append([1, 2])
    table = Table(displayName="Tabla_Otra", ref="A1:B2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(str(xlsx_path))

    wb2 = _load(xlsx_path)
    result = ProcesosParser().extraer(wb2)
    assert result == []


@pytest.mark.parametrize(
    ("uid", "nombre", "codigo", "preal", "index_preal",
     "pint", "index_pint", "alarmas"),
    [
        # (uid, nombre, codigo, preal, index_preal, pint, index_pint, alarmas)
        # Cubre casos borde: uid=0, sin alarmas, alarmas altas. La
        # verificación es de fidelidad campo-a-campo contra el Excel
        # sintético (no se derivan nombres de DB en el DTO).
        (0, "Sin alarmas", "PR0", 0, 0, 0, 0, 0),    # uid=0, sin alarmas
        (1, "Una alarma",  "PR1", 5, 0, 2, 0, 1),    # caso típico
        (1, "Borde 16",    "PR2", 5, 0, 2, 0, 16),   # 16 alarmas
        (1, "Borde 17",    "PR3", 5, 0, 2, 0, 17),   # 17 alarmas
        (1, "Borde 32",    "PR4", 5, 0, 2, 0, 32),   # 32 alarmas
        (1, "Cien alarm",  "PR5", 5, 0, 2, 0, 100),  # alarmas grandes
    ],
)
def test_proceso_plc_8_campos_del_excel(
    tmp_path,
    uid: int,
    nombre: str,
    codigo: str,
    preal: int,
    index_preal: int,
    pint: int,
    index_pint: int,
    alarmas: int,
) -> None:
    """El DTO ``ProcesoPLC`` expone los 8 campos del Excel sin propiedades derivadas.

    Garantiza fidelidad campo-a-campo entre la fila de la tabla
    ``Tabla_Procesos`` y el DTO. Los nombres de DB y otros valores
    derivados NO viven en el DTO: se computan en el consumidor
    (frontend o backend futuro).
    """
    xlsx_path = _save_xlsx_with_procesos_table(
        tmp_path,
        rows=[
            [uid, nombre, codigo, preal, index_preal, pint, index_pint, alarmas],
        ],
    )
    wb = _load(xlsx_path)

    result = ProcesosParser().extraer(wb)
    assert len(result) == 1
    p = result[0]

    # 8 campos del Excel, leídos en el mismo orden que la tabla.
    assert p.uid == uid
    assert p.nombre == nombre
    assert p.codigo == codigo
    assert p.preal == preal
    assert p.index_preal == index_preal
    assert p.pint == pint
    assert p.index_pint == index_pint
    assert p.alarmas == alarmas


def test_fila_sin_uid_se_descarta(tmp_path) -> None:
    """Fila con UID vacío (None) NO aparece en el resultado.

    Razón: el DTO exige ``uid: int`` y ``_safe_int(None) == 0``. Si
    aceptáramos la fila, contaminaría el cache con un proceso ``uid=0``
    falso. Política: descartar (WARNING implícito del builder).
    """
    xlsx_path = _save_xlsx_with_procesos_table(
        tmp_path,
        rows=[
            [None, "Sin UID", "X", 0, 0, 0, 0, 0],  # UID vacío → se descarta
            [1, "Con UID", "PR1", 0, 0, 0, 0, 0],  # válida
        ],
    )
    wb = _load(xlsx_path)

    result = ProcesosParser().extraer(wb)

    # Solo la fila con UID=1 se queda.
    assert len(result) == 1
    assert result[0].uid == 1
    assert result[0].codigo == "PR1"
