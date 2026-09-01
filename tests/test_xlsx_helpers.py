"""Tests del helper compartido ``_xlsx_helpers.py``.

Cubre los cast defensivos (``_safe_str`` / ``_safe_int`` / ``_safe_float``)
y el lector de ``ListObject`` (``extract_list_object_rows``) sobre un
Excel sintético con ``Table`` real añadida vía ``worksheet.add_table``
(riesgo R5 del plan ``_plan/04_excel_cache_phased_plan.md``: las
``ListObject`` solo se pueblan en ``worksheet.tables`` si se usa la
API oficial de openpyxl — escribir filas a mano no basta).
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from areas.alimentacion.infrastructure.parsers._xlsx_helpers import (
    _safe_float,
    _safe_int,
    _safe_str,
    extract_list_object_rows,
)


# ── _safe_str ────────────────────────────────────────────────────────────


def test_safe_str_handles_none_empty_nan() -> None:
    """None, '', 'nan', 'None', 'null' y whitespace → ''."""
    assert _safe_str(None) == ""
    assert _safe_str("") == ""
    assert _safe_str("nan") == ""
    assert _safe_str("NaN") == ""  # case-insensitive
    assert _safe_str("None") == ""
    assert _safe_str("null") == ""
    assert _safe_str("  ") == ""  # whitespace puro


def test_safe_str_preserves_text() -> None:
    """Texto no vacío (incluso con espacios alrededor) se preserva trimmed."""
    assert _safe_str("Proceso 1") == "Proceso 1"
    assert _safe_str("  Hola  ") == "Hola"
    # Distingue "nada" (texto real) de "nan"/"None"/"null" (placeholders).
    assert _safe_str("nada") == "nada"


# ── _safe_int ────────────────────────────────────────────────────────────


def test_safe_int_parses_strings_and_floats() -> None:
    """Numéricos (int/float/str-numérico) se castean; resto → 0."""
    # Formas válidas
    assert _safe_int(5) == 5
    assert _safe_int(5.0) == 5
    assert _safe_int("5") == 5
    assert _safe_int("5.0") == 5
    assert _safe_int("  7  ") == 7  # strip de whitespace
    # bool se trata como int
    assert _safe_int(True) == 1
    assert _safe_int(False) == 0
    # Casos inválidos → default 0
    assert _safe_int(None) == 0
    assert _safe_int("") == 0
    assert _safe_int("nan") == 0
    assert _safe_int("Pendiente") == 0
    # Custom default
    assert _safe_int("Pendiente", default=99) == 99


# ── _safe_float ──────────────────────────────────────────────────────────


def test_safe_float_accepts_comma_decimal() -> None:
    """Coma decimal (formato europeo) se acepta."""
    assert _safe_float("1,5") == 1.5
    assert _safe_float("  0,25  ") == 0.25
    # Punto decimal estándar también funciona
    assert _safe_float("1.5") == 1.5
    # Numéricos nativos
    assert _safe_float(2) == 2.0
    assert _safe_float(2.5) == 2.5
    # bool
    assert _safe_float(True) == 1.0
    # Casos inválidos → default 0.0
    assert _safe_float(None) == 0.0
    assert _safe_float("") == 0.0
    assert _safe_float("texto") == 0.0
    # Custom default
    assert _safe_float("texto", default=-1.0) == -1.0


# ── extract_list_object_rows ─────────────────────────────────────────────


def _make_xlsx_with_table(
    tmp_path,
    sheet_name: str,
    table_name: str,
    headers: list[str],
    rows: list[list],
) -> str:
    """Crea un .xlsx sintético con una ``Table`` real (R5 del plan)."""
    xlsx_path = tmp_path / f"{table_name}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    # Cabeceras
    ws.append(headers)
    # Filas
    for row in rows:
        ws.append(row)
    # Registrar la ListObject con un ref explícito que cubra cabeceras + filas.
    last_col_letter = chr(ord("A") + len(headers) - 1)
    last_row = 1 + len(rows)
    ref = f"A1:{last_col_letter}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(str(xlsx_path))
    return str(xlsx_path)


def test_extract_list_object_rows_returns_dicts(tmp_path) -> None:
    """Crea un .xlsx con Table real, lee y verifica shape de dicts."""
    xlsx_path = _make_xlsx_with_table(
        tmp_path,
        sheet_name="CONFIGURACION",
        table_name="Tabla_Procesos",
        headers=["UID", "Nombre", "Codigo"],
        rows=[
            [1, "Proceso Uno", "PR1"],
            [2, "Proceso Dos", "PR2"],
        ],
    )
    wb = Workbook()  # se ignora; cargamos desde disco para garantizar Table
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)

    rows = extract_list_object_rows(wb, "CONFIGURACION", "Tabla_Procesos")
    assert len(rows) == 2
    assert rows[0] == {"UID": 1, "Nombre": "Proceso Uno", "Codigo": "PR1"}
    assert rows[1] == {"UID": 2, "Nombre": "Proceso Dos", "Codigo": "PR2"}


def test_extract_list_object_rows_sheet_missing_returns_empty(tmp_path) -> None:
    """Si la hoja no existe, devuelve [] (no lanza)."""
    xlsx_path = _make_xlsx_with_table(
        tmp_path,
        sheet_name="OTRA",
        table_name="Tabla_Algo",
        headers=["X"],
        rows=[[1]],
    )
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    assert extract_list_object_rows(wb, "NO_EXISTE", "Tabla_Algo") == []


def test_extract_list_object_rows_table_missing_returns_empty(tmp_path) -> None:
    """Si la tabla no existe en la hoja, devuelve [] (no lanza)."""
    xlsx_path = _make_xlsx_with_table(
        tmp_path,
        sheet_name="CONFIGURACION",
        table_name="Tabla_Procesos",
        headers=["UID"],
        rows=[[1]],
    )
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    assert (
        extract_list_object_rows(wb, "CONFIGURACION", "Tabla_No_Existe") == []
    )


def test_extract_list_object_rows_skips_empty_rows(tmp_path) -> None:
    """Filas completamente vacías se descartan."""
    xlsx_path = _make_xlsx_with_table(
        tmp_path,
        sheet_name="CONFIGURACION",
        table_name="Tabla_Procesos",
        headers=["UID", "Nombre"],
        rows=[
            [1, "Proceso Uno"],
            [None, None],  # fila vacía → se descarta
            [2, "Proceso Dos"],
        ],
    )
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    rows = extract_list_object_rows(wb, "CONFIGURACION", "Tabla_Procesos")
    assert len(rows) == 2
    assert rows[0]["UID"] == 1
    assert rows[1]["UID"] == 2
