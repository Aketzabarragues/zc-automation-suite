"""Tests del endpoint ``POST /api/v1/excel/upload`` wired al cache.

Cubre:
  * El endpoint popula ``state.excel_cache`` y ``state.excel_path``.
  * El response tiene la shape legacy (summary, total_dispositivos,
    dimensiones).
  * Back-compat con la SPA: ``state.dispositivos_<hw>`` se popula
    desde el cache.

Patrón: ``TestClient`` + Excel sintético en ``tmp_path``.
"""
from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path
from typing import Any
from unittest.mock import MagicMock

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from core.application.state import AppState, get_app_state
from core.infrastructure.config_manager import ConfigManager
from core.infrastructure.gateway import TIAProcessGateway
from interfaces.web_server.app import create_app


# ── Configuración JSON fixture ─────────────────────────────────────────


_FULL_CONFIG: dict[str, Any] = {
    "departments": {
        "alimentacion": {
            "global_config_table_name": "000_Config_Dispositivos",
            "tia_folders": {
                "proceso":      "003_Procesos",
                "dispositivos": "2000_Dispositivos",
                "nmax":         "000_Sistema",
            },
            "n_max_catalog": [
                {"name": "N_MAX_DISP_ED",   "value": 10},
                {"name": "N_MAX_DISP_EA",   "value": 10},
                {"name": "N_MAX_DISP_SA",   "value": 10},
                {"name": "N_MAX_DISP_V",    "value": 10},
                {"name": "N_MAX_DISP_M",    "value": 10},
                {"name": "N_MAX_DISP_M_VF", "value": 10},
            ],
            "Dispositivos": {
                "ed": {
                    "db_name": "DB2000_ED", "db_array_name": "ED",
                    "tag_table": "2000_Disp_ED",
                    "config_table": "000_Config_Dispositivos",
                },
                "ea": {
                    "db_name": "DB2001_EA", "db_array_name": "EA",
                    "tag_table": "2000_Disp_EA",
                    "config_table": "000_Config_Dispositivos",
                },
                "sa": {
                    "db_name": "DB2006_SA", "db_array_name": "SA",
                    "tag_table": "2000_Disp_SA",
                    "config_table": "000_Config_Dispositivos",
                },
                "v": {
                    "db_name": "DB2010_V", "db_array_name": "V",
                    "tag_table": "2000_Disp_V",
                    "config_table": "000_Config_Dispositivos",
                },
                "m": {
                    "db_name": "DB2015_M", "db_array_name": "M",
                    "tag_table": "2000_Disp_M",
                    "config_table": "000_Config_Dispositivos",
                },
                "m_vf": {
                    "db_name": "DB2016_M_VF", "db_array_name": "M_VF",
                    "tag_table": "2000_Disp_M_VF",
                    "config_table": "000_Config_Dispositivos",
                },
            },
        }
    }
}


def _write_config(tmp_path: Path) -> Path:
    p = tmp_path / "config.json"
    p.write_text(json.dumps(_FULL_CONFIG), encoding="utf-8")
    return p


# ── Excel sintético ────────────────────────────────────────────────────


def _add_table(
    wb: Workbook, sheet_name: str, table_name: str,
    headers: list[str], rows: list[list],
) -> None:
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    last_col_letter = chr(ord("A") + len(headers) - 1)
    last_row = 1 + len(rows)
    ref = f"A1:{last_col_letter}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)


def _build_minimal_xlsx_bytes() -> bytes:
    """Genera un xlsx en memoria con 1 fila de cada uno de los 6 tipos."""
    wb = Workbook()
    wb.remove(wb.active)
    tables = [
        ("DISP_ED",   "Tabla_Disp_ED",   "ED_001",   1, "V_ED_001"),
        ("DISP_EA",   "Tabla_Disp_EA",   "EA_001",   1, "V_EA_001"),
        ("DISP_SA",   "Tabla_Disp_SA",   "SA_001",   1, "V_SA_001"),
        ("DISP_V",    "Tabla_Disp_V",    "V_001",    1, "V_V_001"),
        ("DISP_M",    "Tabla_Disp_M",    "M_001",    1, "V_M_001"),
        ("DISP_M_VF", "Tabla_Disp_M_VF", "MVF_001",  1, "V_MVF_001"),
    ]
    headers = ["UID", "Numero", "PLC.Tag", "Descripcion"]
    for sheet_name, table_name, uid, numero, tag in tables:
        _add_table(wb, sheet_name, table_name, headers,
                   [[uid, numero, tag, f"Desc {uid}"]])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Fixture: TestClient con dependencias inyectadas ─────────────────────


@pytest.fixture
def app_with_overrides(tmp_path: Path, monkeypatch):
    """App FastAPI con ``ConfigManager`` apuntando al fixture JSON."""
    config_path = _write_config(tmp_path)
    cm = ConfigManager(config_path=config_path)

    # Mockear el gateway (no se usa en el endpoint upload, pero
    # ``create_app`` lo exige).
    gateway = MagicMock(spec=TIAProcessGateway)
    app = create_app(gateway)
    # Override del config_manager con el fixture JSON.
    app.state.config_manager = cm

    # Reset de los campos del AppState.
    state = get_app_state()
    state.reset()
    state.excel_cache = None
    state.excel_path = None
    state.dimensiones = None

    yield app, state
    # Reset al final para que tests posteriores vean estado limpio.
    state.reset()
    state.excel_cache = None
    state.excel_path = None
    state.dimensiones = None


# ── Tests ──────────────────────────────────────────────────────────────


def test_upload_popula_cache(app_with_overrides) -> None:
    """POST /api/v1/excel/upload popula ``state.excel_cache``."""
    app, state = app_with_overrides
    # Sanity: al inicio el cache está vacío.
    assert state.excel_cache is None
    assert state.excel_path is None

    xlsx_bytes = _build_minimal_xlsx_bytes()
    with TestClient(app) as client:
        resp = client.post(
            "/api/v1/excel/upload",
            files={"file": ("test.xlsx", xlsx_bytes, "application/octet-stream")},
        )
    assert resp.status_code == 200
    # Estado populado. ``excel_path`` apunta al tempfile que el
    # endpoint crea (con prefijo ``zcupload_`` y sufijo ``.xlsx``).
    assert state.excel_cache is not None
    assert state.excel_path is not None
    assert state.excel_path.endswith(".xlsx")
    assert "zcupload_" in state.excel_path


def test_upload_response_shape_legacy(app_with_overrides) -> None:
    """Response tiene ``summary``, ``total_dispositivos``, ``dimensiones``."""
    app, state = app_with_overrides
    xlsx_bytes = _build_minimal_xlsx_bytes()
    with TestClient(app) as client:
        resp = client.post(
            "/api/v1/excel/upload",
            files={"file": ("test.xlsx", xlsx_bytes, "application/octet-stream")},
        )
    body = resp.json()
    assert body["ok"] is True
    assert "summary" in body
    assert "total_dispositivos" in body
    assert "dimensiones" in body
    # 6 dispositivos (1 por tipo).
    assert body["total_dispositivos"] == 6
    # ``dimensiones`` tiene los 6 canónicos.
    assert set(body["dimensiones"].keys()) == {
        "num_disp_ed", "num_disp_ea", "num_disp_sa",
        "num_disp_v", "num_disp_m", "num_disp_m_vf",
    }


def test_upload_back_compat_state_dispositivos_ed(app_with_overrides) -> None:
    """``state.dispositivos_ed`` se popula desde el cache (back-compat)."""
    app, state = app_with_overrides
    xlsx_bytes = _build_minimal_xlsx_bytes()
    with TestClient(app) as client:
        resp = client.post(
            "/api/v1/excel/upload",
            files={"file": ("test.xlsx", xlsx_bytes, "application/octet-stream")},
        )
    assert resp.status_code == 200
    # Los 6 atributos legacy están populados.
    for hw in ("ed", "ea", "sa", "v", "m", "m_vf"):
        devices = state.get_devices(hw)
        assert len(devices) == 1, (
            f"Tipo {hw!r} esperaba 1 dispositivo, obtuve {len(devices)}"
        )
    # ``state.dimensiones`` también está populado.
    assert state.dimensiones is not None
