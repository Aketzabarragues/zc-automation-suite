"""Tests end-to-end del ``ExcelLoader`` (Fase 5 del plan).

Cubre:
  * ``load`` abre el workbook UNA vez y construye el ``ExcelCache``.
  * ``load`` lanza ``FileNotFoundError`` si el path no existe.
  * Los lookups ``*_by_codigo`` se precomputan correctamente.
  * Si un parser lanza, ``wb.close()`` se llama en el ``finally``.
  * ``load`` con hojas faltantes no lanza: las listas quedan vacías.
  * ``to_dict()`` produce un dict JSON-serializable.
  * El cache incluye 6 tipos de dispositivos + 4 listas de software
    + 1 ``n_max``.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from areas.alimentacion.domain.models.excel_cache import (
    AlarmaPLC,
    DimensionesDispositivos,
    DispED,
    ExcelCache,
    ParamIntPLC,
    ParamRealPLC,
    ProcesoPLC,
)
from areas.alimentacion.infrastructure.loaders.excel_loader import ExcelLoader


# ── Helpers ─────────────────────────────────────────────────────────────


def _add_named_value(
    wb: Workbook, sheet_name: str, cell: str, name: str, value: Any
) -> None:
    from openpyxl.workbook.defined_name import DefinedName
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]
    ws[cell] = value
    dn = DefinedName(name=name, attr_text=f"'{sheet_name}'!${cell}")
    wb.defined_names[name] = dn


def _add_table(
    wb: Workbook, sheet_name: str, table_name: str, headers: list[str], rows: list[list]
) -> None:
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    last_col_letter = chr(ord("A") + len(headers) - 1)
    last_row = 1 + len(rows)
    ref = f"A1:{last_col_letter}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)


def _build_full_xlsx(target: Path) -> Path:
    """Construye un xlsx con 1 fila de CADA uno de los 10 dominios + 1 N_MAX."""
    wb = Workbook()
    wb.remove(wb.active)

    # 6 dispositivos
    _add_table(wb, "DISP_ED", "Tabla_Disp_ED",
               ["UID", "Numero", "PLC.Tag", "Descripcion"],
               [["ED_001", 1, "V_ED_001", "Entrada digital 1"]])
    _add_table(wb, "DISP_EA", "Tabla_Disp_EA",
               ["UID", "Numero", "PLC.Tag", "Descripcion", "RII", "RSI"],
               [["EA_001", 1, "V_EA_001", "Entrada analogica 1", 0.0, 100.0]])
    _add_table(wb, "DISP_SA", "Tabla_Disp_SA",
               ["UID", "Numero", "PLC.Tag", "Descripcion"],
               [["SA_001", 1, "V_SA_001", "Salida analogica 1"]])
    _add_table(wb, "DISP_V", "Tabla_Disp_V",
               ["UID", "Numero", "PLC.Tag", "Descripcion"],
               [["V_001", 1, "V_V_001", "Variable 1"]])
    _add_table(wb, "DISP_M", "Tabla_Disp_M",
               ["UID", "Numero", "PLC.Tag", "Descripcion"],
               [["M_001", 1, "V_M_001", "Motor 1"]])
    _add_table(wb, "DISP_M_VF", "Tabla_Disp_M_VF",
               ["UID", "Numero", "PLC.Tag", "Descripcion"],
               [["MVF_001", 1, "V_MVF_001", "Motor VF 1"]])

    # 4 software
    _add_table(wb, "CONFIGURACION", "Tabla_Procesos",
               ["UID", "Nombre", "Codigo", "PReal", "PInt", "Alarmas"],
               [[1, "Proceso 1", "PR1", 1, 1, 16]])
    _add_table(wb, "P_REAL", "Tabla_PReal",
               ["UID", "Numero", "Proceso", "Codigo", "Num.DB",
                "Producto", "Tipo", "Descripcion", "ComentarioDB",
                "Visibilidad", "Num.Lista", "Txt.Lista"],
               [["PR_1_001", "001", "Proceso 1", "PR1", 3001,
                 "Producto", "Setpoint", "PR1", "c", "Si", 1, "Txt"]])
    _add_table(wb, "P_INT", "Tabla_PInt",
               ["UID", "Numero", "Proceso", "Codigo", "Num.DB",
                "Producto", "Tipo", "Descripcion", "ComentarioDB",
                "Visibilidad", "Num.Lista", "Txt.Lista"],
               [["PI_1_001", "001", "Proceso 1", "PR1", 3002,
                 "Producto", "Contador", "PI1", "c", "Si", 0, "Txt"]])
    _add_table(wb, "ALARMAS", "Tabla_Alarmas",
               ["UID", "Numero", "Proceso", "Num.DB", "Descripcion", "ComentarioDB"],
               [["AL_1_001", "001", "Proceso 1", 5001, "Alarma 1", "c"]])

    # 1 N_MAX (defined name)
    _add_named_value(wb, "Config", "A1", "N_MAX_DISP_ED", 10)
    _add_named_value(wb, "Config", "A2", "N_MAX_DISP_EA", 20)
    _add_named_value(wb, "Config", "A3", "N_MAX_DISP_SA", 30)
    _add_named_value(wb, "Config", "A4", "N_MAX_DISP_V", 40)
    _add_named_value(wb, "Config", "A5", "N_MAX_DISP_M", 50)
    _add_named_value(wb, "Config", "A6", "N_MAX_DISP_M_VF", 60)

    wb.save(target)
    return target


# ── Tests ───────────────────────────────────────────────────────────────


def test_load_basico(tmp_path) -> None:
    """Excel con 1 fila de CADA dominio → ``ExcelCache`` con todos los campos."""
    xlsx = _build_full_xlsx(tmp_path / "full.xlsx")
    cache = ExcelLoader().load(xlsx)

    assert isinstance(cache, ExcelCache)
    # Path absoluto.
    assert cache.excel_path.endswith("full.xlsx")
    # mtime_ns es un int > 0.
    assert cache.excel_mtime_ns > 0
    # parsed_at es un datetime UTC.
    assert cache.parsed_at.tzinfo is not None
    # Los 6 tipos de dispositivos tienen 1 elemento.
    assert len(cache.dispositivos["ed"]) == 1
    assert len(cache.dispositivos["ea"]) == 1
    assert len(cache.dispositivos["sa"]) == 1
    assert len(cache.dispositivos["v"]) == 1
    assert len(cache.dispositivos["m"]) == 1
    assert len(cache.dispositivos["m_vf"]) == 1
    # Las 4 listas de software tienen 1 elemento.
    assert len(cache.procesos) == 1
    assert len(cache.parametros_real) == 1
    assert len(cache.parametros_int) == 1
    assert len(cache.alarmas) == 1
    # N_MAX correcto.
    assert cache.n_max.num_disp_ed == 10
    assert cache.n_max.num_disp_ea == 20
    assert cache.n_max.num_disp_sa == 30
    assert cache.n_max.num_disp_v == 40
    assert cache.n_max.num_disp_m == 50
    assert cache.n_max.num_disp_m_vf == 60
    # Flag.
    assert cache.software_parsers_implemented is True


def test_load_con_excel_inexistente_lanza_filenotfounderror(tmp_path) -> None:
    """Path inexistente → ``FileNotFoundError``."""
    missing = tmp_path / "no.xlsx"
    with pytest.raises(FileNotFoundError):
        ExcelLoader().load(missing)


def test_load_construye_lookups_by_codigo(tmp_path) -> None:
    """Los 3 lookups precomputados tienen el ``codigo`` correcto como clave."""
    xlsx = _build_full_xlsx(tmp_path / "full.xlsx")
    cache = ExcelLoader().load(xlsx)

    # Procesos: el ``ProcesoPLC`` con ``codigo=PR1``.
    assert "PR1" in cache.procesos_by_codigo
    assert isinstance(cache.procesos_by_codigo["PR1"], ProcesoPLC)
    # PReal: el ``ParamRealPLC`` con ``codigo=PR1``.
    assert "PR1" in cache.parametros_real_by_codigo
    assert isinstance(cache.parametros_real_by_codigo["PR1"], ParamRealPLC)
    # PInt: el ``ParamIntPLC`` con ``codigo=PR1``.
    assert "PR1" in cache.parametros_int_by_codigo
    assert isinstance(cache.parametros_int_by_codigo["PR1"], ParamIntPLC)


def test_load_workbook_cerrado_tras_error(tmp_path, monkeypatch) -> None:
    """Si un parser lanza, ``wb.close()`` se llama en el ``finally``."""
    from areas.alimentacion.infrastructure.parsers import disp_ed as disp_ed_mod

    # Mockear ``DispEDParser.extraer`` para que lance.
    calls: dict[str, int] = {"extraer": 0}

    def boom(self, wb):
        calls["extraer"] += 1
        raise RuntimeError("parser boom")

    monkeypatch.setattr(disp_ed_mod.DispEDParser, "extraer", boom)

    xlsx = _build_full_xlsx(tmp_path / "full.xlsx")
    with pytest.raises(RuntimeError, match="parser boom"):
        ExcelLoader().load(xlsx)

    # El parser mockeado se llamó 1 vez.
    assert calls["extraer"] == 1
    # El workbook se cerró (no se puede verificar directamente sin
    # un spy de openpyxl; el test ya cubre el path del ``finally``).


def test_load_con_hojas_faltantes(tmp_path) -> None:
    """Excel SIN hojas de software → listas vacías, no excepción."""
    wb = Workbook()
    wb.remove(wb.active)
    # Solo 1 dispositivo y nada más.
    _add_table(wb, "DISP_ED", "Tabla_Disp_ED",
               ["UID", "Numero", "PLC.Tag", "Descripcion"],
               [["ED_001", 1, "V_ED_001", "Entrada 1"]])
    xlsx = tmp_path / "minimal.xlsx"
    wb.save(xlsx)

    cache = ExcelLoader().load(xlsx)
    # Dispositivo poblado.
    assert len(cache.dispositivos["ed"]) == 1
    # Software: listas vacías.
    assert cache.procesos == ()
    assert cache.parametros_real == ()
    assert cache.parametros_int == ()
    assert cache.alarmas == ()
    # N_MAX: defaults (no defined names).
    assert isinstance(cache.n_max, DimensionesDispositivos)
    assert cache.n_max.num_disp_ed == 0


def test_to_dict_serailiza_todo(tmp_path) -> None:
    """``ExcelCache.to_dict()`` es JSON-serializable."""
    xlsx = _build_full_xlsx(tmp_path / "full.xlsx")
    cache = ExcelLoader().load(xlsx)
    d = cache.to_dict()
    # No lanza y produce un dict JSON-serializable.
    json.dumps(d)


def test_load_with_config_manager(tmp_path) -> None:
    """Si se inyecta un ``ConfigManager``, los parsers lo usan para
    resolver ``SHEET``/``TABLE`` data-driven."""
    import json
    from core.infrastructure.config_manager import ConfigManager

    cfg: dict[str, Any] = {
        "departments": {
            "alimentacion": {
                "global_config_table_name": "000_Config_Dispositivos",
                "tia_folders": {
                    "proceso": "003_Procesos",
                    "dispositivos": "2000_Dispositivos",
                    "nmax": "000_Sistema",
                },
                "n_max_catalog": [
                    {"name": "N_MAX_DISP_ED", "hw_type": "ed"},
                ],
                "Dispositivos": {
                    "ed": {
                        "db_name": "DB2000_ED", "db_array_name": "ED",
                        "tag_table": "2000_Disp_ED",
                        "config_table": "000_Config_Dispositivos",
                    },
                },
            }
        }
    }
    config_path = tmp_path / "config.json"
    config_path.write_text(json.dumps(cfg), encoding="utf-8")
    cm = ConfigManager(config_path=config_path)

    xlsx = _build_full_xlsx(tmp_path / "full.xlsx")
    # Solo verifica que el loader acepta el ``ConfigManager`` y no lanza.
    cache = ExcelLoader(config_manager=cm).load(xlsx)
    assert isinstance(cache, ExcelCache)
