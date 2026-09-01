"""Tests del ``DispEAParser`` (Fase 5 del plan).

Cubre la extracción de ``Tabla_Disp_EA`` (hoja ``DISP_EA``).
"""
from __future__ import annotations

from areas.alimentacion.domain.models.excel_cache import DispEA
from areas.alimentacion.infrastructure.parsers.disp_ea import DispEAParser

from tests._disp_parser_test_helpers import (
    build_full_row,
    load_workbook_safe,
    save_xlsx_with_disp_table,
)


def test_extrae_disp_ea_basico(tmp_path) -> None:
    """Excel con 1 fila válida → ``DispEA`` con ``rii``/``rsi`` float."""
    xlsx_path = save_xlsx_with_disp_table(
        tmp_path, "EA",
        rows=[build_full_row("EA", UID="EA_001", Numero=1,
                              plc_tag="V_EA_001",
                              UNIDADES="BAR", RII=1.5, RSI=10.0)],
    )
    wb = load_workbook_safe(xlsx_path)
    result = DispEAParser().extraer(wb)
    assert len(result) == 1
    d = result[0]
    assert isinstance(d, DispEA)
    assert d.uid == "EA_001"
    assert d.plc_tag == "V_EA_001"
    assert d.unidades == "BAR"
    assert d.rii == 1.5
    assert d.rsi == 10.0
    assert isinstance(d.rii, float)
    assert isinstance(d.rsi, float)


def test_unidades_unidades_lowercase_accepted(tmp_path) -> None:
    """Acepta tanto ``UNIDADES`` (legacy) como ``Unidades`` por compat."""
    # Cabecera con ``Unidades`` minúscula.
    headers = save_xlsx_with_disp_table(tmp_path, "EA", rows=None)
    # Re-crear con header alternativo
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo

    xlsx_path = tmp_path / "alt.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "DISP_EA"
    headers_alt = [
        "UID", "Numero", "PLC.Tag", "Unidades", "RII", "RSI",
        "Descripcion",
    ]
    ws.append(headers_alt)
    ws.append(["EA_001", 1, "V_EA_001", "KG", 2.0, 20.0, "d"])
    last_col_letter = chr(ord("A") + len(headers_alt) - 1)
    ref = f"A1:{last_col_letter}2"
    table = Table(displayName="Tabla_Disp_EA", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(str(xlsx_path))

    wb2 = load_workbook_safe(xlsx_path)
    result = DispEAParser().extraer(wb2)
    assert len(result) == 1
    assert result[0].unidades == "KG"


def test_hoja_inexistente_devuelve_lista_vacia(tmp_path) -> None:
    xlsx_path = save_xlsx_with_disp_table(
        tmp_path, "EA",
        rows=[build_full_row("EA")],
        sheet_name="OTRA_HOJA",
    )
    wb = load_workbook_safe(xlsx_path)
    assert DispEAParser().extraer(wb) == []


def test_tabla_inexistente_devuelve_lista_vacia(tmp_path) -> None:
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo

    xlsx_path = tmp_path / "otro.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "DISP_EA"
    ws.append(["X", "Y"])
    ws.append([1, 2])
    table = Table(displayName="Tabla_Otra", ref="A1:B2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(str(xlsx_path))

    wb2 = load_workbook_safe(xlsx_path)
    assert DispEAParser().extraer(wb2) == []


def test_fila_sin_uid_se_descarta(tmp_path) -> None:
    """Fila sin UID ni Numero se descarta."""
    xlsx_path = save_xlsx_with_disp_table(
        tmp_path, "EA",
        rows=[
            [None, None, "X", "X", "X", "X", "X", "X"],
            build_full_row("EA", UID="EA_001"),
        ],
    )
    wb = load_workbook_safe(xlsx_path)
    result = DispEAParser().extraer(wb)
    assert len(result) == 1
    assert result[0].uid == "EA_001"
