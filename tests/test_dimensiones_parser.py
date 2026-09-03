"""Tests del ``DimensionesParser`` (Fase 5 del plan).

Cubre:
  * Extracción básica de ``N_MAX_*`` / ``Num_Disp_*`` desde los
    defined names del workbook.
  * Defensa ante prefijos inválidos (se ignoran).
  * N_MAX adicionales del catálogo (no legacy) acaban en ``extras``.
  * Workbook sin ``defined_names`` devuelve instancia vacía.
  * El parser es data-driven: el ``ConfigManager`` opcional
    construye el ``named_range_map`` desde el catálogo.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

from areas.alimentacion.domain.models.excel_cache import DimensionesDispositivos
from areas.alimentacion.infrastructure.parsers.disp_dimensiones import DimensionesParser


# ── Helpers ─────────────────────────────────────────────────────────────


def _write_config(tmp_path: Path) -> Path:
    """Escribe un config.json mínimo con n_max_catalog de 6 hw_types."""
    cfg: dict[str, Any] = {
        "departments": {
            "alimentacion": {
                "global_config_table_name": "000_Config_Dispositivos",
                "tia_folders": {
                    "proceso":      "003_Procesos",
                    "dispositivos": "2000_Dispositivos",
                    "nmax":         "000_Sistema",
                },
                "n_max_catalog": [
                    {"name": "N_MAX_DISP_ED",   "hw_type": "ed"},
                    {"name": "N_MAX_DISP_EA",   "hw_type": "ea"},
                    {"name": "N_MAX_DISP_SA",   "hw_type": "sa"},
                    {"name": "N_MAX_DISP_V",    "hw_type": "v"},
                    {"name": "N_MAX_DISP_M",    "hw_type": "m"},
                    {"name": "N_MAX_DISP_M_VF", "hw_type": "m_vf"},
                ],
            }
        }
    }
    p = tmp_path / "config.json"
    p.write_text(json.dumps(cfg), encoding="utf-8")
    return p


def _add_named_value(
    wb: Workbook,
    sheet_name: str,
    cell: str,
    name: str,
    value: Any,
) -> None:
    """Añade un defined name que apunta a ``sheet!cell`` con ``value``."""
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]
    ws[cell] = value
    # ``localSheetId=None`` → workbook-scoped.
    dn = DefinedName(name=name, attr_text=f"'{sheet_name}'!${cell}")
    wb.defined_names[name] = dn


# ── Tests ───────────────────────────────────────────────────────────────


def test_extrae_n_max_basico(tmp_path) -> None:
    """Excel con 6 defined names ``N_MAX_DISP_X`` → 6 contadores correctos."""
    wb = Workbook()
    # Quitar la hoja por defecto
    wb.remove(wb.active)
    # Cada defined name apunta a una celda distinta (A1, A2, ...).
    values = [("ED", 10, "A1"), ("EA", 20, "A2"), ("SA", 30, "A3"),
              ("V", 40, "A4"), ("M", 50, "A5"), ("M_VF", 60, "A6")]
    for hw, val, cell in values:
        _add_named_value(wb, "Config", cell, f"N_MAX_DISP_{hw}", val)

    d = DimensionesParser().extraer(wb)
    assert isinstance(d, DimensionesDispositivos)
    assert d.num_disp_ed == 10
    assert d.num_disp_ea == 20
    assert d.num_disp_sa == 30
    assert d.num_disp_v == 40
    assert d.num_disp_m == 50
    assert d.num_disp_m_vf == 60


def test_falta_defined_name_devuelve_cero(tmp_path) -> None:
    """Solo 2 defined names presentes → los otros 4 quedan en 0."""
    wb = Workbook()
    wb.remove(wb.active)
    _add_named_value(wb, "Config", "A1", "N_MAX_DISP_ED", 7)
    _add_named_value(wb, "Config", "A2", "N_MAX_DISP_V", 8)

    d = DimensionesParser().extraer(wb)
    assert d.num_disp_ed == 7
    assert d.num_disp_v == 8
    assert d.num_disp_ea == 0
    assert d.num_disp_sa == 0
    assert d.num_disp_m == 0
    assert d.num_disp_m_vf == 0


def test_prefijo_invalido_se_ignora(tmp_path) -> None:
    """Defined names que NO empiezan por ``N_MAX_``/``Num_Disp_`` se ignoran."""
    wb = Workbook()
    wb.remove(wb.active)
    _add_named_value(wb, "Config", "A1", "N_MAX_DISP_ED", 5)
    _add_named_value(wb, "Config", "A2", "OTRA_COSA", 999)
    _add_named_value(wb, "Config", "A3", "Empresa", "Acme")

    d = DimensionesParser().extraer(wb)
    assert d.num_disp_ed == 5
    # ``OTRA_COSA`` y ``Empresa`` no acaban en ``extras`` (no tienen
    # el prefijo correcto).
    assert "OTRA_COSA" not in d.extras
    assert "Empresa" not in d.extras


def test_extras_captura_nmax_adicionales(tmp_path) -> None:
    """Un defined name ``N_MAX_DISP_FF`` (no legacy) acaba en ``extras``."""
    wb = Workbook()
    wb.remove(wb.active)
    _add_named_value(wb, "Config", "A1", "N_MAX_DISP_ED", 5)
    _add_named_value(wb, "Config", "A2", "N_MAX_DISP_FF", 42)

    d = DimensionesParser().extraer(wb)
    assert d.num_disp_ed == 5
    assert d.extras.get("N_MAX_DISP_FF") == 42


def test_workbook_sin_defined_names_devuelve_instancia_vacia(tmp_path) -> None:
    """Workbook sin defined names → ``DimensionesDispositivos()`` vacío."""
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Config")  # hoja sin defined names

    d = DimensionesParser().extraer(wb)
    assert d.num_disp_ed == 0
    assert d.num_disp_ea == 0
    assert d.num_disp_sa == 0
    assert d.num_disp_v == 0
    assert d.num_disp_m == 0
    assert d.num_disp_m_vf == 0
    assert dict(d.extras) == {}


def test_with_config_manager_resolves_data_driven(tmp_path) -> None:
    """Si se inyecta ``ConfigManager``, las entradas del ``n_max_catalog``
    se traducen a ``num_disp_<hw>`` data-driven."""
    from core.infrastructure.config_manager import ConfigManager

    config_path = _write_config(tmp_path)
    cm = ConfigManager(config_path=config_path)

    wb = Workbook()
    wb.remove(wb.active)
    # El ConfigManager mapea ``N_MAX_DISP_ED`` (canónico en catalog) → ``num_disp_ed``.
    _add_named_value(wb, "Config", "A1", "N_MAX_DISP_ED", 11)
    _add_named_value(wb, "Config", "A2", "N_MAX_DISP_V", 33)

    d = DimensionesParser(config_manager=cm).extraer(wb)
    assert d.num_disp_ed == 11
    assert d.num_disp_v == 33
