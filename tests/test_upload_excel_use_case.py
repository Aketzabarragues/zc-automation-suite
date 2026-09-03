"""Tests aislados del ``UploadExcelUseCase``.

Objetivo: verificar que el use case es testeable SIN FastAPI
(``TestClient``). Se instancia directamente con dependencias
explícitas y se invoca ``await use_case.execute(xlsx_path)``.

Cubre:
  * Happy path: parsea un xlsx sintético, popula el cache y el
    ``AppState``, devuelve la shape legacy del response.
  * Sad path: con un xlsx que no existe, lanza ``HTTPException(400)``
    y deja el ``ProgressTracker`` en estado ``error``.

Patrón: ``pytest-asyncio`` + Excel sintético en ``tmp_path`` +
fake ``ExcelCacheManager`` (no toca el singleton global).
"""
from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path
from typing import Any
from unittest.mock import MagicMock

import pytest
from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from areas.alimentacion.application.use_cases.upload_excel import (
    UploadExcelUseCase,
)
from areas.alimentacion.infrastructure.cache import ExcelCacheManager
from core.application.log_buffer import LogBuffer
from core.application.progress_buffer import ProgressTracker
from core.application.state import AppState, get_app_state
from core.infrastructure.config_manager import ConfigManager


# ── Configuración JSON fixture ─────────────────────────────────────────


_FULL_CONFIG: dict[str, Any] = {
    "departments": {
        "alimentacion": {
            "global_config_table_name": "000_Config_Dispositivos",
            "tia_folders": {
                "proceso":      "003_Procesos",
                "dispositivos": "2000_Dispositivos",
                "nmax":         "000_Sistema",
            },
            "n_max_catalog": [
                {"name": "N_MAX_DISP_ED",   "value": 10},
                {"name": "N_MAX_DISP_EA",   "value": 10},
                {"name": "N_MAX_DISP_SA",   "value": 10},
                {"name": "N_MAX_DISP_V",    "value": 10},
                {"name": "N_MAX_DISP_M",    "value": 10},
                {"name": "N_MAX_DISP_M_VF", "value": 10},
            ],
            "Dispositivos": {
                "ed": {
                    "db_name": "DB2000_ED", "db_array_name": "ED",
                    "tag_table": "2000_Disp_ED",
                    "config_table": "000_Config_Dispositivos",
                },
                "ea": {
                    "db_name": "DB2001_EA", "db_array_name": "EA",
                    "tag_table": "2000_Disp_EA",
                    "config_table": "000_Config_Dispositivos",
                },
                "sa": {
                    "db_name": "DB2006_SA", "db_array_name": "SA",
                    "tag_table": "2000_Disp_SA",
                    "config_table": "000_Config_Dispositivos",
                },
                "v": {
                    "db_name": "DB2010_V", "db_array_name": "V",
                    "tag_table": "2000_Disp_V",
                    "config_table": "000_Config_Dispositivos",
                },
                "m": {
                    "db_name": "DB2015_M", "db_array_name": "M",
                    "tag_table": "2000_Disp_M",
                    "config_table": "000_Config_Dispositivos",
                },
                "m_vf": {
                    "db_name": "DB2016_M_VF", "db_array_name": "M_VF",
                    "tag_table": "2000_Disp_M_VF",
                    "config_table": "000_Config_Dispositivos",
                },
            },
        }
    }
}


# ── Excel sintético (copia local de _build_minimal_xlsx_bytes) ────────


def _add_table(
    wb: Workbook, sheet_name: str, table_name: str,
    headers: list[str], rows: list[list],
) -> None:
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    last_col_letter = chr(ord("A") + len(headers) - 1)
    last_row = 1 + len(rows)
    ref = f"A1:{last_col_letter}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)


def _build_minimal_xlsx_bytes() -> bytes:
    """Genera un xlsx en memoria con 1 fila de cada uno de los 6 tipos."""
    wb = Workbook()
    wb.remove(wb.active)
    tables = [
        ("DISP_ED",   "Tabla_Disp_ED",   "ED_001",   1, "V_ED_001"),
        ("DISP_EA",   "Tabla_Disp_EA",   "EA_001",   1, "V_EA_001"),
        ("DISP_SA",   "Tabla_Disp_SA",   "SA_001",   1, "V_SA_001"),
        ("DISP_V",    "Tabla_Disp_V",    "V_001",    1, "V_V_001"),
        ("DISP_M",    "Tabla_Disp_M",    "M_001",    1, "V_M_001"),
        ("DISP_M_VF", "Tabla_Disp_M_VF", "MVF_001",  1, "V_MVF_001"),
    ]
    headers = ["UID", "Numero", "PLC.Tag", "Descripcion"]
    for sheet_name, table_name, uid, numero, tag in tables:
        _add_table(wb, sheet_name, table_name, headers,
                   [[uid, numero, tag, f"Desc {uid}"]])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Fake ExcelCacheManager (no toca el singleton global) ──────────────


class FakeCacheManager:
    """Fake que reemplaza a ``ExcelCacheManager`` en los tests.

    Implementa la API mínima que el use case usa: ``put`` como
    ``classmethod`` async. Acumula el último cache en ``last_cache``
    para inspección.
    """

    last_cache: Any = None
    put_call_count: int = 0

    @classmethod
    async def put(cls, cache: Any) -> None:
        cls.last_cache = cache
        cls.put_call_count += 1

    @classmethod
    def reset(cls) -> None:
        cls.last_cache = None
        cls.put_call_count = 0


# ── Fixtures ───────────────────────────────────────────────────────────


@pytest.fixture
def config_manager(tmp_path: Path) -> ConfigManager:
    """ConfigManager con el JSON fixture del subdominio alimentación."""
    p = tmp_path / "config.json"
    p.write_text(json.dumps(_FULL_CONFIG), encoding="utf-8")
    return ConfigManager(config_path=p)


@pytest.fixture
def state() -> AppState:
    """AppState fresco: devices vacíos, cache y dimensiones a None."""
    s = get_app_state()
    s.reset()
    s.excel_cache = None
    s.excel_path = None
    s.dimensiones = None
    yield s
    # Cleanup tras el test: dejar el estado como lo encontramos.
    s.reset()
    s.excel_cache = None
    s.excel_path = None
    s.dimensiones = None


@pytest.fixture
def progress() -> ProgressTracker:
    """ProgressTracker fresco por test (no usa el Singleton global)."""
    return ProgressTracker()


@pytest.fixture
def use_case(
    config_manager: ConfigManager,
    state: AppState,
    progress: ProgressTracker,
) -> UploadExcelUseCase:
    """Use case con todas las dependencias inyectadas explícitamente."""
    FakeCacheManager.reset()
    return UploadExcelUseCase(
        excel_cache_manager=FakeCacheManager,  # type: ignore[arg-type]
        config_manager=config_manager,
        app_state=state,
        progress_tracker=progress,
        log=MagicMock(spec=LogBuffer),
    )


# ── Tests ──────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_execute_happy_path_puebla_cache_y_state(
    use_case: UploadExcelUseCase,
    state: AppState,
    progress: ProgressTracker,
    tmp_path: Path,
) -> None:
    """Happy path: parsea el xlsx, cachea, popula AppState, devuelve
    la shape legacy del response."""
    xlsx_path = tmp_path / "test.xlsx"
    xlsx_path.write_bytes(_build_minimal_xlsx_bytes())

    # El use case espera que el caller haya hecho begin con los
    # stages correctos. Lo replicamos aquí.
    progress.begin(
        operation="upload_excel",
        label="Cargando Excel: test.xlsx",
        stages=["parsear_excel", "volcar_appstate"],
    )

    result = await use_case.execute(xlsx_path)

    # ── Response shape ────────────────────────────────────────────
    assert result["ok"] is True
    assert "summary" in result
    assert "total_dispositivos" in result
    assert "dimensiones" in result
    # 6 dispositivos (1 por tipo).
    assert result["total_dispositivos"] == 6
    # ``dimensiones`` tiene los 6 canónicos.
    assert set(result["dimensiones"].keys()) == {
        "num_disp_ed", "num_disp_ea", "num_disp_sa",
        "num_disp_v", "num_disp_m", "num_disp_m_vf",
    }
    # ``summary`` tiene 6 keys (DispED, DispEA, ...).
    assert len(result["summary"]) == 6
    assert all(v == 1 for v in result["summary"].values())

    # ── Side effects en AppState ──────────────────────────────────
    assert state.excel_cache is not None
    assert state.excel_path is not None
    # El cache guarda la versión absolute() del path.
    assert state.excel_path.endswith("test.xlsx")
    assert state.dimensiones is not None
    # Los 6 atributos legacy están populados.
    for hw in ("ed", "ea", "sa", "v", "m", "m_vf"):
        devices = state.get_devices(hw)
        assert len(devices) == 1, (
            f"Tipo {hw!r} esperaba 1 dispositivo, obtuve {len(devices)}"
        )

    # ── Side effect en el cache manager ───────────────────────────
    assert FakeCacheManager.put_call_count == 1
    assert FakeCacheManager.last_cache is state.excel_cache


@pytest.mark.asyncio
async def test_execute_invalida_progress_stages(
    use_case: UploadExcelUseCase,
    progress: ProgressTracker,
    tmp_path: Path,
) -> None:
    """Los dos stages pasan a ``done`` tras un ``execute`` exitoso.

    No tocamos ``progress.finish`` (lo hace el handler) — solo
    verificamos que ``start_stage`` + ``finish_stage`` deja los
    stages en estado ``done``.
    """
    xlsx_path = tmp_path / "test.xlsx"
    xlsx_path.write_bytes(_build_minimal_xlsx_bytes())
    progress.begin(
        operation="upload_excel",
        label="Cargando Excel: test.xlsx",
        stages=["parsear_excel", "volcar_appstate"],
    )

    await use_case.execute(xlsx_path)

    snap = progress.snapshot()
    by_id = {s["id"]: s for s in snap.stages}
    assert by_id["parsear_excel"]["status"] == "done"
    assert by_id["volcar_appstate"]["status"] == "done"
    # La operación sigue activa: el handler es quien llama a finish.
    assert snap.active is True
    # El detail del primer stage menciona los 6 dispositivos.
    assert "6" in (by_id["parsear_excel"]["detail"] or "")


@pytest.mark.asyncio
async def test_execute_sin_config_manager_lanza_runtime_error(
    state: AppState,
    progress: ProgressTracker,
    tmp_path: Path,
) -> None:
    """Si el use case no recibe config_manager, ``execute`` falla
    con ``RuntimeError`` claro (no con un ``AttributeError`` opaco).

    Esto protege contra el escenario "se construye el use case sin
    inyectar config_manager y luego se ejecuta accidentalmente".
    """
    uc = UploadExcelUseCase(
        excel_cache_manager=FakeCacheManager,  # type: ignore[arg-type]
        config_manager=None,
        app_state=state,
        progress_tracker=progress,
        log=MagicMock(spec=LogBuffer),
    )
    xlsx_path = tmp_path / "test.xlsx"
    xlsx_path.write_bytes(_build_minimal_xlsx_bytes())

    with pytest.raises(RuntimeError, match="config_manager"):
        await uc.execute(xlsx_path)


@pytest.mark.asyncio
async def test_execute_path_inexistente_lanza_http_exception(
    use_case: UploadExcelUseCase,
    progress: ProgressTracker,
    tmp_path: Path,
) -> None:
    """Sad path: con un ``excel_path`` que no existe, el use case
    lanza ``HTTPException(400)``, deja el progress en error y emite
    log de error.

    Decisión documentada: usamos ``HTTPException`` directamente
    desde el use case (ver docstring de ``UploadExcelUseCase``).
    """
    progress.begin(
        operation="upload_excel",
        label="Cargando Excel: missing.xlsx",
        stages=["parsear_excel", "volcar_appstate"],
    )
    bad_path = tmp_path / "no_existe.xlsx"

    with pytest.raises(HTTPException) as exc_info:
        await use_case.execute(bad_path)

    assert exc_info.value.status_code == 400
    assert "excel_upload failed" in str(exc_info.value.detail)

    # El progress se cerró con error.
    snap = progress.snapshot()
    assert snap.active is False
    assert snap.error is not None
    # El stage parsear_excel quedó en error (lo cerró finish()).
    by_id = {s["id"]: s for s in snap.stages}
    assert by_id["parsear_excel"]["status"] == "error"

    # El use case no llegó a cachear ni a tocar el state.
    assert FakeCacheManager.put_call_count == 0
