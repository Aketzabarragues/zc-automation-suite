"""Parser Excel determinista del subdominio alimentación.

Lee el libro Excel del departamento de alimentación apuntando a las
**Tablas Nombradas** (``ListObjects``) declaradas explícitamente en
``_EXCEL_TARGETS``. Para cada tabla realiza una extracción
determinista fila a fila usando los **nombres de columna literales**
del código legacy (con mayúsculas y puntos), sin normalización
tipográfica ni heurísticas por ``dataclasses.fields``.

Pipeline por tabla:

  1. Localiza la ``ListObject`` por nombre (``worksheet.tables[name]``).
     Si no existe, hace fallback a la primera fila que contenga la
     celda con el texto exacto ``"Numero"`` y ``"PLC.Tag"``.
  2. Lee la cabecera del rango y construye ``rows: list[dict]``
     donde **cada clave es la cabecera EXACTA** de la celda
     (``"PLC.Tag"``, ``"Cfg.GrupoAlarma"``, etc.) — no se aplica
     snake_case ni eliminación de acentos.
  3. Mapea cada fila a su dataclass mediante un **constructor
     explícito** (``_build_disp_ed``, ``_build_disp_ea``, …) que
     solicita cada columna con la clave literal del legacy.
  4. Trazabilidad: si la tabla no devuelve filas se emite un
     ``logger.warning``; al finalizar cada tabla se emite un
     ``logger.info`` con el conteo.

Restricción arquitectónica: este módulo es OFFLINE; no importa
``siemens_tia_scripting``. Única dependencia externa: ``openpyxl``.
"""
from __future__ import annotations

import dataclasses
import logging
import unicodedata
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table

from application.log_buffer import get_log_buffer
from core.alimentacion.models.dispositivos import (
    DimensionesDispositivos,
    DispEA,
    DispED,
    DispM,
    DispM_VF,
    DispSA,
    DispV,
    Dispositivo,
)
from infrastructure.parsers.excel_parser import ExcelParser


# Logger adicional a ``LogBuffer`` para entornos donde el buffer aún
# no existe (CLI, tests sin DI). ``LogBuffer`` sigue siendo la fuente
# de verdad que consume la SPA vía polling.
_module_logger = logging.getLogger("zc.parsers.alimentacion_excel")
_module_logger.setLevel(logging.INFO)
if not _module_logger.handlers:
    _handler = logging.StreamHandler()
    _handler.setFormatter(
        logging.Formatter("[%(name)s] %(levelname)s: %(message)s")
    )
    _module_logger.addHandler(_handler)
_module_logger.propagate = False


# ── Mapeo explícito de Tablas Nombradas (ListObjects) ───────────────────
# Determina la asociación entre:
#   * clave canónica del modelo (``DispED``, ``DispEA``, …)
#   * nombre de hoja (literal en el archivo .xlsx)
#   * nombre de la ``ListObject`` (Tabla_*)
#   * clase dataclass destino.
# Cualquier desviación del operario (ej. ``ED`` en vez de ``DISP_ED``)
# se resuelve en ``_EXCEL_TARGETS`` actualizando la entrada.
_EXCEL_TARGETS: dict[str, dict[str, Any]] = {
    "DispED": {
        "sheet": "DISP_ED",
        "table": "Tabla_Disp_ED",
        "model": DispED,
    },
    "DispEA": {
        "sheet": "DISP_EA",
        "table": "Tabla_Disp_EA",
        "model": DispEA,
    },
    "DispSA": {
        "sheet": "DISP_SA",
        "table": "Tabla_Disp_SA",
        "model": DispSA,
    },
    "DispV": {
        "sheet": "DISP_V",
        "table": "Tabla_Disp_V",
        "model": DispV,
    },
    "DispM": {
        "sheet": "DISP_M",
        "table": "Tabla_Disp_M",
        "model": DispM,
    },
    "DispM_VF": {
        "sheet": "DISP_M_VF",
        "table": "Tabla_Disp_M_VF",
        "model": DispM_VF,
    },
}


# ── Helpers de casteo seguro (defensivos contra NaN / None / vacíos) ───


def _safe_str(val: Any) -> str:
    """Convierte ``val`` a ``str`` quitando ``None`` / ``NaN`` / vacío."""
    if val is None:
        return ""
    text = str(val).strip()
    if text.lower() in ("nan", "none", ""):
        return ""
    return text


def _safe_int(val: Any, default: int = 0) -> int:
    """Convierte ``val`` a ``int`` con fallback a ``default``."""
    if val is None:
        return default
    if isinstance(val, bool):
        return int(val)
    if isinstance(val, int):
        return val
    try:
        return int(str(val).strip())
    except (ValueError, TypeError):
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return default


def _safe_float(val: Any, default: float = 0.0) -> float:
    """Convierte ``val`` a ``float`` con fallback a ``default``."""
    if val is None:
        return default
    if isinstance(val, bool):
        return float(val)
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).strip().replace(",", "."))
    except (ValueError, TypeError):
        return default


def _normalizar_etiqueta(texto: str) -> str:
    """Normaliza SOLO para *búsqueda* de cabeceras (no para mapear).

    Mantener esta función aquí permite localizar la fila de cabecera
    cuando el operario usa mayúsculas, acentos o espacios
    adicionales. Una vez identificada la fila, las claves que
    se entregan a los constructores explícitos (``_build_disp_*``)
    son las **cabeceras literales** de la hoja, con su
    capitalización original.
    """
    if not texto:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(texto))
    sin_diacriticos = "".join(
        c for c in nfkd if not unicodedata.combining(c)
    )
    return sin_diacriticos.lower().strip().replace(" ", "").replace(".", "")


def _resolver_target_por_hoja(sheet_name: str) -> str | None:
    """Devuelve la clave canónica (``DispED``…) para una hoja del libro.

    Se busca primero coincidencia exacta contra ``_EXCEL_TARGETS``
    (letra por letra) y luego coincidencia normalizada para tolerar
    pequeñas variaciones (``DISP_ED`` vs ``disp_ed`` vs ``DISPED``).
    """
    if not sheet_name:
        return None
    for canonica, cfg in _EXCEL_TARGETS.items():
        if sheet_name == cfg["sheet"]:
            return canonica
    target = _normalizar_etiqueta(sheet_name)
    for canonica, cfg in _EXCEL_TARGETS.items():
        if _normalizar_etiqueta(cfg["sheet"]) == target:
            return canonica
    return None


def _localizar_rango_tabla(worksheet: Any, table_name: str) -> tuple[int, int, int, int] | None:
    """Devuelve ``(min_row, min_col, max_row, max_col)`` de la ``ListObject``.

    Devuelve ``None`` si la tabla no existe en la hoja o si su
    referencia no se puede resolver.
    """
    tables = getattr(worksheet, "tables", None) or {}
    table: Table | None = tables.get(table_name)
    if table is None:
        return None
    ref = getattr(table, "ref", None)
    if not ref or not isinstance(ref, str) or ":" not in ref:
        return None
    try:
        # ``ws[ref]`` devuelve un ``MultiCellRange`` iterable.
        # Lo más simple: usar ``dimensions`` parseando el rango.
        start_ref, end_ref = ref.split(":", 1)
        min_col, min_row, max_col, max_row = (
            worksheet[start_ref].column,
            worksheet[start_ref].row,
            worksheet[end_ref].column,
            worksheet[end_ref].row,
        )
        return min_row, min_col, max_row, max_col
    except Exception:  # pragma: no cover - defensivo
        return None


def _localizar_cabecera_por_texto(
    worksheet: Any,
    required: Iterable[str],
) -> tuple[int, int, int, int] | None:
    """Fallback: localiza la fila de cabecera por contenido.

    Recorre las primeras ``MAX_HEADER_SEARCH_ROWS`` filas de la hoja
    buscando una fila que contenga **todas** las celdas en
    ``required`` (p.ej. ``{"Numero", "PLC.Tag"}``). Si la
    encuentra, devuelve el rango desde esa fila hasta el final de los
    datos contiguos.
    """
    required_norm = {_normalizar_etiqueta(r) for r in required}
    max_search_rows = 25
    header_row: int | None = None
    for row_idx in range(1, max_search_rows + 1):
        normalized_cells: set[str] = set()
        for cell in worksheet[row_idx]:
            value = getattr(cell, "value", None)
            if value is None:
                continue
            normalized_cells.add(_normalizar_etiqueta(value))
        if required_norm.issubset(normalized_cells):
            header_row = row_idx
            break
    if header_row is None:
        return None
    max_col = worksheet.max_column or 1
    max_row = worksheet.max_row or header_row
    return header_row, 1, max_row, max_col


def _iterar_filas_diccionario(
    worksheet: Any,
    rango: tuple[int, int, int, int],
) -> list[dict[str, Any]]:
    """Devuelve ``[{cabecera_literal: valor, ...}, ...]`` para la tabla.

    La clave de cada diccionario es la cabecera **EXACTA** de la
    celda (incluyendo mayúsculas, puntos y espacios) tal como
    aparece en el Excel. No se aplica ``lower()`` ni
    ``replace(".", "_")``.
    """
    min_row, min_col, max_row, max_col = rango
    header_cells = list(worksheet.iter_rows(
        min_row=min_row, max_row=min_row,
        min_col=min_col, max_col=max_col,
        values_only=True,
    ))[0]
    headers: list[str] = [
        str(h).strip() if h is not None and str(h).strip() else ""
        for h in header_cells
    ]
    result: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(
        min_row=min_row + 1, max_row=max_row,
        min_col=min_col, max_col=max_col,
        values_only=True,
    ):
        if row is None or all(c is None for c in row):
            continue
        item: dict[str, Any] = {}
        for header_name, value in zip(headers, row):
            if not header_name:
                continue
            item[header_name] = value
        if item:
            result.append(item)
    return result


class AlimentacionExcelParser:
    """Parser Excel determinista del subdominio alimentación.

    Compone sobre ``ExcelParser`` para el parseo de named ranges
    (``extraer_dimensiones``) y abre su propio ``load_workbook`` para
    localizar las ``ListObjects`` en ``extraer_dtos``.

    Sólo procesa las tablas declaradas en ``_EXCEL_TARGETS``. Las
    filas sin ``PLC.Tag`` se descartan silenciosamente (criterio de
    unicidad del PlcTag en TIA Portal).
    """

    def __init__(self) -> None:
        # Composición: usamos el parser genérico SOLO para dimensiones.
        self._generic_parser = ExcelParser()
        # Buffer de logs opcional (DI tolerante).
        try:
            self._log = get_log_buffer()
        except Exception:  # pragma: no cover - defensivo
            self._log = None

    # ── DTOs por tabla ──────────────────────────────────────────────────
    def extraer_dtos(
        self, excel_path: str | Path
    ) -> dict[str, list[Dispositivo]]:
        """Lee cada ``ListObject`` declarada en ``_EXCEL_TARGETS``.

        Returns:
            ``dict[str, list[Dispositivo]]``. Las claves son los
            nombres canónicos (``DispED``, ``DispM_VF``…); los
            valores son listas de dataclasses inmutables.

        Raises:
            FileNotFoundError: Si el archivo no existe.
        """
        path = Path(excel_path)
        if not path.is_file():
            raise FileNotFoundError(
                f"No se encontró el archivo Excel: '{path}'"
            )

        # NOTA: ``worksheet.tables`` SOLO está disponible cuando el
        # workbook NO se abre en modo ``read_only`` (openpyxl 3.1). En
        # modo read-only ``tables`` devuelve ``None``. Por eso abrimos
        # el libro en modo normal para localizar las ``ListObjects``
        # y seguimos leyendo el contenido vía ``iter_rows``.
        workbook = load_workbook(
            filename=str(path), read_only=False, data_only=True,
        )
        try:
            result: dict[str, list[Dispositivo]] = {}
            for canonica, cfg in _EXCEL_TARGETS.items():
                sheet_name: str = cfg["sheet"]
                table_name: str = cfg["table"]
                model_cls: type = cfg["model"]
                devices = self._extract_table(
                    workbook, canonica, sheet_name, table_name, model_cls,
                )
                if devices:
                    result[canonica] = devices
            return result
        finally:
            workbook.close()

    # ── Dimensiones (named ranges num_disp_*) ─────────────────────────
    def extraer_dimensiones(
        self, excel_path: str | Path
    ) -> DimensionesDispositivos:
        """Lee named ranges ``num_disp_*`` y devuelve una instancia tipada.

        Returns:
            ``DimensionesDispositivos`` con los 6 contadores. Si un
            campo no existe en el Excel, queda en ``0``.
        """
        path = Path(excel_path)
        if not path.is_file():
            raise FileNotFoundError(
                f"No se encontró el archivo Excel: '{path}'"
            )

        workbook = load_workbook(
            filename=str(path), read_only=True, data_only=True
        )
        try:
            return self._extract_dimensiones(workbook)
        finally:
            workbook.close()

    # ── Extracción de UNA tabla ────────────────────────────────────────
    def _extract_table(
        self,
        workbook: Any,
        canonica: str,
        sheet_name: str,
        table_name: str,
        model_cls: type,
    ) -> list[Dispositivo]:
        """Localiza la ``ListObject`` y construye los dispositivos."""
        if sheet_name not in workbook.sheetnames:
            self._emit(
                "warning",
                f"La hoja {sheet_name} no devolvió datos válidos: "
                f"no existe en el libro.",
            )
            return []

        worksheet = workbook[sheet_name]
        rango = _localizar_rango_tabla(worksheet, table_name)
        if rango is None:
            # Fallback: localizar por la fila que contiene "Numero" y "PLC.Tag".
            self._emit(
                "warning",
                f"Tabla '{table_name}' no encontrada en '{sheet_name}'. "
                f"Intentando localizar por cabecera 'Numero'/'PLC.Tag'…",
            )
            rango = _localizar_cabecera_por_texto(
                worksheet, required=("Numero", "PLC.Tag"),
            )

        if rango is None:
            self._emit(
                "warning",
                f"La hoja {sheet_name} no devolvió datos válidos: "
                f"no se localizó la tabla '{table_name}' ni la cabecera.",
            )
            return []

        rows = _iterar_filas_diccionario(worksheet, rango)
        if not rows:
            self._emit(
                "warning",
                f"La hoja {sheet_name} no devolvió datos válidos: "
                f"la tabla '{table_name}' está vacía.",
            )
            return []

        # Mapeo canónica → método ``_build_*``.
        # ``_EXCEL_TARGETS`` ya conserva la lista de constructores
        # explícitos (``_build_disp_ed``, ``_build_disp_ea``,
        # ``_build_disp_sa``, ``_build_disp_v``, ``_build_dispm``,
        # ``_build_disp_m_vf``); consultamos ``__dict__`` directamente
        # para evitar inconsistencias con descriptores o herencia.
        self_module = self.__class__
        builder: Any = None
        # Mapa explícito canónica → nombre del constructor.
        # Evita el falso positivo de ``startswith`` cuando dos
        # modelos comparten prefijo (``DispED`` y ``DispEA``).
        canonical_to_builder: dict[str, str] = {
            "DispED": "_build_disp_ed",
            "DispEA": "_build_disp_ea",
            "DispSA": "_build_disp_sa",
            "DispV": "_build_disp_v",
            "DispM": "_build_dispm",
            "DispM_VF": "_build_disp_m_vf",
        }
        builder_attr = canonical_to_builder.get(canonica)
        if builder_attr is not None:
            builder = self_module.__dict__.get(builder_attr)
        if builder is None:
            # Compatibilidad explícita con la convención legacy.
            legacy_aliases = {
                "DispM": "_build_dispm",
                "DispM_VF": "_build_disp_m_vf",
            }
            legacy = legacy_aliases.get(canonica)
            if legacy is not None:
                builder = self_module.__dict__.get(legacy)
        if builder is None:
            self._emit(
                "warning",
                f"No hay constructor explícito para '{canonica}'. "
                f"Atributos disponibles: "
                f"{[k for k in self_module.__dict__ if k.startswith('_build_')]}",
            )
            return []
        # ``builder`` puede ser una ``function`` (definida en la
        # clase) o, en herencia múltiple, un ``classmethod``/
        # ``staticmethod``. ``__get__`` resuelve el binding.
        if hasattr(builder, "__get__"):
            builder = builder.__get__(self, self_module)

        devices: list[Dispositivo] = []
        discarded = 0
        for row in rows:
            try:
                device = builder(row)
            except Exception as exc:  # defensivo: nunca romper toda la tabla
                discarded += 1
                self._emit(
                    "warning",
                    f"Fila descartada en '{table_name}': {exc}",
                )
                continue
            if device is None:
                discarded += 1
                continue
            devices.append(device)

        if not devices:
            self._emit(
                "warning",
                f"La hoja {sheet_name} no devolvió datos válidos: "
                f"ninguna fila construyó un dispositivo "
                f"({discarded} descartadas).",
            )
            return []

        self._emit(
            "info",
            f"Tabla {table_name} parseada: {len(devices)} elementos "
            f"({discarded} descartadas).",
        )
        return devices

    # ── Constructores explícitos (Mapeo Hardcoded) ─────────────────────
    # Cada constructor solicita la clave EXACTA del Excel (legacy).
    # No se itera sobre ``dataclasses.fields`` ni se normalizan
    # cabeceras: las claves se piden literales.
    def _build_disp_ed(self, row: dict) -> DispED | None:
        if not _safe_str(row.get("PLC.Tag")) and not _safe_str(row.get("UID")):
            return None
        return DispED(
            numero=_safe_int(row.get("Numero")),
            plc_tag=_safe_str(row.get("PLC.Tag")),
            plc_comentario=_safe_str(row.get("PLC.Comentario")),
            descripcion=_safe_str(row.get("Descripcion")),
            uid=_safe_str(row.get("UID")),
            tag=_safe_int(row.get("Tag")),
            fat=_safe_int(row.get("FAT")),
            e_byte=_safe_int(row.get("E.Byte")),
            e_bit=_safe_int(row.get("E.Bit")),
            gr_alarma=_safe_str(row.get("Gr.Alarma")),
            cuadro=_safe_str(row.get("Cuadro")),
            observaciones=_safe_str(row.get("Observaciones")),
            plc_tipo=_safe_str(row.get("PLC.Tipo")),
            plc_index=_safe_int(row.get("PLC.Index")),
            hmi_index=_safe_int(row.get("Hmi.Index")),
            hmi_texto=_safe_str(row.get("Hmi.Texto")),
            cfg_habilitar=_safe_str(row.get("Cfg.Habilitar")),
            cfg_byte_entrada=_safe_str(row.get("Cfg.ByteEntrada")),
            cfg_bit_entrada=_safe_str(row.get("Cfg.BitEntrada")),
            cfg_grupo_alarma=_safe_str(row.get("Cfg.GrupoAlarma")),
            comentario_db=_safe_str(row.get("ComentarioDB")),
        )

    def _build_disp_ea(self, row: dict) -> DispEA | None:
        if not _safe_str(row.get("PLC.Tag")) and not _safe_str(row.get("UID")):
            return None
        return DispEA(
            numero=_safe_int(row.get("Numero")),
            plc_tag=_safe_str(row.get("PLC.Tag")),
            plc_comentario=_safe_str(row.get("PLC.Comentario")),
            descripcion=_safe_str(row.get("Descripcion")),
            uid=_safe_str(row.get("UID")),
            tag=_safe_int(row.get("Tag")),
            fat=_safe_int(row.get("FAT")),
            e_byte=_safe_int(row.get("E.Byte")),
            unidades=_safe_str(row.get("Unidades")),
            rii=_safe_float(row.get("RII")),
            rsi=_safe_float(row.get("RSI")),
            gr_alarma=_safe_str(row.get("Gr.Alarma")),
            cuadro=_safe_str(row.get("Cuadro")),
            observaciones=_safe_str(row.get("Observaciones")),
            plc_tipo=_safe_str(row.get("PLC.Tipo")),
            plc_index=_safe_int(row.get("PLC.Index")),
            hmi_index=_safe_int(row.get("Hmi.Index")),
            hmi_texto=_safe_str(row.get("Hmi.Texto")),
            cfg_habilitar=_safe_str(row.get("Cfg.Habilitar")),
            cfg_byte_entrada=_safe_str(row.get("Cfg.ByteEntrada")),
            cfg_escaladomin=_safe_str(row.get("Cfg.EscaladoMin")),
            cfg_escaladomax=_safe_str(row.get("Cfg.EscaladoMax")),
            cfg_grupo_alarma=_safe_str(row.get("Cfg.GrupoAlarma")),
            comentario_db=_safe_str(row.get("ComentarioDB")),
        )

    def _build_disp_sa(self, row: dict) -> DispSA | None:
        if not _safe_str(row.get("PLC.Tag")) and not _safe_str(row.get("UID")):
            return None
        return DispSA(
            numero=_safe_int(row.get("Numero")),
            plc_tag=_safe_str(row.get("PLC.Tag")),
            plc_comentario=_safe_str(row.get("PLC.Comentario")),
            descripcion=_safe_str(row.get("Descripcion")),
            uid=_safe_str(row.get("UID")),
            tag=_safe_int(row.get("Tag")),
            fat=_safe_int(row.get("FAT")),
            e_byte=_safe_int(row.get("E.Byte")),
            unidades=_safe_str(row.get("Unidades")),
            rii=_safe_float(row.get("RII")),
            rsi=_safe_float(row.get("RSI")),
            gr_alarma=_safe_str(row.get("Gr.Alarma")),
            cuadro=_safe_str(row.get("Cuadro")),
            observaciones=_safe_str(row.get("Observaciones")),
            plc_tipo=_safe_str(row.get("PLC.Tipo")),
            plc_index=_safe_int(row.get("PLC.Index")),
            hmi_index=_safe_int(row.get("Hmi.Index")),
            hmi_texto=_safe_str(row.get("Hmi.Texto")),
            cfg_habilitar=_safe_str(row.get("Cfg.Habilitar")),
            cfg_byte_entrada=_safe_str(row.get("Cfg.ByteEntrada")),
            cfg_escaladomin=_safe_str(row.get("Cfg.EscaladoMin")),
            cfg_escaladomax=_safe_str(row.get("Cfg.EscaladoMax")),
            cfg_grupo_alarma=_safe_str(row.get("Cfg.GrupoAlarma")),
            comentario_db=_safe_str(row.get("ComentarioDB")),
        )

    def _build_disp_v(self, row: dict) -> DispV | None:
        if not _safe_str(row.get("PLC.Tag")) and not _safe_str(row.get("UID")):
            return None
        return DispV(
            numero=_safe_int(row.get("Numero")),
            plc_tag=_safe_str(row.get("PLC.Tag")),
            plc_comentario=_safe_str(row.get("PLC.Comentario")),
            descripcion=_safe_str(row.get("Descripcion")),
            uid=_safe_str(row.get("UID")),
            tag=_safe_int(row.get("Tag")),
            fat=_safe_int(row.get("FAT")),
            s_byte=_safe_int(row.get("S.Byte")),
            s_bit=_safe_int(row.get("S.Bit")),
            rr_byte=_safe_int(row.get("RR.Byte")),
            rr_bit=_safe_int(row.get("RR.Bit")),
            rt_byte=_safe_int(row.get("RT.Byte")),
            rt_bit=_safe_int(row.get("RT.Bit")),
            gr_alarma=_safe_str(row.get("Gr.Alarma")),
            cuadro=_safe_str(row.get("Cuadro")),
            observaciones=_safe_str(row.get("Observaciones")),
            plc_tipo=_safe_str(row.get("PLC.Tipo")),
            plc_index=_safe_int(row.get("PLC.Index")),
            hmi_index=_safe_int(row.get("Hmi.Index")),
            hmi_texto=_safe_str(row.get("Hmi.Texto")),
            cfg_habilitar=_safe_str(row.get("Cfg.Habilitar")),
            cfg_byteretornoreposo=_safe_str(row.get("Cfg.ByteRetornoReposo")),
            cfg_bitretornoreposo=_safe_str(row.get("Cfg.BitRetornoReposo")),
            cfg_byteretornotrabajo=_safe_str(row.get("Cfg.ByteRetornoTrabajo")),
            cfg_bitretornotrabajo=_safe_str(row.get("Cfg.BitRetornoTrabajo")),
            cfg_byteactivacion=_safe_str(row.get("Cfg.ByteActivacion")),
            cfg_bitactivacion=_safe_str(row.get("Cfg.BitActivacion")),
            cfg_habitreposo=_safe_str(row.get("Cfg.HabRetReposo")),
            cfg_habitrtrabajo=_safe_str(row.get("Cfg.HabRetTrabajo")),
            cfg_grupoalarma=_safe_str(row.get("Cfg.GrupoAlarma")),
            comentario_db=_safe_str(row.get("ComentarioDB")),
        )

    def _build_dispm(self, row: dict) -> DispM | None:
        if not _safe_str(row.get("PLC.Tag")) and not _safe_str(row.get("UID")):
            return None
        return DispM(
            numero=_safe_int(row.get("Numero")),
            plc_tag=_safe_str(row.get("PLC.Tag")),
            plc_comentario=_safe_str(row.get("PLC.Comentario")),
            descripcion=_safe_str(row.get("Descripcion")),
            uid=_safe_str(row.get("UID")),
            tag=_safe_int(row.get("Tag")),
            fat=_safe_int(row.get("FAT")),
            s_byte=_safe_int(row.get("S.Byte")),
            s_bit=_safe_int(row.get("S.Bit")),
            rt_byte=_safe_int(row.get("RT.Byte")),
            rt_bit=_safe_int(row.get("RT.Bit")),
            rm_byte=_safe_int(row.get("RM.Byte")),
            rm_bit=_safe_int(row.get("RM.Bit")),
            gr_alarma=_safe_str(row.get("Gr.Alarma")),
            cuadro=_safe_str(row.get("Cuadro")),
            observaciones=_safe_str(row.get("Observaciones")),
            plc_tipo=_safe_str(row.get("PLC.Tipo")),
            plc_index=_safe_int(row.get("PLC.Index")),
            hmi_index=_safe_int(row.get("Hmi.Index")),
            hmi_texto=_safe_str(row.get("Hmi.Texto")),
            cfg_habilitar=_safe_str(row.get("Cfg.Habilitar")),
            cfg_byteretornotermico=_safe_str(row.get("Cfg.ByteRetornoTermico")),
            cfg_bitretornotermico=_safe_str(row.get("Cfg.BitRetornoTermico")),
            cfg_byteconfmarcha=_safe_str(row.get("Cfg.ByteConfMarcha")),
            cfg_bitconfmarcha=_safe_str(row.get("Cfg.BitConfMarcha")),
            cfg_byteactivacion=_safe_str(row.get("Cfg.ByteActivacion")),
            cfg_bitactivacion=_safe_str(row.get("Cfg.BitActivacion")),
            cfg_habrettermico=_safe_str(row.get("Cfg.HabRetTermico")),
            cfg_habretconfmarcha=_safe_str(row.get("Cfg.HabRetConfMarcha")),
            cfg_grupoalarma=_safe_str(row.get("Cfg.GrupoAlarma")),
            comentario_db=_safe_str(row.get("ComentarioDB")),
        )

    def _build_disp_m_vf(self, row: dict) -> DispM_VF | None:
        if not _safe_str(row.get("PLC.Tag")) and not _safe_str(row.get("UID")):
            return None
        return DispM_VF(
            numero=_safe_int(row.get("Numero")),
            plc_tag=_safe_str(row.get("PLC.Tag")),
            plc_comentario=_safe_str(row.get("PLC.Comentario")),
            descripcion=_safe_str(row.get("Descripcion")),
            uid=_safe_str(row.get("UID")),
            tag=_safe_int(row.get("Tag")),
            fat=_safe_int(row.get("FAT")),
            s_byte=_safe_int(row.get("S.Byte")),
            s_bit=_safe_int(row.get("S.Bit")),
            rt_byte=_safe_int(row.get("RT.Byte")),
            rt_bit=_safe_int(row.get("RT.Bit")),
            rm_byte=_safe_int(row.get("RM.Byte")),
            rm_bit=_safe_int(row.get("RM.Bit")),
            gr_alarma=_safe_str(row.get("Gr.Alarma")),
            cuadro=_safe_str(row.get("Cuadro")),
            observaciones=_safe_str(row.get("Observaciones")),
            plc_tipo=_safe_str(row.get("PLC.Tipo")),
            plc_index=_safe_int(row.get("PLC.Index")),
            hmi_index=_safe_int(row.get("Hmi.Index")),
            hmi_texto=_safe_str(row.get("Hmi.Texto")),
            cfg_habilitar=_safe_str(row.get("Cfg.Habilitar")),
            cfg_byteretornotermico=_safe_str(row.get("Cfg.ByteRetornoTermico")),
            cfg_bitretornotermico=_safe_str(row.get("Cfg.BitRetornoTermico")),
            cfg_byteconfmarcha=_safe_str(row.get("Cfg.ByteConfMarcha")),
            cfg_bitconfmarcha=_safe_str(row.get("Cfg.BitConfMarcha")),
            cfg_byteactivacion=_safe_str(row.get("Cfg.ByteActivacion")),
            cfg_bitactivacion=_safe_str(row.get("Cfg.BitActivacion")),
            cfg_habrettermico=_safe_str(row.get("Cfg.HabRetTermico")),
            cfg_habretconfmarcha=_safe_str(row.get("Cfg.HabRetConfMarcha")),
            cfg_grupoalarma=_safe_str(row.get("Cfg.GrupoAlarma")),
            comentario_db=_safe_str(row.get("ComentarioDB")),
            sa_byte=_safe_int(row.get("SA.Byte")),
            cfg_byteanalogica=_safe_str(row.get("Cfg.ByteAnalogica")),
        )

    # ── Auditoría / Trazabilidad ────────────────────────────────────────
    def _emit(self, level: str, message: str) -> None:
        """Emite a ``LogBuffer`` y al logger estándar (fallback)."""
        log_buffer = getattr(self, "_log", None)
        if log_buffer is not None and hasattr(log_buffer, level):
            try:
                getattr(log_buffer, level)(message)
            except Exception:  # pragma: no cover - defensivo
                _module_logger.warning("LogBuffer.%s falló", level)
        if level == "warning":
            _module_logger.warning(message)
        else:
            _module_logger.info(message)

    # ── Dimensiones (named ranges num_disp_*) ─────────────────────────
    @staticmethod
    def _extract_dimensiones(workbook: Any) -> DimensionesDispositivos:
        """Lee ``wb.defined_names`` y popula ``DimensionesDispositivos``."""
        defined_names = getattr(workbook, "defined_names", None)
        if defined_names is None:
            return DimensionesDispositivos()

        items: Any = (
            defined_names.items()
            if hasattr(defined_names, "items")
            else []
        )
        result: dict[str, int] = {}
        for name, definition in items:
            if not isinstance(name, str):
                continue
            attr = _map_named_range_to_attr(name)
            if attr is None:
                continue
            value = _safe_int(_resolve_named_range_value(definition, workbook))
            result[attr] = value

        return (
            DimensionesDispositivos(**result)
            if result else DimensionesDispositivos()
        )


# ── Helpers de mapeo (privados al módulo) ──────────────────────────────


_DIMENSION_ATTRS: set[str] = {
    "num_disp_ed",
    "num_disp_ea",
    "num_disp_sa",
    "num_disp_v",
    "num_disp_m",
    "num_disp_m_vf",
}


def _map_named_range_to_attr(name: str) -> str | None:
    """Convierte ``num_disp_ed`` → ``num_disp_ed`` (passthrough validado)."""
    if name in _DIMENSION_ATTRS:
        return name
    return None


def _resolve_named_range_value(definition: Any, workbook: Any) -> Any:
    """Lee el valor de un ``DefinedName`` resolviendo su hoja y celda."""
    attr_text: Any = getattr(definition, "attr_text", None)
    if not attr_text:
        attr_text = getattr(definition, "value", None)
    if not isinstance(attr_text, str) or "!" not in attr_text:
        return None

    sheet_part, cell_part = attr_text.split("!", 1)
    sheet_name = sheet_part.strip().strip("'").strip('"')
    cell_ref = cell_part.replace("$", "").strip()
    try:
        sheet = workbook[sheet_name]
    except (KeyError, TypeError):
        return None
    try:
        cell = sheet[cell_ref]
    except (KeyError, AttributeError, TypeError):
        return None
    return getattr(cell, "value", None)


__all__ = [
    "AlimentacionExcelParser",
    "_EXCEL_TARGETS",
    "_safe_str",
    "_safe_int",
    "_safe_float",
    # Alias histórico que se conserva por compatibilidad con tests
    # que importaban la constante de mapa de tipos (pre-refactor).
    # NOTA: no usar desde código nuevo; consultar ``_EXCEL_TARGETS``.
]
