"""Helpers compartidos para los parsers del Excel corporativo.

Este módulo concentra utilidades usadas por **todos** los parsers de
software y de hardware del subdominio ``alimentacion`` (los mini parsers
creados en las Fases 1-5 del plan
``_plan/04_excel_cache_phased_plan.md``). Su objetivo es eliminar
duplicación de código entre parsers (cada uno abría el workbook,
localizaba la ``ListObject``, iteraba filas, descartaba vacías...) y
unificar la **semántica defensiva** del cast de valores (``_safe_str``
/ ``_safe_int`` / ``_safe_float``).

Convención del helper:
    * ``_safe_str`` **NUNCA** devuelve ``None`` (devuelve ``""``). Esto
      es coherente con los DTOs ``frozen=True`` del subdominio (todos
      tienen ``str = ""`` como default) y con el comportamiento del
      parser consolidado ``AlimentacionExcelParser`` (que también
      normaliza a ``""``).
    * ``_safe_int`` y ``_safe_float`` aceptan ``None``/``bool``/
      ``int``/``float``/``str`` numérico. ``bool`` se trata como
      ``int``/``float`` para no perder ``True``/``False`` legítimos.
    * ``extract_list_object_rows`` es la única puerta de entrada a las
      ``ListObject`` (Tablas Nombradas). Si la hoja o la tabla no
      existen, devuelve ``[]`` (no lanza) — política coherente con el
      R1 del plan.

Restricción arquitectónica: este módulo es OFFLINE; no importa
``siemens_tia_scripting``. Única dependencia externa: ``openpyxl``.
"""
from __future__ import annotations

import logging
from typing import Any

from openpyxl import Workbook
from openpyxl.utils.cell import range_boundaries


logger = logging.getLogger(__name__)


# ── Cast defensivo de valores ────────────────────────────────────────────


def _safe_str(val: Any) -> str:
    """Convierte ``val`` a ``str``. NUNCA devuelve ``None``.

    Reglas:
        * ``None`` → ``""``
        * ``"nan"`` / ``"NaN"`` / ``"None"`` / ``"null"`` (case
          insensitive) → ``""``
        * Cadena vacía o solo whitespace → ``""``
        * Resto → ``str(val).strip()``

    Esta semántica es coherente con los DTOs ``frozen=True`` del
    subdominio (todos tienen ``str = ""`` como default) y garantiza
    que las claves ``cfg_*`` y los nombres de PlcTag lleguen
    literales a la SPA sin ``None`` que rompan el template.
    """
    if val is None:
        return ""
    text = str(val).strip()
    if text.lower() in ("nan", "none", "null", ""):
        return ""
    return text


def _safe_int(val: Any, default: int = 0) -> int:
    """Convierte ``val`` a ``int`` con fallback a ``default``.

    Acepta:
        * ``None`` → ``default``
        * ``bool`` → ``int(val)`` (``True`` → 1, ``False`` → 0)
        * ``int`` → ``val`` (sin cambios)
        * ``str`` numérico (``"5"``, ``"5.0"``) → ``int``
        * ``float`` → ``int(val)`` (truncado, no redondeado)

    Cualquier otro caso (ej. ``"Pendiente"``) devuelve ``default`` en
    lugar de lanzar. Esto es **defensivo**: un Excel malformado no debe
    tumbar la carga.
    """
    if val is None:
        return default
    if isinstance(val, bool):
        return int(val)
    if isinstance(val, int):
        return val
    try:
        return int(str(val).strip())
    except (ValueError, TypeError):
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return default


def _safe_float(val: Any, default: float = 0.0) -> float:
    """Convierte ``val`` a ``float`` con fallback a ``default``.

    Acepta coma decimal (``"1,5"`` → ``1.5``) — el Excel del
    operario usa coma como separador decimal en algunas hojas
    (típicamente las de rangos/escalado de ``DispEA`` / ``DispSA``).
    """
    if val is None:
        return default
    if isinstance(val, bool):
        return float(val)
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).strip().replace(",", "."))
    except (ValueError, TypeError):
        return default


def _safe_num_lista(value: Any) -> int | str:
    """None/NaN/vacío → 0. Numérico → int. Texto no numérico → str.

    Helper **específico** de los parsers de parámetros (Fase 2 y 3
    del plan ``_plan/04_excel_cache_phased_plan.md``). La columna
    ``Num.Lista`` del Excel corporativo puede contener **dos clases
    de valores**:

        * Valores numéricos (``0``, ``1``, ``2``, …) que el operario
          usa como índice de selección en una lista HMI.
        * Texto literal (``"N/A"``, ``"TODOS"``, …) que el operario
          usa como marcador semántico (sin lista asociada, todos
          los items, etc.).

    Reglas:
        * ``None`` / ``""`` / ``"nan"`` / ``"None"`` / ``"null"`` /
          whitespace puro → ``0`` (interpretado como "sin lista").
        * ``int`` / ``float`` / ``str`` numérico (``"5"``,
          ``"5.0"``) → ``int`` (truncado, no redondeado).
        * Cualquier otro texto (``"N/A"``, ``"TODOS"``, …) se
          preserva literal como ``str``.

    Diferencia con ``_safe_int``: ``_safe_int`` cae a ``0`` para
    texto no numérico, lo cual sería destructivo para los
    marcadores semánticos del operario. Aquí se preserva el texto
    para que el DTO ``ParamRealPLC.num_lista: int | str`` refleje
    fielmente el Excel.
    """
    cleaned = _safe_str(value)
    if not cleaned:
        return 0
    try:
        return int(float(cleaned))
    except (ValueError, TypeError):
        return cleaned


# ── Extracción de ListObjects (Tablas Nombradas) ─────────────────────────


def extract_list_object_rows(
    wb: Workbook,
    sheet: str,
    table: str,
) -> list[dict[str, Any]]:
    """Lee la ``ListObject`` ``table`` de la hoja ``sheet`` y devuelve filas como dicts.

    Args:
        wb: workbook de openpyxl **ya abierto** (no se cierra aquí;
            la responsabilidad es del caller / loader).
        sheet: nombre literal de la hoja (case-sensitive en openpyxl).
        table: nombre de la ``ListObject`` (Tabla_*).

    Returns:
        Lista de ``{cabecera_literal: valor, ...}`` con una entrada
        por fila de datos. Las filas completamente vacías
        (``all(cell is None)``) se descartan. Las cabeceras vacías
        (``""`` o ``None``) se omiten del dict.

    Política defensiva (R1 del plan):
        * Si la hoja no existe → ``[]``.
        * Si la tabla no existe en la hoja → ``[]``.
        * Si el ``ref`` de la tabla es inválido → ``[]``.
        * Si ``range_boundaries`` lanza → ``[]``.

    **NUNCA** lanza: un Excel malformado o una hoja renombrada
    simplemente devuelve lista vacía. El caller (parser) emite WARNING
    si lo necesita, pero ``extract_list_object_rows`` no se queja.
    """
    if sheet not in wb.sheetnames:
        return []
    worksheet = wb[sheet]
    tables = getattr(worksheet, "tables", None) or {}
    table_obj = tables.get(table)
    if table_obj is None:
        return []
    ref = getattr(table_obj, "ref", None)
    if not ref or not isinstance(ref, str) or ":" not in ref:
        return []

    try:
        min_col, min_row, max_col, max_row = range_boundaries(ref)
    except Exception:  # pragma: no cover - defensivo
        return []

    # ── Cabeceras literales (1ª fila del rango) ───────────────────────
    header_cells = next(
        worksheet.iter_rows(
            min_row=min_row,
            max_row=min_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )
    )
    headers: list[str] = [
        str(h).strip() if h is not None and str(h).strip() else ""
        for h in header_cells
    ]

    # ── Datos (resto del rango) ───────────────────────────────────────
    rows: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(
        min_row=min_row + 1,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    ):
        if row is None or all(c is None for c in row):
            continue
        item: dict[str, Any] = {}
        for header_name, value in zip(headers, row):
            if not header_name:
                continue
            item[header_name] = value
        if item:
            rows.append(item)
    return rows


__all__ = [
    "_safe_str",
    "_safe_int",
    "_safe_float",
    "_safe_num_lista",
    "extract_list_object_rows",
]
