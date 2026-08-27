"""Parser Excel determinista del subdominio alimentación.

Lee el libro Excel del departamento de alimentación apuntando a las
**Tablas Nombradas** (``ListObjects``) declaradas explícitamente en
``_EXCEL_TARGETS`` (módulo) o, si se inyecta un ``ConfigManager``, en
su ``excel_target`` por ``hw_type``. Para cada tabla realiza una
extracción determinista fila a fila usando los **nombres de columna
literales** del código legacy (con mayúsculas y puntos), sin
normalización tipográfica ni heurísticas por ``dataclasses.fields``.

Pipeline por tabla:

  1. Localiza la ``ListObject`` por nombre (``worksheet.tables[name]``).
  2. Resuelve el rango con ``openpyxl.utils.cell.range_boundaries``
     (la API oficial para parsear ``table.ref``).
  3. Construye ``rows: list[dict]`` con clave = cabecera **EXACTA**
     de la celda (``"PLC.Tag"``, ``"Cfg.GrupoAlarma"``, …).
  4. Mapea cada fila a su dataclass con un constructor explícito
     (``_build_disp_ed``, ``_build_disp_ea``, …) que solicita
     cada columna con la clave literal del legacy.
  5. Emite trazas vía ``LogBuffer`` (SPA) + logger estándar.

Restricción arquitectónica: este módulo es OFFLINE; no importa
``siemens_tia_scripting``. Única dependencia externa: ``openpyxl``.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from areas.alimentacion.domain.models.dispositivos import (
    DimensionesDispositivos,
    DispEA,
    DispED,
    DispM,
    DispM_VF,
    DispSA,
    DispV,
    Dispositivo,
)
from core.application.log_buffer import get_log_buffer
from core.infrastructure.config_manager import ConfigManager
from core.infrastructure.parsers.excel_parser import ExcelParser


# Logger adicional a ``LogBuffer`` para entornos donde el buffer aún
# no existe (CLI, tests sin DI). ``LogBuffer`` sigue siendo la fuente
# de verdad que consume la SPA vía polling.
_module_logger = logging.getLogger("zc.parsers.alimentacion_excel")
_module_logger.setLevel(logging.INFO)
if not _module_logger.handlers:
    _handler = logging.StreamHandler()
    _handler.setFormatter(
        logging.Formatter("[%(name)s] %(levelname)s: %(message)s")
    )
    _module_logger.addHandler(_handler)
_module_logger.propagate = False


# ── Mapeo explícito de Tablas Nombradas (ListObjects) ───────────────────
# Determina la asociación entre:
#   * clave canónica del modelo (``DispED``, ``DispEA``, …)
#   * nombre de hoja (literal en el archivo .xlsx)
#   * nombre de la ``ListObject`` (Tabla_*)
# Cualquier desviación del operario (ej. ``ED`` en vez de ``DISP_ED``)
# se resuelve en ``_EXCEL_TARGETS`` actualizando la entrada.
#
# ESTOS VALORES SON LOS **DEFAULTS LEGACY**. Si el parser se
# construye con un ``ConfigManager``, el dict de instancia
# ``self._excel_targets`` se reescribe desde
# ``ConfigManager.get_excel_target_for(hw_type)`` para soportar
# overrides del config y futuros hw_types (sd/m_sina/tq/tq_ae).
_EXCEL_TARGETS: dict[str, dict[str, str]] = {
    "DispED":   {"sheet": "DISP_ED",   "table": "Tabla_Disp_ED"},
    "DispEA":   {"sheet": "DISP_EA",   "table": "Tabla_Disp_EA"},
    "DispSA":   {"sheet": "DISP_SA",   "table": "Tabla_Disp_SA"},
    "DispV":    {"sheet": "DISP_V",    "table": "Tabla_Disp_V"},
    "DispM":    {"sheet": "DISP_M",    "table": "Tabla_Disp_M"},
    "DispM_VF": {"sheet": "DISP_M_VF", "table": "Tabla_Disp_M_VF"},
}


# Mapa por defecto de named ranges N_MAX / num_disp_* → atributo
# legacy. Se usa como fallback cuando el parser se construye sin
# ``ConfigManager`` (modo histórico) o cuando
# ``_extract_dimensiones`` se llama sin ``named_range_map``.
_DEFAULT_NAMED_RANGE_MAP: dict[str, str] = {
    "N_MAX_DISP_ED":   "num_disp_ed",
    "N_MAX_DISP_EA":   "num_disp_ea",
    "N_MAX_DISP_SA":   "num_disp_sa",
    "N_MAX_DISP_V":    "num_disp_v",
    "N_MAX_DISP_M":    "num_disp_m",
    "N_MAX_DISP_M_VF": "num_disp_m_vf",
    "num_disp_ed":     "num_disp_ed",
    "num_disp_ea":     "num_disp_ea",
    "num_disp_sa":     "num_disp_sa",
    "num_disp_v":      "num_disp_v",
    "num_disp_m":      "num_disp_m",
    "num_disp_m_vf":   "num_disp_m_vf",
}


# ── Helpers de casteo seguro (defensivos contra NaN / None / vacíos) ───


def _safe_str(val: Any) -> str:
    """Convierte ``val`` a ``str`` quitando ``None`` / ``NaN`` / vacío.

    A diferencia de la versión legacy, NUNCA devuelve ``None``: la
    convención del parser actual es siempre ``str`` (con ``""`` como
    valor por defecto) para que las claves de los ``cfg_*`` lleguen
    literales a la SPA sin ``None`` que rompan el template.
    """
    if val is None:
        return ""
    text = str(val).strip()
    if text.lower() in ("nan", "none", ""):
        return ""
    return text


def _safe_int(val: Any, default: int = 0) -> int:
    """Convierte ``val`` a ``int`` con fallback a ``default``.

    Acepta ``None``, ``bool``, ``int``, ``float``, ``str`` numérico
    (``"5"``, ``"5.0"``). Cualquer string no numérico (ej.
    ``"Pendiente"``) devuelve ``default`` en lugar de lanzar.
    """
    if val is None:
        return default
    if isinstance(val, bool):
        return int(val)
    if isinstance(val, int):
        return val
    try:
        return int(str(val).strip())
    except (ValueError, TypeError):
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return default


def _safe_float(val: Any, default: float = 0.0) -> float:
    """Convierte ``val`` a ``float`` con fallback a ``default``.

    Acepta ``None``, ``bool``, ``int``, ``float`` y ``str`` en
    formato decimal con punto o coma (``"1,5"`` → ``1.5``).
    """
    if val is None:
        return default
    if isinstance(val, bool):
        return float(val)
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).strip().replace(",", "."))
    except (ValueError, TypeError):
        return default


def _extract_table_rows(
    worksheet: Worksheet,
    table_name: str,
) -> list[dict[str, Any]]:
    """Lee la ``ListObject`` indicada y devuelve ``[{col: val, ...}]``.

    Args:
        worksheet: Hoja de openpyxl ya abierta.
        table_name: Nombre de la ``ListObject`` (ej. ``Tabla_Disp_ED``).

    Returns:
        Lista de diccionarios donde **cada clave es la cabecera
        LITERAL** de la celda (mayúsculas, puntos, espacios). Si la
        tabla no existe, está vacía o su ``ref`` es inválido, devuelve
        ``[]``.
    """
    tables = getattr(worksheet, "tables", None) or {}
    table = tables.get(table_name)
    if table is None:
        return []
    ref = getattr(table, "ref", None)
    if not ref or not isinstance(ref, str) or ":" not in ref:
        return []

    try:
        min_col, min_row, max_col, max_row = range_boundaries(ref)
    except Exception:  # pragma: no cover - defensivo
        return []

    # ── Cabeceras literales (1ª fila del rango) ───────────────────────
    header_cells = next(
        worksheet.iter_rows(
            min_row=min_row,
            max_row=min_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )
    )
    headers: list[str] = [
        str(h).strip() if h is not None and str(h).strip() else ""
        for h in header_cells
    ]

    # ── Datos (resto del rango) ───────────────────────────────────────
    rows: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(
        min_row=min_row + 1,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    ):
        if row is None or all(c is None for c in row):
            continue
        item: dict[str, Any] = {}
        for header_name, value in zip(headers, row):
            if not header_name:
                continue
            item[header_name] = value
        if item:
            rows.append(item)
    return rows


# Mapa explícito canónica → builder del modelo.
# Evita colisiones de prefijos (``DispED`` vs ``DispEA``).
_CANONICAL_TO_BUILDER: dict[str, str] = {
    "DispED":   "_build_disp_ed",
    "DispEA":   "_build_disp_ea",
    "DispSA":   "_build_disp_sa",
    "DispV":    "_build_disp_v",
    "DispM":    "_build_dispm",
    "DispM_VF": "_build_disp_m_vf",
}


class AlimentacionExcelParser:
    """Parser Excel determinista del subdominio alimentación.

    Compone sobre ``ExcelParser`` para el parseo de named ranges
    (``extraer_dimensiones``) y abre su propio ``load_workbook`` para
    localizar las ``ListObjects`` en ``extraer_dtos``.

    Sólo procesa las tablas declaradas en ``_EXCEL_TARGETS`` (o en
    el override del ``ConfigManager`` si se inyecta uno). Las filas
    sin ``UID`` ni ``Numero`` se descartan silenciosamente (criterio
    de unicidad del PlcTag en TIA Portal).

    Args:
        config_manager: ``ConfigManager`` opcional. Si se pasa, el
            parser consulta ``get_excel_target_for(hw_type)`` y
            ``get_nmax_entry(name)`` para soportar overrides del
            config y futuros hw_types (sd/m_sina/tq/tq_ae) sin tocar
            código. Si es ``None`` (default), se instancia un
            ``ConfigManager`` por defecto apuntando al JSON del repo
            para garantizar la ruta data-driven; como fallback
            defensivo (p.ej. tests que parchean ``cwd``), se usa
            ``_EXCEL_TARGETS`` y ``_DEFAULT_NAMED_RANGE_MAP``.
    """

    def __init__(
        self,
        config_manager: ConfigManager | None = None,
    ) -> None:
        # Composición: usamos el parser genérico SOLO para dimensiones.
        self._generic_parser = ExcelParser()
        # Config: si no se inyecta, instanciamos el default apuntando
        # al JSON del repo. Esto garantiza que ``_build_named_range_map``
        # SIEMPRE produce las claves ``Num_Disp_*`` correctas (el
        # default estático NO las incluye: ver R1 del plan).
        if config_manager is None:
            try:
                self._config_manager = ConfigManager()
            except FileNotFoundError:
                # Fallback defensivo (p.ej. tests con cwd temporal):
                # el parser sigue funcionando con los defaults legacy,
                # pero las hojas con nombres ``Num_Disp_*`` no se
                # mapearán a ``num_disp_*`` (quedan en ``extras``).
                self._config_manager = None
        else:
            self._config_manager = config_manager
        # Buffer de logs opcional (DI tolerante).
        try:
            self._log = get_log_buffer()
        except Exception:  # pragma: no cover - defensivo
            self._log = None
        # Resoluciones data-driven.
        self._excel_targets: dict[str, dict[str, str]] = (
            self._build_excel_targets()
        )
        self._named_range_map: dict[str, str] = (
            self._build_named_range_map()
        )

    # ── Resolución de maps desde ConfigManager (o defaults) ───────────

    def _build_excel_targets(self) -> dict[str, dict[str, str]]:
        """Devuelve ``{canonica: {sheet, table}}`` desde config o defaults.

        Si hay ``ConfigManager``, itera ``list_hw_types_active()`` y
        llama ``get_excel_target_for(hw)`` para cada uno. Si no,
        retorna una copia de ``_EXCEL_TARGETS``.
        """
        if self._config_manager is None:
            return dict(_EXCEL_TARGETS)
        targets: dict[str, dict[str, str]] = {}
        for hw in self._config_manager.list_hw_types_active():
            t = self._config_manager.get_excel_target_for(hw)
            if t is None:
                continue
            canonica = t.get("canonical", "")
            if canonica:
                targets[canonica] = {
                    "sheet": t.get("sheet", ""),
                    "table": t.get("table", ""),
                }
        # Si el ConfigManager no devolvió nada (caso patológico),
        # caemos a los defaults.
        return targets or dict(_EXCEL_TARGETS)

    def _build_named_range_map(self) -> dict[str, str]:
        """Devuelve ``{nombre_nmax: nombre_attr_legacy}`` para extraer_dimensiones.

        Si hay ``ConfigManager``, itera ``list_nmax_active()`` y
        resuelve ``hw_type`` → ``num_disp_<hw>``. Si no, usa
        ``_DEFAULT_NAMED_RANGE_MAP``.
        """
        if self._config_manager is None:
            return dict(_DEFAULT_NAMED_RANGE_MAP)
        mapping: dict[str, str] = {}
        for nmax_name in self._config_manager.list_nmax_active():
            entry = self._config_manager.get_nmax_entry(nmax_name) or {}
            hw = entry.get("hw_type", "")
            if not hw:
                continue
            attr = f"num_disp_{hw}"
            mapping[nmax_name] = attr
            # Aceptamos también la forma legacy (``Num_Disp_X``).
            excel_nr = entry.get("excel_named_range", "")
            if excel_nr:
                mapping[excel_nr] = attr
            # Y la forma minúscula ``num_disp_x``.
            mapping[attr] = attr
        return mapping

    # ── DTOs por tabla ──────────────────────────────────────────────────
    def extraer_dtos(
        self, excel_path: str | Path
    ) -> dict[str, list[Dispositivo]]:
        """Lee cada ``ListObject`` declarada en ``_excel_targets``.

        Returns:
            ``dict[str, list[Dispositivo]]``. Las claves son los
            nombres canónicos (``DispED``, ``DispM_VF``…); los
            valores son listas de dataclasses inmutables.

        Raises:
            FileNotFoundError: Si el archivo no existe.
        """
        path = Path(excel_path)
        if not path.is_file():
            raise FileNotFoundError(
                f"No se encontró el archivo Excel: '{path}'"
            )

        # ``worksheet.tables`` SOLO está disponible en modo NO
        # ``read_only`` (openpyxl 3.1). Por eso abrimos el libro en
        # modo normal para localizar las ``ListObjects`` y seguimos
        # leyendo el contenido vía ``iter_rows``.
        workbook = load_workbook(
            filename=str(path), read_only=False, data_only=True,
        )
        try:
            result: dict[str, list[Dispositivo]] = {}
            for canonica, cfg in self._excel_targets.items():
                sheet_name: str = cfg["sheet"]
                table_name: str = cfg["table"]
                devices = self._extract_table(
                    workbook, canonica, sheet_name, table_name,
                )
                if devices:
                    result[canonica] = devices
            return result
        finally:
            workbook.close()

    # ── Extracción de UNA tabla ────────────────────────────────────────
    def _extract_table(
        self,
        workbook: Any,
        canonica: str,
        sheet_name: str,
        table_name: str,
    ) -> list[Dispositivo]:
        """Localiza la ``ListObject`` y construye los dispositivos."""
        if sheet_name not in workbook.sheetnames:
            self._emit(
                "warning",
                f"Tabla {table_name} no encontrada o vacía: "
                f"la hoja '{sheet_name}' no existe en el libro.",
            )
            return []

        worksheet = workbook[sheet_name]
        rows = _extract_table_rows(worksheet, table_name)
        if not rows:
            self._emit(
                "warning",
                f"Tabla {table_name} no encontrada o vacía en '{sheet_name}'.",
            )
            return []

        builder_attr = _CANONICAL_TO_BUILDER.get(canonica)
        if builder_attr is None:
            self._emit(
                "warning",
                f"No hay constructor explícito para '{canonica}'.",
            )
            return []
        builder = getattr(self, builder_attr, None)
        if builder is None:
            self._emit(
                "warning",
                f"No hay constructor explícito para '{canonica}'.",
            )
            return []

        devices: list[Dispositivo] = []
        discarded = 0
        for row in rows:
            try:
                device = builder(row)
            except Exception as exc:  # defensivo: nunca romper la tabla
                discarded += 1
                self._emit(
                    "warning",
                    f"Fila descartada en '{table_name}': {exc}",
                )
                continue
            if device is None:
                discarded += 1
                continue
            devices.append(device)

        if not devices:
            self._emit(
                "warning",
                f"Tabla {table_name} no encontrada o vacía: "
                f"ninguna fila construyó un dispositivo "
                f"({discarded} descartadas).",
            )
            return []

        self._emit(
            "info",
            f"Tabla {table_name} parseada: {len(devices)} elementos "
            f"({discarded} descartadas).",
        )
        return devices

    # ── Constructores explícitos (Mapeo Hardcoded Legacy) ─────────────
    # Cada constructor solicita la clave EXACTA del Excel heredada
    # del código legacy (``_legacy_reference/ZC_ALM_TOOLS/infrastructure
    # /parsers/hardware/disp_*.py``). No se itera sobre
    # ``dataclasses.fields`` ni se normalizan cabeceras: las claves
    # se piden literales (con puntos y mayúsculas).

    def _build_disp_ed(self, row: dict) -> DispED | None:
        """Replica 1:1 del ``DispEDParser.extraer()`` legacy."""
        if not row.get("UID") and not row.get("Numero"):
            return None
        return DispED(
            numero=_safe_int(row.get("Numero")),
            plc_tag=_safe_str(row.get("PLC.Tag")),
            plc_comentario=_safe_str(row.get("PLC.Comentario")),
            descripcion=_safe_str(row.get("Descripcion")),
            uid=_safe_str(row.get("UID")),
            tag=_safe_str(row.get("Tag")),
            fat=_safe_str(row.get("FAT")),
            e_byte=_safe_int(row.get("E.Byte")),
            e_bit=_safe_int(row.get("E.Bit")),
            gr_alarma=_safe_int(row.get("Gr.Alarma")),
            cuadro=_safe_str(row.get("Cuadro")),
            observaciones=_safe_str(row.get("Observaciones")),
            plc_tipo=_safe_str(row.get("PLC.Tipo")),
            plc_index=_safe_int(row.get("PLC.Index")),
            hmi_index=_safe_int(row.get("Hmi.Index")),
            hmi_texto=_safe_str(row.get("Hmi.Texto")),
            cfg_habilitar=_safe_str(row.get("Cfg.Habilitar")),
            cfg_byte_entrada=_safe_str(row.get("Cfg.ByteEntrada")),
            cfg_bit_entrada=_safe_str(row.get("Cfg.BitEntrada")),
            cfg_grupo_alarma=_safe_str(row.get("Cfg.GrupoAlarma")),
            comentario_db=_safe_str(row.get("ComentarioDB")),
        )

    def _build_disp_ea(self, row: dict) -> DispEA | None:
        """Replica 1:1 del ``DispEAParser.extraer()`` legacy.

        ``RII`` y ``RSI`` se castean con ``_safe_float`` (en legacy
        eran ``float(int(...))`` por culpa del ``astype(int)`` global
        del ``BaseParser``: aquí el casteo es directo y limpio).
        """
        if not row.get("UID") and not row.get("Numero"):
            return None
        return DispEA(
            numero=_safe_int(row.get("Numero")),
            plc_tag=_safe_str(row.get("PLC.Tag")),
            plc_comentario=_safe_str(row.get("PLC.Comentario")),
            descripcion=_safe_str(row.get("Descripcion")),
            uid=_safe_str(row.get("UID")),
            tag=_safe_str(row.get("Tag")),
            fat=_safe_str(row.get("FAT")),
            e_byte=_safe_int(row.get("E.Byte")),
            # El legacy usa la clave "UNIDADES" en MAYÚSCULAS. Aceptamos
            # también "Unidades" por compatibilidad hacia delante.
            unidades=_safe_str(
                row.get("UNIDADES")
                if row.get("UNIDADES") is not None
                else row.get("Unidades")
            ),
            rii=_safe_float(row.get("RII")),
            rsi=_safe_float(row.get("RSI")),
            gr_alarma=_safe_int(row.get("Gr.Alarma")),
            cuadro=_safe_str(row.get("Cuadro")),
            observaciones=_safe_str(row.get("Observaciones")),
            plc_tipo=_safe_str(row.get("PLC.Tipo")),
            plc_index=_safe_int(row.get("PLC.Index")),
            hmi_index=_safe_int(row.get("Hmi.Index")),
            hmi_texto=_safe_str(row.get("Hmi.Texto")),
            cfg_habilitar=_safe_str(row.get("Cfg.Habilitar")),
            cfg_byte_entrada=_safe_str(row.get("Cfg.ByteEntrada")),
            cfg_escaladomin=_safe_str(row.get("Cfg.EscaladoMin")),
            cfg_escaladomax=_safe_str(row.get("Cfg.EscaladoMax")),
            cfg_grupo_alarma=_safe_str(row.get("Cfg.GrupoAlarma")),
            comentario_db=_safe_str(row.get("ComentarioDB")),
        )

    def _build_disp_sa(self, row: dict) -> DispSA | None:
        """Replica 1:1 del ``DispSAParser.extraer()`` legacy."""
        if not row.get("UID") and not row.get("Numero"):
            return None
        return DispSA(
            numero=_safe_int(row.get("Numero")),
            plc_tag=_safe_str(row.get("PLC.Tag")),
            plc_comentario=_safe_str(row.get("PLC.Comentario")),
            descripcion=_safe_str(row.get("Descripcion")),
            uid=_safe_str(row.get("UID")),
            tag=_safe_str(row.get("Tag")),
            fat=_safe_str(row.get("FAT")),
            e_byte=_safe_int(row.get("E.Byte")),
            unidades=_safe_str(
                row.get("UNIDADES")
                if row.get("UNIDADES") is not None
                else row.get("Unidades")
            ),
            rii=_safe_float(row.get("RII")),
            rsi=_safe_float(row.get("RSI")),
            gr_alarma=_safe_int(row.get("Gr.Alarma")),
            cuadro=_safe_str(row.get("Cuadro")),
            observaciones=_safe_str(row.get("Observaciones")),
            plc_tipo=_safe_str(row.get("PLC.Tipo")),
            plc_index=_safe_int(row.get("PLC.Index")),
            hmi_index=_safe_int(row.get("Hmi.Index")),
            hmi_texto=_safe_str(row.get("Hmi.Texto")),
            cfg_habilitar=_safe_str(row.get("Cfg.Habilitar")),
            cfg_byte_entrada=_safe_str(row.get("Cfg.ByteEntrada")),
            cfg_escaladomin=_safe_str(row.get("Cfg.EscaladoMin")),
            cfg_escaladomax=_safe_str(row.get("Cfg.EscaladoMax")),
            cfg_grupo_alarma=_safe_str(row.get("Cfg.GrupoAlarma")),
            comentario_db=_safe_str(row.get("ComentarioDB")),
        )

    def _build_disp_v(self, row: dict) -> DispV | None:
        """Replica 1:1 del ``DispVParser.extraer()`` legacy."""
        if not row.get("UID") and not row.get("Numero"):
            return None
        return DispV(
            numero=_safe_int(row.get("Numero")),
            plc_tag=_safe_str(row.get("PLC.Tag")),
            plc_comentario=_safe_str(row.get("PLC.Comentario")),
            descripcion=_safe_str(row.get("Descripcion")),
            uid=_safe_str(row.get("UID")),
            tag=_safe_str(row.get("Tag")),
            fat=_safe_str(row.get("FAT")),
            s_byte=_safe_int(row.get("S.Byte")),
            s_bit=_safe_int(row.get("S.Bit")),
            rr_byte=_safe_int(row.get("RR.Byte")),
            rr_bit=_safe_int(row.get("RR.Bit")),
            rt_byte=_safe_int(row.get("RT.Byte")),
            rt_bit=_safe_int(row.get("RT.Bit")),
            gr_alarma=_safe_int(row.get("Gr.Alarma")),
            cuadro=_safe_str(row.get("Cuadro")),
            observaciones=_safe_str(row.get("Observaciones")),
            plc_tipo=_safe_str(row.get("PLC.Tipo")),
            plc_index=_safe_int(row.get("PLC.Index")),
            hmi_index=_safe_int(row.get("Hmi.Index")),
            hmi_texto=_safe_str(row.get("Hmi.Texto")),
            cfg_habilitar=_safe_str(row.get("Cfg.Habilitar")),
            cfg_byteretornoreposo=_safe_str(row.get("Cfg.ByteRetornoReposo")),
            cfg_bitretornoreposo=_safe_str(row.get("Cfg.BitRetornoReposo")),
            cfg_byteretornotrabajo=_safe_str(row.get("Cfg.ByteRetornoTrabajo")),
            cfg_bitretornotrabajo=_safe_str(row.get("Cfg.BitRetornoTrabajo")),
            cfg_byteactivacion=_safe_str(row.get("Cfg.ByteActivacion")),
            cfg_bitactivacion=_safe_str(row.get("Cfg.BitActivacion")),
            cfg_habitreposo=_safe_str(row.get("Cfg.HabRetReposo")),
            cfg_habitrtrabajo=_safe_str(row.get("Cfg.HabRetTrabajo")),
            cfg_grupoalarma=_safe_str(row.get("Cfg.GrupoAlarma")),
            comentario_db=_safe_str(row.get("ComentarioDB")),
        )

    def _build_dispm(self, row: dict) -> DispM | None:
        """Replica 1:1 del ``DispMParser.extraer()`` legacy."""
        if not row.get("UID") and not row.get("Numero"):
            return None
        return DispM(
            numero=_safe_int(row.get("Numero")),
            plc_tag=_safe_str(row.get("PLC.Tag")),
            plc_comentario=_safe_str(row.get("PLC.Comentario")),
            descripcion=_safe_str(row.get("Descripcion")),
            uid=_safe_str(row.get("UID")),
            tag=_safe_str(row.get("Tag")),
            fat=_safe_str(row.get("FAT")),
            s_byte=_safe_int(row.get("S.Byte")),
            s_bit=_safe_int(row.get("S.Bit")),
            rt_byte=_safe_int(row.get("RT.Byte")),
            rt_bit=_safe_int(row.get("RT.Bit")),
            rm_byte=_safe_int(row.get("RM.Byte")),
            rm_bit=_safe_int(row.get("RM.Bit")),
            gr_alarma=_safe_int(row.get("Gr.Alarma")),
            cuadro=_safe_str(row.get("Cuadro")),
            observaciones=_safe_str(row.get("Observaciones")),
            plc_tipo=_safe_str(row.get("PLC.Tipo")),
            plc_index=_safe_int(row.get("PLC.Index")),
            hmi_index=_safe_int(row.get("Hmi.Index")),
            hmi_texto=_safe_str(row.get("Hmi.Texto")),
            cfg_habilitar=_safe_str(row.get("Cfg.Habilitar")),
            cfg_byteretornotermico=_safe_str(row.get("Cfg.ByteRetornoTermico")),
            cfg_bitretornotermico=_safe_str(row.get("Cfg.BitRetornoTermico")),
            cfg_byteconfmarcha=_safe_str(row.get("Cfg.ByteConfMarcha")),
            cfg_bitconfmarcha=_safe_str(row.get("Cfg.BitConfMarcha")),
            cfg_byteactivacion=_safe_str(row.get("Cfg.ByteActivacion")),
            cfg_bitactivacion=_safe_str(row.get("Cfg.BitActivacion")),
            cfg_habrettermico=_safe_str(row.get("Cfg.HabRetTermico")),
            cfg_habretconfmarcha=_safe_str(row.get("Cfg.HabRetConfMarcha")),
            cfg_grupoalarma=_safe_str(row.get("Cfg.GrupoAlarma")),
            comentario_db=_safe_str(row.get("ComentarioDB")),
        )

    def _build_disp_m_vf(self, row: dict) -> DispM_VF | None:
        """Replica 1:1 del ``DispMVFarser.extraer()`` legacy."""
        if not row.get("UID") and not row.get("Numero"):
            return None
        return DispM_VF(
            numero=_safe_int(row.get("Numero")),
            plc_tag=_safe_str(row.get("PLC.Tag")),
            plc_comentario=_safe_str(row.get("PLC.Comentario")),
            descripcion=_safe_str(row.get("Descripcion")),
            uid=_safe_str(row.get("UID")),
            tag=_safe_str(row.get("Tag")),
            fat=_safe_str(row.get("FAT")),
            s_byte=_safe_int(row.get("S.Byte")),
            s_bit=_safe_int(row.get("S.Bit")),
            rt_byte=_safe_int(row.get("RT.Byte")),
            rt_bit=_safe_int(row.get("RT.Bit")),
            rm_byte=_safe_int(row.get("RM.Byte")),
            rm_bit=_safe_int(row.get("RM.Bit")),
            sa_byte=_safe_int(row.get("SA.Byte")),
            gr_alarma=_safe_int(row.get("Gr.Alarma")),
            cuadro=_safe_str(row.get("Cuadro")),
            observaciones=_safe_str(row.get("Observaciones")),
            plc_tipo=_safe_str(row.get("PLC.Tipo")),
            plc_index=_safe_int(row.get("PLC.Index")),
            hmi_index=_safe_int(row.get("Hmi.Index")),
            hmi_texto=_safe_str(row.get("Hmi.Texto")),
            cfg_habilitar=_safe_str(row.get("Cfg.Habilitar")),
            cfg_byteretornotermico=_safe_str(row.get("Cfg.ByteRetornoTermico")),
            cfg_bitretornotermico=_safe_str(row.get("Cfg.BitRetornoTermico")),
            cfg_byteconfmarcha=_safe_str(row.get("Cfg.ByteConfMarcha")),
            cfg_bitconfmarcha=_safe_str(row.get("Cfg.BitConfMarcha")),
            cfg_byteactivacion=_safe_str(row.get("Cfg.ByteActivacion")),
            cfg_bitactivacion=_safe_str(row.get("Cfg.BitActivacion")),
            cfg_byteanalogica=_safe_str(row.get("Cfg.ByteAnalogica")),
            cfg_habrettermico=_safe_str(row.get("Cfg.HabRetTermico")),
            cfg_habretconfmarcha=_safe_str(row.get("Cfg.HabRetConfMarcha")),
            cfg_grupoalarma=_safe_str(row.get("Cfg.GrupoAlarma")),
            comentario_db=_safe_str(row.get("ComentarioDB")),
        )

    # ── Dimensiones (named ranges num_disp_*) ─────────────────────────
    def extraer_dimensiones(
        self, excel_path: str | Path
    ) -> DimensionesDispositivos:
        """Lee named ranges ``num_disp_*`` y devuelve una instancia tipada.

        Returns:
            ``DimensionesDispositivos`` con los 6 contadores. Si un
            campo no existe en el Excel, queda en ``0``. Los N_MAX
            adicionales del ``n_max_catalog`` que el Excel defina
            como named range acaban en ``DimensionesDispositivos.extras``.
        """
        path = Path(excel_path)
        if not path.is_file():
            raise FileNotFoundError(
                f"No se encontró el archivo Excel: '{path}'"
            )

        workbook = load_workbook(
            filename=str(path), read_only=True, data_only=True
        )
        try:
            return self._extract_dimensiones(workbook, self._named_range_map)
        finally:
            workbook.close()

    @staticmethod
    def _extract_dimensiones(
        workbook: Any,
        named_range_map: dict[str, str] | None = None,
    ) -> DimensionesDispositivos:
        """Lee ``wb.defined_names`` y popula ``DimensionesDispositivos``.

        ``named_range_map`` se construye desde el ``ConfigManager`` o,
        si es ``None``, se usa ``_DEFAULT_NAMED_RANGE_MAP`` (los 6 legacy).
        """
        defined_names = getattr(workbook, "defined_names", None)
        if defined_names is None:
            return DimensionesDispositivos()

        items: Any = (
            defined_names.items()
            if hasattr(defined_names, "items")
            else []
        )

        # Mapa por defecto (6 legacy) si no se inyecta uno data-driven.
        if named_range_map is None:
            named_range_map = _DEFAULT_NAMED_RANGE_MAP

        result: dict[str, int] = {}
        extras: dict[str, int] = {}
        for name, definition in items:
            if not isinstance(name, str):
                continue
            attr = named_range_map.get(name)
            if attr is not None:
                value = _safe_int(_resolve_value(definition, workbook))
                result[attr] = value
            else:
                # Si el named range no es de los legacy, intentar leerlo
                # como N_MAX directo (data-driven): p.ej. un Excel que
                # defina ``N_MAX_DISP_FF`` → acaba en ``extras``.
                if name.startswith("N_MAX_DISP_") or name.startswith("Num_Disp_"):
                    v = _safe_int(_resolve_value(definition, workbook))
                    if v:
                        extras[name] = v

        if result:
            kwargs = dict(result)
            if extras:
                kwargs["extras"] = extras
            return DimensionesDispositivos(**kwargs)
        if extras:
            return DimensionesDispositivos(extras=extras)
        return DimensionesDispositivos()

    # ── Auditoría / Trazabilidad ────────────────────────────────────────
    def _emit(self, level: str, message: str) -> None:
        """Emite a ``LogBuffer`` y al logger estándar (fallback)."""
        log_buffer = getattr(self, "_log", None)
        if log_buffer is not None and hasattr(log_buffer, level):
            try:
                getattr(log_buffer, level)(message)
            except Exception:  # pragma: no cover - defensivo
                _module_logger.warning("LogBuffer.%s falló", level)
        if level == "warning":
            _module_logger.warning(message)
        else:
            _module_logger.info(message)


# ── Helpers de mapeo de named ranges (privados al módulo) ──────────────


def _resolve_value(definition: Any, workbook: Any) -> Any:
    """Lee el valor de un ``DefinedName`` resolviendo su hoja y celda."""
    # Compatibilidad: openpyxl 3.1+ expone ``attr_text``; las
    # versiones anteriores usaban ``value``. Aceptamos ambos.
    attr_text: Any = (
        getattr(definition, "attr_text", None)
        or getattr(definition, "value", None)
    )
    if not isinstance(attr_text, str) or "!" not in attr_text:
        return None

    sheet_part, cell_part = attr_text.split("!", 1)
    sheet_name = sheet_part.strip().strip("'").strip('"')
    cell_ref = cell_part.replace("$", "").strip()
    try:
        sheet = workbook[sheet_name]
    except (KeyError, TypeError):
        return None
    try:
        # ``destinations`` es la API moderna (openpyxl 3.1).
        # Si está disponible, devuelve (sheet, coord) directamente.
        destinations = getattr(definition, "destinations", None)
        if destinations is not None:
            dest = list(destinations)
            if dest:
                sheet_name, coord = dest[0]
                cell = workbook[sheet_name][coord]
                if isinstance(cell, tuple):
                    cell = cell[0][0] if isinstance(cell[0], tuple) else cell[0]
                return getattr(cell, "value", None)
        # Fallback: parsear ``attr_text``.
        cell = sheet[cell_ref]
    except (KeyError, AttributeError, TypeError):
        return None
    return getattr(cell, "value", None)


__all__ = [
    "AlimentacionExcelParser",
    "_EXCEL_TARGETS",
    "_CANONICAL_TO_BUILDER",
    "_safe_str",
    "_safe_int",
    "_safe_float",
    "_extract_table_rows",
]
