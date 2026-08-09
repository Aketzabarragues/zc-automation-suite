"""Smoke test determinista del parser.

Genera un libro .xlsx con:
  * Una hoja ``DISP_ED`` con una ``ListObject`` ``Tabla_Disp_ED``
    (flujo principal: búsqueda por tabla).
  * Una hoja ``DISP_V`` SIN ``ListObject`` para forzar el fallback
    por cabecera literal (``Numero`` + ``PLC.Tag``).
  * Una hoja ``DISP_EA`` con la ``ListObject`` vacía (cero filas
    tras la cabecera) para verificar el WARNING de tabla vacía.
  * Una hoja ``HOJADESC`` con nombre desconocido (debe ignorarse sin
    romper la carga).

Tras ``extraer_dtos`` se valida que:
  * ``DispED`` mapea correctamente columnas con puntos
    (``PLC.Tag``, ``Cfg.GrupoAlarma``…) y conserva ``cfg_*``
    como ``str`` (no se pierden ni se normalizan).
  * ``DispV`` se construye vía fallback (sin ListObject).
  * ``DISP_EA`` no aparece en el resultado (WARNING emitido).
  * ``HOJADESC`` no aparece en el resultado.
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from application.log_buffer import get_log_buffer
from infrastructure.alimentacion.parsers.alimentacion_excel_parser import (
    AlimentacionExcelParser,
)


ED_HEADERS = [
    "Numero", "PLC.Tag", "PLC.Comentario", "Descripcion", "UID", "Tag",
    "FAT", "E.Byte", "E.Bit", "Gr.Alarma", "Cuadro", "Observaciones",
    "PLC.Tipo", "PLC.Index", "Hmi.Index", "Hmi.Texto",
    "Cfg.Habilitar", "Cfg.ByteEntrada", "Cfg.BitEntrada",
    "Cfg.GrupoAlarma", "ComentarioDB",
]
ED_ROW = [
    1, "ED_TAG_1", "Comentario PLC", "Sensor 1", "UID-ED-1", 10,
    0, 0, 1, "1", "Cuadro A", "OK",
    "Bool", 0, 1, "Texto HMI 1",
    "cfg_habilitar_line := TRUE;",
    "DB_Entrada.byte := 0;",
    "DB_Entrada.bit := 1;",
    "Cfg.GrupoAlarma := 5;",
    "/* Comentario DB */",
]

V_HEADERS = [
    "Numero", "PLC.Tag", "PLC.Comentario", "Descripcion", "UID", "Tag",
    "FAT", "S.Byte", "S.Bit", "RR.Byte", "RR.Bit", "RT.Byte", "RT.Bit",
    "Gr.Alarma", "Cuadro", "Observaciones",
    "PLC.Tipo", "PLC.Index", "Hmi.Index", "Hmi.Texto",
    "Cfg.Habilitar", "Cfg.ByteRetornoReposo", "Cfg.BitRetornoReposo",
    "Cfg.ByteRetornoTrabajo", "Cfg.BitRetornoTrabajo",
    "Cfg.ByteActivacion", "Cfg.BitActivacion",
    "Cfg.HabRetReposo", "Cfg.HabRetTrabajo", "Cfg.GrupoAlarma",
    "ComentarioDB",
]
V_ROW = [
    1, "V_TAG_1", "", "Variable 1", "UID-V-1", 0,
    0, 0, 0, 0, 0, 0, 0,
    "", "Cuadro B", "",
    "Int", 0, 0, "",
    "cfg_habilitar := TRUE;",
    "b := 0;", "bit := 0;",
    "b := 0;", "bit := 0;",
    "b := 0;", "bit := 0;",
    "hab := TRUE;", "hab := TRUE;",
    "grupo := 0;",
    "",
]


def _col_letter(idx_1based: int) -> str:
    """``1 -> A``, ``27 -> AA``. Rango soportado por las pruebas."""
    n = idx_1based
    s = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        s = chr(65 + rem) + s
    return s


def _build_workbook() -> Path:
    wb = Workbook()

    # DISP_ED con ListObject
    ws_ed = wb.active
    ws_ed.title = "DISP_ED"
    ws_ed.append(ED_HEADERS)
    ws_ed.append(ED_ROW)
    last_col_ed = _col_letter(len(ED_HEADERS))
    table_ed = Table(
        displayName="Tabla_Disp_ED",
        ref=f"A1:{last_col_ed}2",
    )
    table_ed.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws_ed.add_table(table_ed)

    # DISP_V sin ListObject (forzar fallback por cabecera)
    ws_v = wb.create_sheet("DISP_V")
    ws_v.append(V_HEADERS)
    ws_v.append(V_ROW)

    # DISP_EA con ListObject VACÍA (cabecera sin filas)
    ws_ea = wb.create_sheet("DISP_EA")
    ws_ea.append(ED_HEADERS)  # sólo cabecera, ninguna fila
    last_col_ea = _col_letter(len(ED_HEADERS))
    table_ea = Table(
        displayName="Tabla_Disp_EA",
        ref=f"A1:{last_col_ea}1",
    )
    table_ea.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws_ea.add_table(table_ea)

    # Hoja desconocida: debe ignorarse.
    ws_x = wb.create_sheet("HOJADESC")
    ws_x.append(["dummy", "data"])

    tmp = Path(tempfile.gettempdir()) / "zc_test_table.xlsx"
    wb.save(tmp)
    wb.close()
    return tmp


def main() -> int:
    sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]
    log = get_log_buffer()
    log.clear()

    path = _build_workbook()
    try:
        parser = AlimentacionExcelParser()
        result = parser.extraer_dtos(path)
    finally:
        path.unlink(missing_ok=True)

    # 1) DispED por ListObject.
    assert "DispED" in result, f"DispED no detectado: {result}"
    ed_devices = result["DispED"]
    assert len(ed_devices) == 1, ed_devices
    ed = ed_devices[0]
    assert ed.plc_tag == "ED_TAG_1"
    assert ed.plc_comentario == "Comentario PLC"
    assert ed.descripcion == "Sensor 1"
    assert ed.uid == "UID-ED-1"
    assert ed.e_byte == 0 and ed.e_bit == 1
    # Crítico: los campos ``cfg_*`` deben preservarse literales.
    assert ed.cfg_habilitar == "cfg_habilitar_line := TRUE;"
    assert ed.cfg_byte_entrada == "DB_Entrada.byte := 0;"
    assert ed.cfg_bit_entrada == "DB_Entrada.bit := 1;"
    assert ed.cfg_grupo_alarma == "Cfg.GrupoAlarma := 5;"
    assert ed.comentario_db == "/* Comentario DB */"
    assert ed.gr_alarma == "1"
    assert ed.plc_tipo == "Bool"
    assert ed.plc_index == 0
    assert ed.hmi_index == 1
    assert ed.hmi_texto == "Texto HMI 1"

    # 2) DispV por fallback (sin ListObject).
    assert "DispV" in result, f"DispV no detectado: {result}"
    v_devices = result["DispV"]
    assert len(v_devices) == 1, v_devices
    v = v_devices[0]
    assert v.plc_tag == "V_TAG_1"
    assert v.uid == "UID-V-1"
    assert v.cfg_habilitar == "cfg_habilitar := TRUE;"
    assert v.cfg_byteretornoreposo == "b := 0;"
    assert v.cfg_habitrtrabajo == "hab := TRUE;"
    assert v.cfg_grupoalarma == "grupo := 0;"

    # 3) DispEA: tabla vacía → no debe aparecer, pero el WARNING sí.
    assert "DispEA" not in result, result
    assert any(
        "DISP_EA" in e["message"] and e["level"] == "warning"
        for e in log.snapshot()
    ), log.snapshot()

    # 4) Hoja desconocida ignorada.
    assert "HOJADESC" not in result
    assert "HOJADESC" not in str(result.keys())

    # 5) Trazas de éxito.
    joined = " | ".join(e["message"] for e in log.snapshot())
    assert "Tabla Tabla_Disp_ED parseada" in joined, joined
    assert "Tabla Tabla_Disp_V parseada" in joined, joined

    print("OK: parser determinista valida ListObject, fallback, tabla vacía y cfg_* literales.")
    print("Entradas de log:")
    for entry in log.snapshot():
        print(f"  [{entry['level']}] {entry['message']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
