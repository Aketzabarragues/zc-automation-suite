"""Parser de Excel para extraer el estado *deseado* de las dimensiones.

Recorre los nombres definidos (``defined_names``) del libro y extrae
el valor de los nombres que comienzan por ``N_MAX_`` o ``Num_Disp_``.

Restricción arquitectónica: este parser es OFFLINE; no importa
``siemens_tia_scripting``. La única librería externa usada es
``openpyxl``.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

_PREFIXES: tuple[str, ...] = ("N_MAX_", "Num_Disp_")


class ExcelParser:
    """Lee un .xlsx y extrae ``{nombre_constante: valor}`` para los
    nombres definidos que comienzan por ``N_MAX_`` o ``Num_Disp_``.

    Las celdas corruptas o no casteables a ``int`` se descartan
    silenciosamente (defensa frente a entradas manuales erróneas).
    """

    def extraer_dimensiones(self, excel_path: str | Path) -> dict[str, int]:
        """Devuelve ``{nombre_constante: valor}`` desde el Excel.

        Args:
            excel_path: Ruta al archivo .xlsx.

        Returns:
            ``dict[str, int]`` con los nombres definidos que pasaron
            los filtros (prefijo válido + casteo a ``int`` exitoso).

        Raises:
            FileNotFoundError: Si el archivo no existe.
        """
        path = Path(excel_path)
        if not path.is_file():
            raise FileNotFoundError(
                f"No se encontró el archivo Excel: '{path}'"
            )

        workbook = load_workbook(
            filename=str(path), read_only=True, data_only=True
        )
        try:
            return self._extract_defined_names(workbook)
        finally:
            workbook.close()

    @staticmethod
    def _extract_defined_names(workbook: Any) -> dict[str, int]:
        """Filtra ``wb.defined_names`` por prefijo y extrae valores enteros."""
        result: dict[str, int] = {}
        defined_names = getattr(workbook, "defined_names", None)
        if defined_names is None:
            return result

        # ``defined_names`` es un ``DefinedNameDict`` con ``.items()``.
        items: Any = (
            defined_names.items()
            if hasattr(defined_names, "items")
            else []
        )
        for name, definition in items:
            if not isinstance(name, str):
                continue
            if not any(name.startswith(prefix) for prefix in _PREFIXES):
                continue
            value = _resolve_value(definition, workbook)
            if value is None:
                continue
            try:
                result[name] = int(value)
            except (ValueError, TypeError):
                # Descartar celdas corruptas o tipos no enteros.
                continue
        return result

    def extraer_dtos(self, excel_path: str | Path) -> dict[str, list[dict[str, Any]]]:
        """Lee cada hoja del Excel y agrupa filas por nombre de hoja.

        Convención asumida:
          - Cada hoja representa un **tipo de dispositivo** (ej. ``"DispED"``,
            ``"DispV"``, ``"Motores"``).
          - La primera fila es la cabecera; las filas siguientes son
            instancias (un DTO por fila).
          - Se preservan todos los nombres de columna literalmente
            (espacios y mayúsculas intactos); el caller decide cómo
            mapear ``"nombre"`` / ``"direccion"`` o equivalentes.

        Args:
            excel_path: Ruta al archivo .xlsx.

        Returns:
            ``{nombre_hoja: [ {col1: val, col2: val, ...}, ... ]}``.
            Las hojas vacías o sin filas se omiten.

        Raises:
            FileNotFoundError: Si el archivo no existe.
        """
        path = Path(excel_path)
        if not path.is_file():
            raise FileNotFoundError(
                f"No se encontró el archivo Excel: '{path}'"
            )

        workbook = load_workbook(
            filename=str(path), read_only=True, data_only=True
        )
        try:
            return self._extract_dtos_from_workbook(workbook)
        finally:
            workbook.close()

    @staticmethod
    def _extract_dtos_from_workbook(workbook: Any) -> dict[str, list[dict[str, Any]]]:
        """Itera las hojas y devuelve filas como dicts keyed por header."""
        result: dict[str, list[dict[str, Any]]] = {}
        sheet_names: list[str] = list(getattr(workbook, "sheetnames", []) or [])
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            rows_iter = sheet.iter_rows(values_only=True)
            try:
                first_row = next(rows_iter)
            except StopIteration:
                continue
            header: list[str] = [
                str(h).strip() if h is not None and str(h).strip() else f"col_{i}"
                for i, h in enumerate(first_row)
            ]
            sheet_data: list[dict[str, Any]] = []
            for row in rows_iter:
                if row is None or all(c is None for c in row):
                    continue
                item: dict[str, Any] = {}
                for col_name, value in zip(header, row):
                    item[col_name] = value
                sheet_data.append(item)
            if sheet_data:
                result[sheet_name] = sheet_data
        return result


def _resolve_value(definition: Any, workbook: Any) -> Any:
    """Lee el valor de un ``DefinedName`` resolviendo su hoja y celda.

    Returns:
        El valor de la celda referenciada o ``None`` si no se pudo
        resolver la hoja/celda.
    """
    # openpyxl 3.1+: ``attr_text``. Versiones anteriores: ``value``.
    attr_text: Any = getattr(definition, "attr_text", None)
    if not attr_text:
        attr_text = getattr(definition, "value", None)
    if not isinstance(attr_text, str) or "!" not in attr_text:
        return None

    sheet_part, cell_part = attr_text.split("!", 1)
    # Quitar comillas que pueden rodear el nombre de la hoja.
    sheet_name = sheet_part.strip().strip("'").strip('"')
    cell_ref = cell_part.replace("$", "").strip()
    try:
        sheet = workbook[sheet_name]
    except (KeyError, TypeError):
        return None
    try:
        cell = sheet[cell_ref]
    except (KeyError, AttributeError, TypeError):
        return None
    return getattr(cell, "value", None)
